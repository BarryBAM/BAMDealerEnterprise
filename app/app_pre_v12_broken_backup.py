
import os
import sqlite3
import re
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
PROJECT_DIR = APP_DIR.parent
DB_DIR = PROJECT_DIR / "database"
UPLOAD_DIR = PROJECT_DIR / "uploads"
BACKUP_DIR = PROJECT_DIR / "backups"
REPORT_DIR = PROJECT_DIR / "reports"

for folder in (DB_DIR, UPLOAD_DIR, BACKUP_DIR, REPORT_DIR):
    folder.mkdir(parents=True, exist_ok=True)

DB_PATH = DB_DIR / "bam_motor_group.db"

app = Flask(__name__)
app.secret_key = os.environ.get("BAM_SECRET_KEY", "change-this-secret-key-before-production")
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "pdf"}

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def ensure_column(conn, table, column, definition):
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")



def init_db():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        display_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'owner'
    );

    CREATE TABLE IF NOT EXISTS vehicles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stock_no TEXT UNIQUE NOT NULL,
        status TEXT NOT NULL DEFAULT 'In Stock',
        purchase_date TEXT,
        make TEXT NOT NULL,
        model TEXT NOT NULL,
        variant TEXT,
        year INTEGER,
        vin TEXT UNIQUE,
        registration TEXT,
        odometer_km INTEGER,
        colour TEXT,
        purchase_price_inc_gst REAL NOT NULL DEFAULT 0,
        purchase_gst REAL NOT NULL DEFAULT 0,
        barry_contribution REAL NOT NULL DEFAULT 0,
        matt_contribution REAL NOT NULL DEFAULT 0,
        rego_expiry TEXT,
        photo_filename TEXT,
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        expense_date TEXT,
        category TEXT NOT NULL,
        description TEXT NOT NULL,
        supplier TEXT,
        paid_by TEXT NOT NULL,
        cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        receipt_filename TEXT,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER UNIQUE NOT NULL,
        sale_date TEXT,
        buyer_name TEXT,
        buyer_phone TEXT,
        sale_price_inc_gst REAL NOT NULL DEFAULT 0,
        sale_gst REAL NOT NULL DEFAULT 0,
        advertising_cost REAL NOT NULL DEFAULT 0,
        transfer_cost REAL NOT NULL DEFAULT 0,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contact_type TEXT,
        name TEXT NOT NULL,
        phone TEXT,
        email TEXT,
        address TEXT,
        licence_no TEXT,
        notes TEXT
    );

    CREATE TABLE IF NOT EXISTS consumables (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        purchase_date TEXT,
        item TEXT NOT NULL,
        category TEXT,
        supplier TEXT,
        purchased_by TEXT,
        qty_purchased REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        qty_used REAL NOT NULL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS vehicle_photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        caption TEXT,
        uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS job_cards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        job_date TEXT,
        category TEXT,
        description TEXT NOT NULL,
        supplier TEXT,
        paid_by TEXT,
        estimated_cost REAL NOT NULL DEFAULT 0,
        actual_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'Open',
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS job_card_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_card_id INTEGER NOT NULL,
        old_status TEXT,
        new_status TEXT NOT NULL,
        changed_by TEXT,
        changed_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        note TEXT,
        FOREIGN KEY(job_card_id) REFERENCES job_cards(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS vehicle_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        document_type TEXT NOT NULL,
        filename TEXT NOT NULL,
        description TEXT,
        uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS service_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        service_date TEXT,
        odometer_km INTEGER,
        service_type TEXT,
        description TEXT NOT NULL,
        supplier TEXT,
        cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        paid_by TEXT,
        next_service_date TEXT,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS parts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        part_number TEXT UNIQUE,
        part_name TEXT NOT NULL,
        category TEXT,
        supplier TEXT,
        quantity_on_hand REAL NOT NULL DEFAULT 0,
        reorder_level REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount_per_unit REAL NOT NULL DEFAULT 0,
        storage_location TEXT,
        notes TEXT
    );

    CREATE TABLE IF NOT EXISTS part_usage (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        part_id INTEGER NOT NULL,
        vehicle_id INTEGER NOT NULL,
        job_card_id INTEGER,
        usage_date TEXT,
        quantity_used REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        paid_by TEXT,
        notes TEXT,
        FOREIGN KEY(part_id) REFERENCES parts(id),
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE,
        FOREIGN KEY(job_card_id) REFERENCES job_cards(id) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT,
        action TEXT NOT NULL,
        entity_type TEXT,
        entity_id INTEGER,
        details TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS reminders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER,
        reminder_date TEXT NOT NULL,
        reminder_type TEXT NOT NULL,
        title TEXT NOT NULL,
        notes TEXT,
        completed INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER,
        task_date TEXT,
        due_date TEXT,
        title TEXT NOT NULL,
        category TEXT,
        assigned_to TEXT,
        priority TEXT NOT NULL DEFAULT 'Normal',
        status TEXT NOT NULL DEFAULT 'Open',
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );
    """)

    # Upgrade columns from older BAM databases.
    ensure_column(conn, "vehicles", "ppsr_number", "TEXT")
    ensure_column(conn, "vehicles", "roadworthy_status", "TEXT DEFAULT 'Not Checked'")
    ensure_column(conn, "vehicles", "service_due_date", "TEXT")
    ensure_column(conn, "vehicles", "service_history", "TEXT")
    ensure_column(conn, "vehicles", "asking_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "advertisement_title", "TEXT")
    ensure_column(conn, "vehicles", "advertisement_description", "TEXT")

    ensure_column(conn, "vehicles", "featured_photo_id", "INTEGER")
    ensure_column(conn, "vehicles", "estimated_sale_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "minimum_sale_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "valuation_notes", "TEXT")

    ensure_column(conn, "vehicles", "advertised_date", "TEXT")
    ensure_column(conn, "vehicles", "ready_for_sale_date", "TEXT")
    ensure_column(conn, "vehicles", "target_profit", "REAL DEFAULT 0")

    ensure_column(conn, "sales", "invoice_number", "TEXT")
    ensure_column(conn, "sales", "buyer_email", "TEXT")
    ensure_column(conn, "sales", "buyer_address", "TEXT")

    ensure_column(conn, "sales", "deposit_amount", "REAL DEFAULT 0")
    ensure_column(conn, "sales", "deposit_date", "TEXT")
    ensure_column(conn, "sales", "payment_method", "TEXT")
    ensure_column(conn, "sales", "trade_in_description", "TEXT")
    ensure_column(conn, "sales", "trade_in_value", "REAL DEFAULT 0")
    ensure_column(conn, "sales", "warranty_type", "TEXT")
    ensure_column(conn, "sales", "warranty_expiry", "TEXT")
    ensure_column(conn, "sales", "contract_number", "TEXT")

    conn.executescript("""
        CREATE TABLE IF NOT EXISTS parts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_number TEXT UNIQUE NOT NULL,
            vehicle_id INTEGER,
            vehicle_stock_no TEXT,
            vin TEXT,
            make TEXT,
            model TEXT,
            year INTEGER,
            category TEXT,
            subcategory TEXT,
            part_name TEXT NOT NULL,
            description TEXT,
            condition TEXT DEFAULT 'Used',
            quantity INTEGER DEFAULT 1,
            location TEXT,
            purchase_price REAL DEFAULT 0,
            selling_price REAL DEFAULT 0,
            supplier TEXT,
            status TEXT DEFAULT 'In Stock',
            engine_code TEXT,
            transmission_code TEXT,
            barcode TEXT,
            notes TEXT,
            date_added TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (vehicle_id) REFERENCES vehicles(id)
        );

        CREATE TABLE IF NOT EXISTS part_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            caption TEXT,
            is_featured INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (part_id)
                REFERENCES parts(id)
                ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS part_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL,
            quantity INTEGER DEFAULT 1,
            customer_name TEXT,
            customer_phone TEXT,
            customer_email TEXT,
            sale_price REAL DEFAULT 0,
            freight_cost REAL DEFAULT 0,
            payment_method TEXT,
            warranty TEXT,
            invoice_number TEXT,
            sale_date TEXT DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (part_id) REFERENCES parts(id)
        );
    """)
ensure_column(conn, "parts", "vehicle_id", "INTEGER")
ensure_column(conn, "parts", "vehicle_stock_no", "TEXT")
ensure_column(conn, "parts", "vin", "TEXT")
ensure_column(conn, "parts", "make", "TEXT")
ensure_column(conn, "parts", "model", "TEXT")
ensure_column(conn, "parts", "year", "INTEGER")
ensure_column(conn, "parts", "subcategory", "TEXT")
ensure_column(conn, "parts", "description", "TEXT")
ensure_column(conn, "parts", "condition", "TEXT DEFAULT 'Used'")
ensure_column(conn, "parts", "selling_price", "REAL DEFAULT 0")
ensure_column(conn, "parts", "status", "TEXT DEFAULT 'In Stock'")
ensure_column(conn, "parts", "engine_code", "TEXT")
ensure_column(conn, "parts", "transmission_code", "TEXT")
ensure_column(conn, "parts", "barcode", "TEXT")
ensure_column(conn, "parts", "date_added", "TEXT")
    count = conn.execute(
        "SELECT COUNT(*) AS c FROM users"
    ).fetchone()["c"]

    if count == 0:
        conn.execute(
            "INSERT INTO users(username,password_hash,display_name,role) VALUES(?,?,?,?)",
            ("barry", generate_password_hash("ChangeMe123!"), "Barry", "owner"),
        )
        conn.execute(
            "INSERT INTO users(username,password_hash,display_name,role) VALUES(?,?,?,?)",
            ("matt", generate_password_hash("ChangeMe123!"), "Matt", "owner"),
        )

    conn.commit()
    conn.close()



def next_stock_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT stock_no FROM vehicles WHERE stock_no LIKE 'BAM-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["stock_no"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"BAM-{highest + 1:05d}"


def next_invoice_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT invoice_number FROM sales WHERE invoice_number LIKE 'INV-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["invoice_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"INV-{highest + 1:05d}"



def log_action(action, entity_type=None, entity_id=None, details=None):
    conn = db()
    conn.execute("""
        INSERT INTO audit_log(user_name,action,entity_type,entity_id,details)
        VALUES(?,?,?,?,?)
    """, (
        session.get("display_name"),
        action,
        entity_type,
        entity_id,
        details,
    ))
    conn.commit()
    conn.close()


def next_contract_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT contract_number FROM sales WHERE contract_number LIKE 'CON-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["contract_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"CON-{highest + 1:05d}"


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def save_upload(file):
    if not file or not file.filename:
        return None
    if not allowed_file(file.filename):
        raise ValueError("Unsupported file type.")
    name = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(file.filename)}"
    file.save(UPLOAD_DIR / name)
    return name

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip().lower()
        password = request.form["password"]
        conn = db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["display_name"] = user["display_name"]
            return redirect(url_for("dashboard"))
        flash("Incorrect username or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    conn = db()

    metrics = conn.execute("""
        SELECT
          COUNT(*) AS total_vehicles,
          SUM(CASE WHEN status='Sold' THEN 1 ELSE 0 END) AS sold,
          SUM(CASE WHEN status='BER' THEN 1 ELSE 0 END) AS ber,
          SUM(CASE WHEN status NOT IN ('Sold','BER') THEN 1 ELSE 0 END) AS active_stock,
          SUM(CASE WHEN status NOT IN ('Sold','BER') THEN purchase_price_inc_gst ELSE 0 END) AS stock_value,
          SUM(purchase_price_inc_gst) AS purchase_total
        FROM vehicles
    """).fetchone()

    expense_total = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM expenses").fetchone()["v"]
    job_total = conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) AS v FROM job_cards").fetchone()["v"]
    service_total = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM service_entries").fetchone()["v"]
    parts_total = conn.execute("SELECT COALESCE(SUM(quantity_used*unit_cost_inc_gst),0) AS v FROM part_usage").fetchone()["v"]
    sales_total = conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) AS v FROM sales").fetchone()["v"]

    gst_paid = conn.execute("SELECT COALESCE(SUM(purchase_gst),0) AS v FROM vehicles").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM expenses").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM job_cards").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM service_entries").fetchone()["v"]
    gst_collected = conn.execute("SELECT COALESCE(SUM(sale_gst),0) AS v FROM sales").fetchone()["v"]

    barry_total = conn.execute("SELECT COALESCE(SUM(barry_contribution),0) AS v FROM vehicles").fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM expenses").fetchone()["v"]
    barry_total += conn.execute(
        "SELECT COALESCE(SUM(CASE "
        "WHEN paid_by='Barry' THEN CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END "
        "WHEN paid_by='Shared' THEN (CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)/2 "
        "ELSE 0 END),0) AS v FROM job_cards"
    ).fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM service_entries").fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN quantity_used*unit_cost_inc_gst WHEN paid_by='Shared' THEN quantity_used*unit_cost_inc_gst/2 ELSE 0 END),0) AS v FROM part_usage").fetchone()["v"]

    matt_total = conn.execute("SELECT COALESCE(SUM(matt_contribution),0) AS v FROM vehicles").fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM expenses").fetchone()["v"]
    matt_total += conn.execute(
        "SELECT COALESCE(SUM(CASE "
        "WHEN paid_by='Matt' THEN CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END "
        "WHEN paid_by='Shared' THEN (CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)/2 "
        "ELSE 0 END),0) AS v FROM job_cards"
    ).fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM service_entries").fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN quantity_used*unit_cost_inc_gst WHEN paid_by='Shared' THEN quantity_used*unit_cost_inc_gst/2 ELSE 0 END),0) AS v FROM part_usage").fetchone()["v"]

    status_counts = conn.execute("""
        SELECT status, COUNT(*) AS total
        FROM vehicles
        GROUP BY status
        ORDER BY status
    """).fetchall()

    today = date.today()
    rego_30 = (today + timedelta(days=30)).isoformat()
    rego_60 = (today + timedelta(days=60)).isoformat()
    rego_90 = (today + timedelta(days=90)).isoformat()

    rego_alerts = conn.execute("""
        SELECT id,stock_no,make,model,registration,rego_expiry
        FROM vehicles
        WHERE rego_expiry IS NOT NULL AND rego_expiry!=''
          AND rego_expiry<=?
          AND status NOT IN ('Sold','BER')
        ORDER BY rego_expiry
        LIMIT 8
    """, (rego_90,)).fetchall()

    service_alerts = conn.execute("""
        SELECT id,stock_no,make,model,service_due_date
        FROM vehicles
        WHERE service_due_date IS NOT NULL AND service_due_date!=''
          AND service_due_date<=?
          AND status NOT IN ('Sold','BER')
        ORDER BY service_due_date
        LIMIT 8
    """, (rego_90,)).fetchall()

    recent_vehicles = conn.execute("""
        SELECT v.*,
          COALESCE((SELECT SUM(cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
          COALESCE((SELECT SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)
                    FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
          COALESCE((SELECT sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v
        ORDER BY v.id DESC
        LIMIT 8
    """).fetchall()

    monthly_rows = conn.execute("""
        SELECT substr(sale_date,1,7) AS month,
               COUNT(*) AS vehicles_sold,
               COALESCE(SUM(sale_price_inc_gst),0) AS sales_total
        FROM sales
        WHERE sale_date IS NOT NULL AND sale_date!=''
        GROUP BY substr(sale_date,1,7)
        ORDER BY month DESC
        LIMIT 12
    """).fetchall()
    monthly_rows = list(reversed(monthly_rows))

    next_stock = next_stock_number(conn)

    today_text = today.isoformat()
    open_tasks = conn.execute("""
        SELECT t.*,v.stock_no,v.make,v.model
        FROM tasks t
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        WHERE t.status!='Completed'
        ORDER BY CASE t.priority WHEN 'Urgent' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
                 COALESCE(t.due_date,'9999-12-31'),t.id
        LIMIT 8
    """).fetchall()

    attention_vehicles = conn.execute("""
        SELECT v.*,
          CAST(julianday(?) - julianday(COALESCE(v.purchase_date,substr(v.created_at,1,10))) AS INTEGER) AS days_in_stock,
          COALESCE((SELECT COUNT(*) FROM reminders r WHERE r.vehicle_id=v.id AND r.completed=0),0) AS open_reminders,
          COALESCE((SELECT COUNT(*) FROM tasks t WHERE t.vehicle_id=v.id AND t.status!='Completed'),0) AS open_tasks
        FROM vehicles v
        WHERE v.status NOT IN ('Sold','BER')
        ORDER BY open_reminders DESC,open_tasks DESC,days_in_stock DESC
        LIMIT 8
    """, (today_text,)).fetchall()

    advertised_count = conn.execute(
        "SELECT COUNT(*) AS c FROM vehicles WHERE status='Advertised'"
    ).fetchone()["c"]

    sold_this_month = conn.execute("""
        SELECT COUNT(*) AS c FROM sales
        WHERE substr(sale_date,1,7)=substr(?,1,7)
    """, (today_text,)).fetchone()["c"]

    conn.close()

    net_profit = sales_total - (metrics["purchase_total"] or 0) - expense_total - job_total - service_total - parts_total

    status_labels = [row["status"] for row in status_counts]
    status_values = [row["total"] for row in status_counts]
    month_labels = [row["month"] for row in monthly_rows]
    month_values = [row["sales_total"] for row in monthly_rows]

    return render_template(
        "dashboard.html",
        metrics=metrics,
        expense_total=expense_total,
        job_total=job_total,
        service_total=service_total,
        parts_total=parts_total,
        sales_total=sales_total,
        gst_paid=gst_paid,
        gst_collected=gst_collected,
        net_profit=net_profit,
        barry_total=barry_total,
        matt_total=matt_total,
        rego_alerts=rego_alerts,
        service_alerts=service_alerts,
        recent_vehicles=recent_vehicles,
        next_stock=next_stock,
        status_labels=status_labels,
        status_values=status_values,
        month_labels=month_labels,
        month_values=month_values,
        today=today.isoformat(),
        open_tasks=open_tasks,
        attention_vehicles=attention_vehicles,
        advertised_count=advertised_count,
        sold_this_month=sold_this_month,
    )

@app.route("/vehicles")
@login_required
def vehicle_list():
    q = request.args.get("q", "").strip()
    conn = db()
    if q:
        rows = conn.execute("""
            SELECT * FROM vehicles
            WHERE stock_no LIKE ? OR make LIKE ? OR model LIKE ? OR vin LIKE ? OR registration LIKE ?
            ORDER BY id DESC
        """, tuple([f"%{q}%"] * 5)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM vehicles ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("vehicles.html", vehicles=rows, q=q)

@app.route("/vehicles/new", methods=["GET", "POST"])
@login_required
def vehicle_new():
    conn = db()
    suggested_stock = next_stock_number(conn)
    conn.close()
    if request.method == "POST":
        try:
            photo = save_upload(request.files.get("photo"))
            price = float(request.form.get("purchase_price_inc_gst") or 0)
            gst = round(price / 11, 2)
            conn = db()
            stock_no = request.form.get("stock_no", "").strip() or next_stock_number(conn)
            conn.execute("""
                INSERT INTO vehicles(
                    stock_no,status,purchase_date,make,model,variant,year,vin,registration,
                    odometer_km,colour,purchase_price_inc_gst,purchase_gst,
                    barry_contribution,matt_contribution,rego_expiry,photo_filename,notes,
                    ppsr_number,roadworthy_status,service_due_date,service_history
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                stock_no,
                request.form["status"],
                request.form.get("purchase_date"),
                request.form["make"].strip(),
                request.form["model"].strip(),
                request.form.get("variant"),
                request.form.get("year") or None,
                request.form.get("vin") or None,
                request.form.get("registration"),
                request.form.get("odometer_km") or None,
                request.form.get("colour"),
                price, gst,
                float(request.form.get("barry_contribution") or 0),
                float(request.form.get("matt_contribution") or 0),
                request.form.get("rego_expiry"),
                photo,
                request.form.get("notes"),
                request.form.get("ppsr_number"),
                request.form.get("roadworthy_status") or "Not Checked",
                request.form.get("service_due_date"),
                request.form.get("service_history"),
            ))
            conn.commit()
            conn.close()
            flash("Vehicle added.", "success")
            return redirect(url_for("vehicle_list"))
        except (sqlite3.IntegrityError, ValueError) as exc:
            flash(str(exc), "error")
    return render_template("vehicle_form.html", suggested_stock=suggested_stock)

@app.route("/vehicles/<int:vehicle_id>")
@login_required
def vehicle_detail(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    expenses = conn.execute(
        "SELECT * FROM expenses WHERE vehicle_id=? ORDER BY expense_date DESC,id DESC",
        (vehicle_id,)
    ).fetchall()
    job_cards = conn.execute(
        "SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY job_date DESC,id DESC",
        (vehicle_id,)
    ).fetchall()
    job_history_rows = conn.execute("""
        SELECT h.*, j.vehicle_id
        FROM job_card_history h
        JOIN job_cards j ON j.id=h.job_card_id
        WHERE j.vehicle_id=?
        ORDER BY h.changed_at DESC, h.id DESC
    """, (vehicle_id,)).fetchall()
    photos = conn.execute(
        "SELECT * FROM vehicle_photos WHERE vehicle_id=? ORDER BY CASE WHEN id=(SELECT featured_photo_id FROM vehicles WHERE id=?) THEN 0 ELSE 1 END,id DESC",
        (vehicle_id, vehicle_id)
    ).fetchall()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()

    if not vehicle:
        return "Vehicle not found", 404

    expense_total = sum(row["cost_inc_gst"] for row in expenses)
    job_total = sum(
        row["actual_cost_inc_gst"] if row["actual_cost_inc_gst"] > 0 else row["estimated_cost"]
        for row in job_cards
    )
    sale_price = sale["sale_price_inc_gst"] if sale else 0
    selling_costs = (sale["advertising_cost"] + sale["transfer_cost"]) if sale else 0
    profit = sale_price - vehicle["purchase_price_inc_gst"] - expense_total - job_total - selling_costs

    return render_template(
        "vehicle_detail.html",
        vehicle=vehicle,
        expenses=expenses,
        job_cards=job_cards,
        job_history_rows=job_history_rows,
        photos=photos,
        sale=sale,
        expense_total=expense_total,
        job_total=job_total,
        profit=profit,
    )

@app.route("/vehicles/<int:vehicle_id>/expense", methods=["POST"])
@login_required
def expense_add(vehicle_id):
    try:
        receipt = save_upload(request.files.get("receipt"))
        cost = float(request.form.get("cost_inc_gst") or 0)
        gst = round(cost / 11, 2)
        conn = db()
        conn.execute("""
            INSERT INTO expenses(vehicle_id,expense_date,category,description,supplier,paid_by,
                                 cost_inc_gst,gst_amount,receipt_filename,notes)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (vehicle_id, request.form.get("expense_date"), request.form["category"],
              request.form["description"], request.form.get("supplier"),
              request.form["paid_by"], cost, gst, receipt, request.form.get("notes")))
        conn.commit()
        conn.close()
        flash("Expense added.", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

@app.route("/vehicles/<int:vehicle_id>/photos", methods=["POST"])
@login_required
def photo_add(vehicle_id):
    files = request.files.getlist("photos")
    caption = request.form.get("caption")
    saved = 0
    conn = db()
    try:
        for file in files:
            filename = save_upload(file)
            if filename:
                conn.execute(
                    "INSERT INTO vehicle_photos(vehicle_id,filename,caption) VALUES(?,?,?)",
                    (vehicle_id, filename, caption),
                )
                saved += 1
        conn.commit()
        flash(f"{saved} photo(s) added.", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    finally:
        conn.close()
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/job-card", methods=["POST"])
@login_required
def job_card_add(vehicle_id):
    actual = float(request.form.get("actual_cost_inc_gst") or 0)
    gst = round(actual / 11, 2)
    conn = db()
    cursor = conn.execute("""
        INSERT INTO job_cards(
            vehicle_id,job_date,category,description,supplier,paid_by,
            estimated_cost,actual_cost_inc_gst,gst_amount,status,notes
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
    """, (
        vehicle_id,
        request.form.get("job_date"),
        request.form.get("category"),
        request.form.get("description"),
        request.form.get("supplier"),
        request.form.get("paid_by"),
        float(request.form.get("estimated_cost") or 0),
        actual,
        gst,
        request.form.get("status") or "Open",
        request.form.get("notes"),
    ))
    job_id = cursor.lastrowid
    conn.execute("""
        INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
        VALUES(?,?,?,?,?)
    """, (
        job_id, None, request.form.get("status") or "Open",
        session.get("display_name"), "Job card created"
    ))
    conn.commit()
    conn.close()
    flash("Job card added.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))



@app.route("/job-cards/<int:job_id>/edit", methods=["GET", "POST"])
@login_required
def job_card_edit(job_id):
    conn = db()
    job = conn.execute("SELECT * FROM job_cards WHERE id=?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return "Job card not found", 404

    if request.method == "POST":
        old_status = job["status"]
        new_status = request.form.get("status") or old_status
        actual = float(request.form.get("actual_cost_inc_gst") or 0)
        gst = round(actual / 11, 2)

        conn.execute("""
            UPDATE job_cards
            SET job_date=?, category=?, description=?, supplier=?, paid_by=?,
                estimated_cost=?, actual_cost_inc_gst=?, gst_amount=?, status=?, notes=?
            WHERE id=?
        """, (
            request.form.get("job_date"),
            request.form.get("category"),
            request.form.get("description"),
            request.form.get("supplier"),
            request.form.get("paid_by"),
            float(request.form.get("estimated_cost") or 0),
            actual,
            gst,
            new_status,
            request.form.get("notes"),
            job_id,
        ))

        change_note = request.form.get("change_note") or "Job card updated"
        conn.execute("""
            INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
            VALUES(?,?,?,?,?)
        """, (
            job_id, old_status, new_status,
            session.get("display_name"), change_note
        ))
        conn.commit()
        vehicle_id = job["vehicle_id"]
        conn.close()
        flash("Job card updated.", "success")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    history = conn.execute(
        "SELECT * FROM job_card_history WHERE job_card_id=? ORDER BY changed_at DESC,id DESC",
        (job_id,)
    ).fetchall()
    conn.close()
    return render_template("job_card_edit.html", job=job, history=history)


@app.route("/job-cards/<int:job_id>/reopen", methods=["POST"])
@login_required
def job_card_reopen(job_id):
    conn = db()
    job = conn.execute("SELECT * FROM job_cards WHERE id=?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return "Job card not found", 404

    old_status = job["status"]
    new_status = request.form.get("status") or "Open"
    note = request.form.get("note") or "Job card reopened"

    conn.execute("UPDATE job_cards SET status=? WHERE id=?", (new_status, job_id))
    conn.execute("""
        INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
        VALUES(?,?,?,?,?)
    """, (
        job_id, old_status, new_status,
        session.get("display_name"), note
    ))
    conn.commit()
    vehicle_id = job["vehicle_id"]
    conn.close()
    flash(f"Job card reopened as {new_status}.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/sale", methods=["POST"])
@login_required
def sale_add(vehicle_id):
    price = float(request.form.get("sale_price_inc_gst") or 0)
    gst = round(price / 11, 2)
    conn = db()
    existing = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    invoice_number = existing["invoice_number"] if existing and existing["invoice_number"] else next_invoice_number(conn)
    contract_number = existing["contract_number"] if existing and existing["contract_number"] else next_contract_number(conn)

    conn.execute("""
        INSERT INTO sales(
            vehicle_id,sale_date,buyer_name,buyer_phone,buyer_email,buyer_address,
            sale_price_inc_gst,sale_gst,advertising_cost,transfer_cost,invoice_number,
            deposit_amount,deposit_date,payment_method,trade_in_description,trade_in_value,
            warranty_type,warranty_expiry,contract_number,notes
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(vehicle_id) DO UPDATE SET
          sale_date=excluded.sale_date,
          buyer_name=excluded.buyer_name,
          buyer_phone=excluded.buyer_phone,
          buyer_email=excluded.buyer_email,
          buyer_address=excluded.buyer_address,
          sale_price_inc_gst=excluded.sale_price_inc_gst,
          sale_gst=excluded.sale_gst,
          advertising_cost=excluded.advertising_cost,
          transfer_cost=excluded.transfer_cost,
          invoice_number=COALESCE(sales.invoice_number,excluded.invoice_number),
          deposit_amount=excluded.deposit_amount,
          deposit_date=excluded.deposit_date,
          payment_method=excluded.payment_method,
          trade_in_description=excluded.trade_in_description,
          trade_in_value=excluded.trade_in_value,
          warranty_type=excluded.warranty_type,
          warranty_expiry=excluded.warranty_expiry,
          contract_number=COALESCE(sales.contract_number,excluded.contract_number),
          notes=excluded.notes
    """, (
        vehicle_id,
        request.form.get("sale_date"),
        request.form.get("buyer_name"),
        request.form.get("buyer_phone"),
        request.form.get("buyer_email"),
        request.form.get("buyer_address"),
        price,
        gst,
        float(request.form.get("advertising_cost") or 0),
        float(request.form.get("transfer_cost") or 0),
        invoice_number,
        float(request.form.get("deposit_amount") or 0),
        request.form.get("deposit_date"),
        request.form.get("payment_method"),
        request.form.get("trade_in_description"),
        float(request.form.get("trade_in_value") or 0),
        request.form.get("warranty_type"),
        request.form.get("warranty_expiry"),
        contract_number,
        request.form.get("notes"),
    ))
    conn.execute("UPDATE vehicles SET status='Sold' WHERE id=?", (vehicle_id,))
    conn.commit()
    conn.close()
    log_action("Sale recorded", "vehicle", vehicle_id, f"{invoice_number} / {contract_number}")
    flash(f"Sale recorded. Invoice {invoice_number} and contract {contract_number} created.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/import-excel", methods=["GET", "POST"])
@login_required
def import_excel():
    result = None
    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            flash("Please choose an .xlsx Excel workbook.", "error")
            return redirect(url_for("import_excel"))

        temp_path = BASE_DIR / f"import_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(uploaded.filename)}"
        uploaded.save(temp_path)
        imported = 0
        skipped = 0
        warnings = []
        workbook = None

        try:
            workbook = load_workbook(temp_path, data_only=True)
            sheet = None
            for candidate in ["Vehicles", "Vehicle Inventory"]:
                if candidate in workbook.sheetnames:
                    sheet = workbook[candidate]
                    break
            if sheet is None:
                raise ValueError("The workbook needs a sheet named Vehicles or Vehicle Inventory.")

            headers = {}
            header_row = None
            for row_number in range(1, min(sheet.max_row, 15) + 1):
                values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
                if any(str(v).strip() in ("Stock No.", "Stock Number") for v in values if v is not None):
                    header_row = row_number
                    headers = {
                        str(sheet.cell(row_number, col).value).strip(): col
                        for col in range(1, sheet.max_column + 1)
                        if sheet.cell(row_number, col).value is not None
                    }
                    break
            if not header_row:
                raise ValueError("Could not find the vehicle heading row.")

            def value(row, *names):
                for name in names:
                    col = headers.get(name)
                    if col:
                        return sheet.cell(row, col).value
                return None

            conn = db()
            for row in range(header_row + 1, sheet.max_row + 1):
                make = str(value(row, "Make") or "").strip()
                model = str(value(row, "Model") or "").strip()
                if not make and not model:
                    continue

                stock = str(value(row, "Stock No.", "Stock Number") or "").strip()
                if not stock:
                    stock = next_stock_number(conn)

                existing = conn.execute(
                    "SELECT id FROM vehicles WHERE stock_no=?",
                    (stock,)
                ).fetchone()
                if existing:
                    skipped += 1
                    continue

                vin = str(value(row, "VIN") or "").strip() or None
                if vin and conn.execute("SELECT id FROM vehicles WHERE vin=?", (vin,)).fetchone():
                    warnings.append(f"{stock}: duplicate VIN skipped")
                    skipped += 1
                    continue

                price = clean_number(value(row, "Purchase Price Inc GST", "Purchase Price"))
                gst = clean_number(value(row, "Purchase GST"))
                if not gst and price:
                    gst = round(price / 11, 2)

                purchase_date = value(row, "Purchase Date")
                rego_expiry = value(row, "Registration Expiry", "Rego Expiry")
                for_date = lambda x: x.date().isoformat() if hasattr(x, "date") else (str(x) if x else None)

                conn.execute("""
                    INSERT INTO vehicles(
                        stock_no,status,purchase_date,make,model,variant,year,vin,registration,
                        odometer_km,colour,purchase_price_inc_gst,purchase_gst,
                        barry_contribution,matt_contribution,rego_expiry,photo_filename,notes
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    stock,
                    value(row, "Status") or "In Stock",
                    for_date(purchase_date),
                    make,
                    model,
                    value(row, "Variant"),
                    int(clean_number(value(row, "Year"))) or None,
                    vin,
                    value(row, "Registration"),
                    int(clean_number(value(row, "Odometer km", "Odometer"))) or None,
                    value(row, "Colour"),
                    price,
                    gst,
                    clean_number(value(row, "Barry Contribution", "Barry Purchase Contribution")),
                    clean_number(value(row, "Matt Contribution", "Matt Purchase Contribution")),
                    for_date(rego_expiry),
                    value(row, "Main Photo Link"),
                    value(row, "Notes"),
                ))
                imported += 1

            conn.commit()
            conn.close()
            result = {"imported": imported, "skipped": skipped, "warnings": warnings}
            flash(f"Import completed: {imported} vehicle(s) added, {skipped} skipped.", "success")
        except Exception as exc:
            flash(f"Import failed: {exc}", "error")
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except PermissionError:
                    # Windows may briefly retain the file handle. The next import
                    # uses a new temporary name, so leaving this temp copy is safe.
                    pass

    return render_template("import_excel.html", result=result)


@app.route("/vehicles/<int:vehicle_id>/report")
@login_required
def vehicle_report(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    expenses = conn.execute("SELECT * FROM expenses WHERE vehicle_id=? ORDER BY expense_date,id", (vehicle_id,)).fetchall()
    jobs = conn.execute("SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY job_date,id", (vehicle_id,)).fetchall()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404
    expense_total = sum(x["cost_inc_gst"] for x in expenses)
    job_total = sum((x["actual_cost_inc_gst"] if x["actual_cost_inc_gst"] > 0 else x["estimated_cost"]) for x in jobs)
    sale_price = sale["sale_price_inc_gst"] if sale else 0
    selling_costs = (sale["advertising_cost"] + sale["transfer_cost"]) if sale else 0
    profit = sale_price - vehicle["purchase_price_inc_gst"] - expense_total - job_total - selling_costs
    return render_template("vehicle_report.html", vehicle=vehicle, expenses=expenses, jobs=jobs,
                           sale=sale, expense_total=expense_total, job_total=job_total, profit=profit)

@app.route("/backup")
@login_required
def backup_database():
    backup_dir = BACKUP_DIR
    backup_dir.mkdir(exist_ok=True)
    backup_name = f"bam_motor_group_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    backup_path = backup_dir / backup_name
    source = sqlite3.connect(DB_PATH)
    target = sqlite3.connect(backup_path)
    with target:
        source.backup(target)
    source.close()
    target.close()
    return send_file(backup_path, as_attachment=True, download_name=backup_name)


@app.route("/contacts", methods=["GET", "POST"])
@login_required
def contacts_page():
    conn = db()
    if request.method == "POST":
        conn.execute("""
            INSERT INTO contacts(contact_type,name,phone,email,address,licence_no,notes)
            VALUES(?,?,?,?,?,?,?)
        """, (
            request.form.get("contact_type"),
            request.form.get("name"),
            request.form.get("phone"),
            request.form.get("email"),
            request.form.get("address"),
            request.form.get("licence_no"),
            request.form.get("notes"),
        ))
        conn.commit()
        flash("Contact saved.", "success")
        return redirect(url_for("contacts_page"))

    q = request.args.get("q", "").strip()
    if q:
        rows = conn.execute("""
            SELECT * FROM contacts
            WHERE name LIKE ? OR phone LIKE ? OR email LIKE ? OR contact_type LIKE ?
            ORDER BY name
        """, tuple([f"%{q}%"] * 4)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM contacts ORDER BY name").fetchall()
    conn.close()
    return render_template("contacts.html", contacts=rows, q=q)


@app.route("/reports/profit")
@login_required
def profit_report():
    conn = db()
    rows = conn.execute("""
        SELECT v.id,v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price,
               COALESCE((SELECT s.advertising_cost+s.transfer_cost FROM sales s WHERE s.vehicle_id=v.id),0) AS selling_costs
        FROM vehicles v
        ORDER BY v.id DESC
    """).fetchall()
    conn.close()
    report_rows = []
    for row in rows:
        total_cost = row["purchase_price_inc_gst"] + row["expenses"] + row["jobs"] + row["selling_costs"]
        profit = row["sale_price"] - total_cost
        report_rows.append(dict(row) | {"total_cost": total_cost, "profit": profit})
    totals = {
        "purchase": sum(r["purchase_price_inc_gst"] for r in report_rows),
        "expenses": sum(r["expenses"] for r in report_rows),
        "jobs": sum(r["jobs"] for r in report_rows),
        "sales": sum(r["sale_price"] for r in report_rows),
        "profit": sum(r["profit"] for r in report_rows),
    }
    return render_template("profit_report.html", rows=report_rows, totals=totals)


@app.route("/vehicles/<int:vehicle_id>/invoice")
@login_required
def sale_invoice(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Sale invoice is not available until a sale is recorded.", 404
    sale_ex_gst = sale["sale_price_inc_gst"] - sale["sale_gst"]
    return render_template("sale_invoice.html", vehicle=vehicle, sale=sale, sale_ex_gst=sale_ex_gst)


@app.route("/export/vehicles.csv")
@login_required
def export_vehicles_csv():
    conn = db()
    rows = conn.execute("""
        SELECT stock_no,status,purchase_date,year,make,model,variant,vin,registration,
               odometer_km,purchase_price_inc_gst,purchase_gst,barry_contribution,
               matt_contribution,rego_expiry,ppsr_number,roadworthy_status,service_due_date
        FROM vehicles ORDER BY id
    """).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(rows[0].keys() if rows else [
        "stock_no","status","purchase_date","year","make","model","variant","vin",
        "registration","odometer_km","purchase_price_inc_gst","purchase_gst",
        "barry_contribution","matt_contribution","rego_expiry","ppsr_number",
        "roadworthy_status","service_due_date"
    ])
    for row in rows:
        writer.writerow(tuple(row))
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=BAM_vehicle_export.csv"},
    )


@app.route("/export/profit.csv")
@login_required
def export_profit_csv():
    conn = db()
    rows = conn.execute("""
        SELECT v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v ORDER BY v.id
    """).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Stock","Year","Make","Model","Status","Purchase","Expenses","Jobs","Sale","Profit/Loss"])
    for row in rows:
        profit = row["sale_price"] - row["purchase_price_inc_gst"] - row["expenses"] - row["jobs"]
        writer.writerow([
            row["stock_no"],row["year"],row["make"],row["model"],row["status"],
            row["purchase_price_inc_gst"],row["expenses"],row["jobs"],row["sale_price"],profit
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=BAM_profit_export.csv"},
    )


@app.route("/parts", methods=["GET", "POST"])
@login_required
def parts_page():
    conn = db()
    if request.method == "POST":
        quantity = float(request.form.get("quantity_on_hand") or 0)
        unit_cost = float(request.form.get("unit_cost_inc_gst") or 0)
        gst = round(unit_cost / 11, 2)
        conn.execute("""
            INSERT INTO parts(
                part_number,part_name,category,supplier,quantity_on_hand,reorder_level,
                unit_cost_inc_gst,gst_amount_per_unit,storage_location,notes
            ) VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (
            request.form.get("part_number") or None,
            request.form.get("part_name"),
            request.form.get("category"),
            request.form.get("supplier"),
            quantity,
            float(request.form.get("reorder_level") or 0),
            unit_cost,
            gst,
            request.form.get("storage_location"),
            request.form.get("notes"),
        ))
        conn.commit()
        flash("Part added.", "success")
        return redirect(url_for("parts_page"))

    q = request.args.get("q", "").strip()
    if q:
        rows = conn.execute("""
            SELECT * FROM parts
            WHERE part_number LIKE ? OR part_name LIKE ? OR category LIKE ? OR supplier LIKE ?
            ORDER BY part_name
        """, tuple([f"%{q}%"] * 4)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM parts ORDER BY part_name").fetchall()
    conn.close()
    return render_template("parts.html", parts=rows, q=q)


@app.route("/vehicles/<int:vehicle_id>/use-part", methods=["POST"])
@login_required
def use_part(vehicle_id):
    part_id = int(request.form.get("part_id"))
    quantity = float(request.form.get("quantity_used") or 0)
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()

    if not part:
        conn.close()
        flash("Part not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    if quantity <= 0:
        conn.close()
        flash("Quantity must be greater than zero.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    if quantity > part["quantity_on_hand"]:
        conn.close()
        flash("Not enough stock available.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute("""
        INSERT INTO part_usage(
            part_id,vehicle_id,job_card_id,usage_date,quantity_used,unit_cost_inc_gst,paid_by,notes
        ) VALUES(?,?,?,?,?,?,?,?)
    """, (
        part_id,
        vehicle_id,
        int(request.form.get("job_card_id")) if request.form.get("job_card_id") else None,
        request.form.get("usage_date"),
        quantity,
        part["unit_cost_inc_gst"],
        request.form.get("paid_by"),
        request.form.get("notes"),
    ))
    conn.execute(
        "UPDATE parts SET quantity_on_hand=quantity_on_hand-? WHERE id=?",
        (quantity, part_id),
    )
    conn.commit()
    conn.close()
    flash("Part allocated to vehicle.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/search")
@login_required
def global_search():
    q = request.args.get("q", "").strip()
    conn = db()
    vehicles = []
    contacts = []
    if q:
        pattern = f"%{q}%"
        vehicles = conn.execute("""
            SELECT * FROM vehicles
            WHERE stock_no LIKE ? OR vin LIKE ? OR registration LIKE ? OR make LIKE ? OR model LIKE ?
            ORDER BY id DESC
        """, (pattern, pattern, pattern, pattern, pattern)).fetchall()
        contacts = conn.execute("""
            SELECT * FROM contacts
            WHERE name LIKE ? OR phone LIKE ? OR email LIKE ?
            ORDER BY name
        """, (pattern, pattern, pattern)).fetchall()
    conn.close()
    return render_template("global_search.html", q=q, vehicles=vehicles, contacts=contacts)


@app.route("/reports/finance")
@login_required
def finance_summary():
    conn = db()
    rows = conn.execute("""
        SELECT v.id,v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,v.barry_contribution,v.matt_contribution,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT SUM(s.cost_inc_gst) FROM service_entries s WHERE s.vehicle_id=v.id),0) AS services,
               COALESCE((SELECT SUM(u.quantity_used*u.unit_cost_inc_gst) FROM part_usage u WHERE u.vehicle_id=v.id),0) AS parts,
               COALESCE((SELECT sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v
        ORDER BY v.id DESC
    """).fetchall()
    conn.close()

    report = []
    total_barry = 0
    total_matt = 0
    total_profit = 0
    money_in_stock = 0

    for r in rows:
        total_cost = r["purchase_price_inc_gst"] + r["expenses"] + r["jobs"] + r["services"] + r["parts"]
        profit = r["sale_price"] - total_cost
        if r["status"] not in ("Sold", "BER"):
            money_in_stock += total_cost
        total_barry += r["barry_contribution"]
        total_matt += r["matt_contribution"]
        total_profit += profit
        report.append(dict(r) | {"total_cost": total_cost, "profit": profit})

    return render_template(
        "finance_summary.html",
        rows=report,
        total_barry=total_barry,
        total_matt=total_matt,
        total_profit=total_profit,
        money_in_stock=money_in_stock,
        barry_profit_share=total_profit/2,
        matt_profit_share=total_profit/2,
    )


@app.route("/reports/performance")
@login_required
def performance_report():
    conn = db()
    monthly = conn.execute("""
        SELECT substr(sale_date,1,7) AS month,
               COUNT(*) AS vehicles_sold,
               SUM(sale_price_inc_gst) AS sales_total
        FROM sales
        WHERE sale_date IS NOT NULL AND sale_date!=''
        GROUP BY substr(sale_date,1,7)
        ORDER BY month
    """).fetchall()

    makes = conn.execute("""
        SELECT make, COUNT(*) AS total
        FROM vehicles
        GROUP BY make
        ORDER BY total DESC, make
        LIMIT 10
    """).fetchall()
    conn.close()

    max_sales = max([r["sales_total"] or 0 for r in monthly], default=1)
    return render_template("performance_report.html", monthly=monthly, makes=makes, max_sales=max_sales)


@app.route("/vehicles/<int:vehicle_id>/contract")
@login_required
def sale_contract(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Sales contract is not available until a sale is recorded.", 404
    balance_due = sale["sale_price_inc_gst"] - sale["deposit_amount"] - sale["trade_in_value"]
    return render_template(
        "sale_contract.html",
        vehicle=vehicle,
        sale=sale,
        balance_due=balance_due,
    )


@app.route("/customers")
@login_required
def customer_history():
    conn = db()
    q = request.args.get("q", "").strip()
    if q:
        pattern = f"%{q}%"
        rows = conn.execute("""
            SELECT s.*,v.stock_no,v.year,v.make,v.model,v.registration
            FROM sales s
            JOIN vehicles v ON v.id=s.vehicle_id
            WHERE s.buyer_name LIKE ? OR s.buyer_phone LIKE ? OR s.buyer_email LIKE ?
            ORDER BY s.sale_date DESC,s.id DESC
        """, (pattern, pattern, pattern)).fetchall()
    else:
        rows = conn.execute("""
            SELECT s.*,v.stock_no,v.year,v.make,v.model,v.registration
            FROM sales s
            JOIN vehicles v ON v.id=s.vehicle_id
            ORDER BY s.sale_date DESC,s.id DESC
        """).fetchall()
    conn.close()
    return render_template("customer_history.html", rows=rows, q=q)


@app.route("/reports/bas")
@login_required
def bas_report():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    conn = db()

    clauses = []
    params = []
    if start:
        clauses.append("date_value>=?")
        params.append(start)
    if end:
        clauses.append("date_value<=?")
        params.append(end)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""

    sales_rows = conn.execute(f"""
        SELECT sale_date AS date_value,sale_price_inc_gst,sale_gst
        FROM sales
        {where.replace('date_value','sale_date')}
    """, params).fetchall()

    vehicle_rows = conn.execute(f"""
        SELECT purchase_date AS date_value,purchase_price_inc_gst,purchase_gst
        FROM vehicles
        {where.replace('date_value','purchase_date')}
    """, params).fetchall()

    expense_rows = conn.execute(f"""
        SELECT expense_date AS date_value,cost_inc_gst,gst_amount
        FROM expenses
        {where.replace('date_value','expense_date')}
    """, params).fetchall()

    job_rows = conn.execute(f"""
        SELECT job_date AS date_value,
               CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END AS cost_value,
               gst_amount
        FROM job_cards
        {where.replace('date_value','job_date')}
    """, params).fetchall()

    service_rows = conn.execute(f"""
        SELECT service_date AS date_value,cost_inc_gst,gst_amount
        FROM service_entries
        {where.replace('date_value','service_date')}
    """, params).fetchall()

    part_rows = conn.execute(f"""
        SELECT usage_date AS date_value,quantity_used*unit_cost_inc_gst AS cost_value,
               (quantity_used*unit_cost_inc_gst)/11.0 AS gst_amount
        FROM part_usage
        {where.replace('date_value','usage_date')}
    """, params).fetchall()
    conn.close()

    sales_total = sum(r["sale_price_inc_gst"] for r in sales_rows)
    gst_collected = sum(r["sale_gst"] for r in sales_rows)
    purchase_total = sum(r["purchase_price_inc_gst"] for r in vehicle_rows)
    gst_vehicle = sum(r["purchase_gst"] for r in vehicle_rows)
    expense_total = sum(r["cost_inc_gst"] for r in expense_rows)
    gst_expenses = sum(r["gst_amount"] for r in expense_rows)
    job_total = sum(r["cost_value"] for r in job_rows)
    gst_jobs = sum(r["gst_amount"] for r in job_rows)
    service_total = sum(r["cost_inc_gst"] for r in service_rows)
    gst_services = sum(r["gst_amount"] for r in service_rows)
    parts_total = sum(r["cost_value"] for r in part_rows)
    gst_parts = sum(r["gst_amount"] for r in part_rows)

    gst_paid = gst_vehicle + gst_expenses + gst_jobs + gst_services + gst_parts
    net_gst = gst_collected - gst_paid

    return render_template(
        "bas_report.html",
        start=start,
        end=end,
        sales_total=sales_total,
        gst_collected=gst_collected,
        purchase_total=purchase_total,
        expense_total=expense_total,
        job_total=job_total,
        service_total=service_total,
        parts_total=parts_total,
        gst_paid=gst_paid,
        net_gst=net_gst,
    )


@app.route("/audit")
@login_required
def audit_page():
    conn = db()
    rows = conn.execute("""
        SELECT * FROM audit_log
        ORDER BY created_at DESC,id DESC
        LIMIT 500
    """).fetchall()
    conn.close()
    return render_template("audit_log.html", rows=rows)


@app.route("/quick-add-vehicle", methods=["POST"])
@login_required
def quick_add_vehicle():
    conn = db()
    stock_no = next_stock_number(conn)
    price = float(request.form.get("purchase_price_inc_gst") or 0)
    gst = round(price / 11, 2)

    cursor = conn.execute("""
        INSERT INTO vehicles(
            stock_no,status,purchase_date,make,model,variant,year,vin,registration,
            odometer_km,colour,purchase_price_inc_gst,purchase_gst,
            barry_contribution,matt_contribution,rego_expiry,photo_filename,notes,
            ppsr_number,roadworthy_status,service_due_date,service_history
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        stock_no,
        request.form.get("status") or "In Stock",
        request.form.get("purchase_date"),
        request.form.get("make"),
        request.form.get("model"),
        request.form.get("variant"),
        int(request.form.get("year") or 0) or None,
        request.form.get("vin") or None,
        request.form.get("registration"),
        int(request.form.get("odometer_km") or 0) or None,
        request.form.get("colour"),
        price,
        gst,
        float(request.form.get("barry_contribution") or 0),
        float(request.form.get("matt_contribution") or 0),
        request.form.get("rego_expiry"),
        None,
        request.form.get("notes"),
        request.form.get("ppsr_number"),
        request.form.get("roadworthy_status") or "Not Checked",
        request.form.get("service_due_date"),
        request.form.get("service_history"),
    ))
    vehicle_id = cursor.lastrowid
    conn.commit()
    conn.close()

    log_action("Quick vehicle added", "vehicle", vehicle_id, stock_no)
    flash(f"{stock_no} added successfully.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/photos/featured/<int:photo_id>", methods=["POST"])
@login_required
def set_featured_photo(vehicle_id, photo_id):
    conn = db()
    photo = conn.execute(
        "SELECT * FROM vehicle_photos WHERE id=? AND vehicle_id=?",
        (photo_id, vehicle_id),
    ).fetchone()
    if not photo:
        conn.close()
        flash("Photo not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute(
        "UPDATE vehicles SET featured_photo_id=?,photo_filename=? WHERE id=?",
        (photo_id, photo["filename"], vehicle_id),
    )
    conn.commit()
    conn.close()
    log_action("Featured photo changed", "vehicle", vehicle_id, photo["filename"])
    flash("Featured photo updated.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/photos/delete/<int:photo_id>", methods=["POST"])
@login_required
def delete_vehicle_photo(vehicle_id, photo_id):
    conn = db()
    photo = conn.execute(
        "SELECT * FROM vehicle_photos WHERE id=? AND vehicle_id=?",
        (photo_id, vehicle_id),
    ).fetchone()
    if not photo:
        conn.close()
        flash("Photo not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute("DELETE FROM vehicle_photos WHERE id=?", (photo_id,))
    vehicle = conn.execute("SELECT featured_photo_id FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if vehicle and vehicle["featured_photo_id"] == photo_id:
        replacement = conn.execute(
            "SELECT id,filename FROM vehicle_photos WHERE vehicle_id=? ORDER BY id DESC LIMIT 1",
            (vehicle_id,),
        ).fetchone()
        conn.execute(
            "UPDATE vehicles SET featured_photo_id=?,photo_filename=? WHERE id=?",
            (replacement["id"] if replacement else None, replacement["filename"] if replacement else None, vehicle_id),
        )
    conn.commit()
    conn.close()

    try:
        file_path = UPLOAD_DIR / photo["filename"]
        if file_path.exists():
            file_path.unlink()
    except OSError:
        pass

    log_action("Vehicle photo deleted", "vehicle", vehicle_id, photo["filename"])
    flash("Photo deleted.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/valuation", methods=["GET", "POST"])
@login_required
def vehicle_valuation(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    expenses = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM expenses WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    jobs = conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) AS v FROM job_cards WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    services = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM service_entries WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    parts = conn.execute("SELECT COALESCE(SUM(quantity_used*unit_cost_inc_gst),0) AS v FROM part_usage WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]

    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    if request.method == "POST":
        conn.execute("""
            UPDATE vehicles
            SET estimated_sale_price=?,minimum_sale_price=?,valuation_notes=?
            WHERE id=?
        """, (
            float(request.form.get("estimated_sale_price") or 0),
            float(request.form.get("minimum_sale_price") or 0),
            request.form.get("valuation_notes"),
            vehicle_id,
        ))
        conn.commit()
        vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
        log_action("Vehicle valuation updated", "vehicle", vehicle_id, request.form.get("valuation_notes"))
        flash("Valuation saved.", "success")

    conn.close()
    total_cost = vehicle["purchase_price_inc_gst"] + expenses + jobs + services + parts
    estimated_profit = vehicle["estimated_sale_price"] - total_cost
    minimum_profit = vehicle["minimum_sale_price"] - total_cost
    return render_template(
        "vehicle_valuation.html",
        vehicle=vehicle,
        expenses=expenses,
        jobs=jobs,
        services=services,
        parts=parts,
        total_cost=total_cost,
        estimated_profit=estimated_profit,
        minimum_profit=minimum_profit,
    )


@app.route("/vehicles/<int:vehicle_id>/purchase-agreement")
@login_required
def purchase_agreement(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404
    return render_template("purchase_agreement.html", vehicle=vehicle)


@app.route("/vehicles/<int:vehicle_id>/receipt")
@login_required
def vehicle_receipt(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Receipt is available after a sale is recorded.", 404
    balance_paid = sale["sale_price_inc_gst"] - sale["deposit_amount"] - sale["trade_in_value"]
    return render_template(
        "vehicle_receipt.html",
        vehicle=vehicle,
        sale=sale,
        balance_paid=balance_paid,
    )


@app.route("/vehicles/<int:vehicle_id>/reminders", methods=["POST"])
@login_required
def add_reminder(vehicle_id):
    conn = db()
    conn.execute("""
        INSERT INTO reminders(vehicle_id,reminder_date,reminder_type,title,notes)
        VALUES(?,?,?,?,?)
    """, (
        vehicle_id,
        request.form.get("reminder_date"),
        request.form.get("reminder_type"),
        request.form.get("title"),
        request.form.get("notes"),
    ))
    conn.commit()
    conn.close()
    flash("Reminder added.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/reminders/<int:reminder_id>/complete", methods=["POST"])
@login_required
def complete_reminder(reminder_id):
    conn = db()
    reminder = conn.execute("SELECT * FROM reminders WHERE id=?", (reminder_id,)).fetchone()
    if not reminder:
        conn.close()
        return "Reminder not found", 404
    conn.execute("UPDATE reminders SET completed=1 WHERE id=?", (reminder_id,))
    conn.commit()
    vehicle_id = reminder["vehicle_id"]
    conn.close()
    flash("Reminder completed.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/reminders")
@login_required
def reminders_page():
    conn = db()
    rows = conn.execute("""
        SELECT r.*,v.stock_no,v.make,v.model
        FROM reminders r
        LEFT JOIN vehicles v ON v.id=r.vehicle_id
        ORDER BY r.completed,r.reminder_date,r.id
    """).fetchall()
    conn.close()
    return render_template("reminders.html", rows=rows)


@app.route("/tasks", methods=["GET", "POST"])
@login_required
def tasks_page():
    conn = db()
    if request.method == "POST":
        conn.execute("""
            INSERT INTO tasks(
                vehicle_id,task_date,due_date,title,category,assigned_to,priority,status,notes
            ) VALUES(?,?,?,?,?,?,?,?,?)
        """, (
            int(request.form.get("vehicle_id")) if request.form.get("vehicle_id") else None,
            request.form.get("task_date"),
            request.form.get("due_date"),
            request.form.get("title"),
            request.form.get("category"),
            request.form.get("assigned_to"),
            request.form.get("priority") or "Normal",
            request.form.get("status") or "Open",
            request.form.get("notes"),
        ))
        conn.commit()
        conn.close()
        flash("Task added.", "success")
        return redirect(url_for("tasks_page"))

    rows = conn.execute("""
        SELECT t.*,v.stock_no,v.make,v.model
        FROM tasks t
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        ORDER BY CASE WHEN t.status='Completed' THEN 1 ELSE 0 END,
                 CASE t.priority WHEN 'Urgent' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
                 COALESCE(t.due_date,'9999-12-31'),t.id DESC
    """).fetchall()
    vehicles = conn.execute("""
        SELECT id,stock_no,make,model FROM vehicles
        WHERE status NOT IN ('Sold','BER')
        ORDER BY stock_no
    """).fetchall()
    conn.close()
    return render_template("tasks.html", rows=rows, vehicles=vehicles)


@app.route("/tasks/<int:task_id>/complete", methods=["POST"])
@login_required
def complete_task(task_id):
    conn = db()
    conn.execute("UPDATE tasks SET status='Completed' WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    flash("Task completed.", "success")
    return redirect(request.referrer or url_for("tasks_page"))


@app.route("/vehicles/<int:vehicle_id>/prepare-for-sale", methods=["POST"])
@login_required
def prepare_for_sale(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    today_text = date.today().isoformat()
    tasks = [
        ("Workshop", "Complete mechanical inspection", "Barry", "High"),
        ("Compliance", "Confirm PPSR and registration details", "Barry", "High"),
        ("Presentation", "Complete detailing and final clean", "Matt", "Normal"),
        ("Photography", "Upload complete vehicle photo set", "Matt", "Normal"),
        ("Advertising", "Create advertisement and window card", "Barry", "Normal"),
        ("Sales", "Confirm asking price and minimum sale price", "Barry", "High"),
    ]
    for category, title, assigned_to, priority in tasks:
        exists = conn.execute("""
            SELECT id FROM tasks
            WHERE vehicle_id=? AND title=? AND status!='Completed'
        """, (vehicle_id, title)).fetchone()
        if not exists:
            conn.execute("""
                INSERT INTO tasks(vehicle_id,task_date,due_date,title,category,assigned_to,priority,status)
                VALUES(?,?,?,?,?,?,?,'Open')
            """, (
                vehicle_id,
                today_text,
                (date.today() + timedelta(days=7)).isoformat(),
                title,
                category,
                assigned_to,
                priority,
            ))

    conn.execute(
        "UPDATE vehicles SET status='Being Repaired' WHERE id=? AND status='In Stock'",
        (vehicle_id,),
    )
    conn.commit()
    conn.close()
    log_action("Prepare for sale workflow created", "vehicle", vehicle_id, vehicle["stock_no"])
    flash("Prepare-for-sale tasks created.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/finance-calculator", methods=["GET", "POST"])
@login_required
def finance_calculator():
    result = None
    values = {
        "vehicle_price": request.form.get("vehicle_price", ""),
        "deposit": request.form.get("deposit", ""),
        "trade_in": request.form.get("trade_in", ""),
        "annual_rate": request.form.get("annual_rate", ""),
        "term_years": request.form.get("term_years", "5"),
        "balloon": request.form.get("balloon", "0"),
    }
    if request.method == "POST":
        price = float(values["vehicle_price"] or 0)
        deposit = float(values["deposit"] or 0)
        trade_in = float(values["trade_in"] or 0)
        annual_rate = float(values["annual_rate"] or 0) / 100
        years = int(float(values["term_years"] or 0))
        balloon = float(values["balloon"] or 0)
        principal = max(price - deposit - trade_in, 0)
        months = max(years * 12, 1)
        monthly_rate = annual_rate / 12

        financed_before_balloon = max(principal - balloon, 0)
        if monthly_rate > 0:
            monthly_payment = financed_before_balloon * (
                monthly_rate * (1 + monthly_rate) ** months
            ) / ((1 + monthly_rate) ** months - 1)
        else:
            monthly_payment = financed_before_balloon / months

        total_monthly = monthly_payment * months
        total_repaid = total_monthly + balloon
        total_interest = total_repaid - principal
        result = {
            "principal": principal,
            "monthly": monthly_payment,
            "weekly": monthly_payment * 12 / 52,
            "fortnightly": monthly_payment * 12 / 26,
            "total_repaid": total_repaid,
            "total_interest": total_interest,
            "balloon": balloon,
        }

    return render_template("finance_calculator.html", result=result, values=values)


@app.route("/reports/stock-age")
@login_required
def stock_age_report():
    conn = db()
    today_text = date.today().isoformat()
    rows = conn.execute("""
        SELECT v.*,
          CAST(julianday(?) - julianday(COALESCE(v.purchase_date,substr(v.created_at,1,10))) AS INTEGER) AS days_in_stock,
          COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
          COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                    FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs
        FROM vehicles v
        WHERE v.status NOT IN ('Sold','BER')
        ORDER BY days_in_stock DESC
    """, (today_text,)).fetchall()
    conn.close()
    return render_template("stock_age_report.html", rows=rows)


@app.route("/vehicles/<int:vehicle_id>/advertisement-pro")
@login_required
def advertisement_pro(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    photos = conn.execute(
        "SELECT * FROM vehicle_photos WHERE vehicle_id=? ORDER BY id",
        (vehicle_id,),
    ).fetchall()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404

    title = vehicle["advertisement_title"] or (
        f"{vehicle['year'] or ''} {vehicle['make']} {vehicle['model']} {vehicle['variant'] or ''}".strip()
    )
    description = vehicle["advertisement_description"] or (
        f"{title}\n\n"
        f"• {vehicle['odometer_km'] or 0:,} km\n"
        f"• Registration: {vehicle['registration'] or 'Not listed'}\n"
        f"• Colour: {vehicle['colour'] or 'Not listed'}\n"
        f"• Roadworthy: {vehicle['roadworthy_status'] or 'Not checked'}\n\n"
        f"{vehicle['notes'] or ''}\n\n"
        "Contact BAM Motor Group — Buy • Sell • Trade."
    )
    marketplace = description
    carsales = description + "\n\nStock Number: " + vehicle["stock_no"]
    gumtree = description + "\n\nInspection by appointment."
    return render_template(
        "advertisement_pro.html",
        vehicle=vehicle,
        photos=photos,
        title=title,
        marketplace=marketplace,
        carsales=carsales,
        gumtree=gumtree,
    )

@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.template_filter("money")
def money(value):
    try:
        return f"${float(value or 0):,.2f}"
    except Exception:
        return "$0.00"

if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="127.0.0.1", port=5000)
