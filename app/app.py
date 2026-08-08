
import os
import sqlite3
import re
import socket
import io
import csv
import json
import shutil
import zipfile
import urllib.request
import urllib.parse
import urllib.error
import statistics
from datetime import datetime, date, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
PROJECT_DIR = APP_DIR.parent
BASE_DIR = PROJECT_DIR

# Azure App Service keeps /home persistent between restarts. Local Windows use
# continues to default to the project folders, so the same build works both ways.
DATA_ROOT = Path(os.environ.get("BAM_DATA_ROOT", str(PROJECT_DIR))).expanduser().resolve()
DB_DIR = Path(os.environ.get("BAM_DB_DIR", str(DATA_ROOT / "database"))).expanduser().resolve()
UPLOAD_DIR = Path(os.environ.get("BAM_UPLOAD_DIR", str(DATA_ROOT / "uploads"))).expanduser().resolve()
BACKUP_DIR = Path(os.environ.get("BAM_BACKUP_DIR", str(DATA_ROOT / "backups"))).expanduser().resolve()
REPORT_DIR = Path(os.environ.get("BAM_REPORT_DIR", str(DATA_ROOT / "reports"))).expanduser().resolve()

for folder in (DB_DIR, UPLOAD_DIR, BACKUP_DIR, REPORT_DIR):
    folder.mkdir(parents=True, exist_ok=True)

DB_PATH = Path(os.environ.get("BAM_SQLITE_PATH", str(DB_DIR / "bam_motor_group.db"))).expanduser().resolve()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
app.secret_key = os.environ.get("BAM_SECRET_KEY", "change-this-secret-key-before-production")
app.config.update(
    MAX_CONTENT_LENGTH=int(os.environ.get("BAM_MAX_UPLOAD_MB", "100")) * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("BAM_SECURE_COOKIES", "0") == "1",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=int(os.environ.get("BAM_SESSION_HOURS", "12"))),
)

APP_VERSION = "24.1"
APP_NAME = "BAM Dealer Enterprise Cloud"
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5").strip() or "gpt-5"
VIN_DECODER_URL = os.environ.get("BAM_VIN_DECODER_URL", "https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVinValuesExtended/{vin}?format=json")
WORKSHOP_PROVIDER_NAME = os.environ.get("BAM_WORKSHOP_PROVIDER_NAME", "eManualOnline").strip() or "eManualOnline"
WORKSHOP_PORTAL_URL = os.environ.get("BAM_WORKSHOP_PORTAL_URL", "https://www.emanualonline.com/cars/").strip() or "https://www.emanualonline.com/cars/"
DEFAULT_LABOUR_RATE = float(os.environ.get("BAM_LABOUR_RATE", "145") or 145)
EMAIL_WEBMAIL_URL = os.environ.get("BAM_EMAIL_WEBMAIL_URL", "").strip()
EMAIL_ADDRESS = os.environ.get("BAM_EMAIL_ADDRESS", "").strip()

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "pdf"}
BACKUP_EXTENSIONS = {"zip"}


@app.context_processor
def inject_app_identity():
    return {"app_version": APP_VERSION, "app_name": APP_NAME}

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def ensure_column(conn, table, column, definition):
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")



def init_db():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        display_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'owner'
    );

    CREATE TABLE IF NOT EXISTS vehicles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stock_no TEXT UNIQUE NOT NULL,
        status TEXT NOT NULL DEFAULT 'In Stock',
        purchase_date TEXT,
        make TEXT NOT NULL,
        model TEXT NOT NULL,
        variant TEXT,
        year INTEGER,
        vin TEXT UNIQUE,
        registration TEXT,
        odometer_km INTEGER,
        colour TEXT,
        purchase_price_inc_gst REAL NOT NULL DEFAULT 0,
        purchase_gst REAL NOT NULL DEFAULT 0,
        barry_contribution REAL NOT NULL DEFAULT 0,
        matt_contribution REAL NOT NULL DEFAULT 0,
        rego_expiry TEXT,
        photo_filename TEXT,
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        expense_date TEXT,
        category TEXT NOT NULL,
        description TEXT NOT NULL,
        supplier TEXT,
        paid_by TEXT NOT NULL,
        cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        receipt_filename TEXT,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER UNIQUE NOT NULL,
        sale_date TEXT,
        buyer_name TEXT,
        buyer_phone TEXT,
        sale_price_inc_gst REAL NOT NULL DEFAULT 0,
        sale_gst REAL NOT NULL DEFAULT 0,
        advertising_cost REAL NOT NULL DEFAULT 0,
        transfer_cost REAL NOT NULL DEFAULT 0,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS contacts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        contact_type TEXT,
        name TEXT NOT NULL,
        phone TEXT,
        email TEXT,
        address TEXT,
        licence_no TEXT,
        notes TEXT
    );

    CREATE TABLE IF NOT EXISTS consumables (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        purchase_date TEXT,
        item TEXT NOT NULL,
        category TEXT,
        supplier TEXT,
        purchased_by TEXT,
        qty_purchased REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        qty_used REAL NOT NULL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS vehicle_photos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        filename TEXT NOT NULL,
        caption TEXT,
        uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS job_cards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        job_date TEXT,
        category TEXT,
        description TEXT NOT NULL,
        supplier TEXT,
        paid_by TEXT,
        estimated_cost REAL NOT NULL DEFAULT 0,
        actual_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        status TEXT NOT NULL DEFAULT 'Open',
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS job_card_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_card_id INTEGER NOT NULL,
        old_status TEXT,
        new_status TEXT NOT NULL,
        changed_by TEXT,
        changed_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        note TEXT,
        FOREIGN KEY(job_card_id) REFERENCES job_cards(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS vehicle_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        document_type TEXT NOT NULL,
        filename TEXT NOT NULL,
        description TEXT,
        uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS service_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER NOT NULL,
        service_date TEXT,
        odometer_km INTEGER,
        service_type TEXT,
        description TEXT NOT NULL,
        supplier TEXT,
        cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount REAL NOT NULL DEFAULT 0,
        paid_by TEXT,
        next_service_date TEXT,
        notes TEXT,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS parts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        part_number TEXT UNIQUE,
        part_name TEXT NOT NULL,
        category TEXT,
        supplier TEXT,
        quantity_on_hand REAL NOT NULL DEFAULT 0,
        reorder_level REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        gst_amount_per_unit REAL NOT NULL DEFAULT 0,
        storage_location TEXT,
        notes TEXT
    );

    CREATE TABLE IF NOT EXISTS part_usage (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        part_id INTEGER NOT NULL,
        vehicle_id INTEGER NOT NULL,
        job_card_id INTEGER,
        usage_date TEXT,
        quantity_used REAL NOT NULL DEFAULT 0,
        unit_cost_inc_gst REAL NOT NULL DEFAULT 0,
        paid_by TEXT,
        notes TEXT,
        FOREIGN KEY(part_id) REFERENCES parts(id),
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE,
        FOREIGN KEY(job_card_id) REFERENCES job_cards(id) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT,
        action TEXT NOT NULL,
        entity_type TEXT,
        entity_id INTEGER,
        details TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS reminders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER,
        reminder_date TEXT NOT NULL,
        reminder_type TEXT NOT NULL,
        title TEXT NOT NULL,
        notes TEXT,
        completed INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        vehicle_id INTEGER,
        task_date TEXT,
        due_date TEXT,
        title TEXT NOT NULL,
        category TEXT,
        assigned_to TEXT,
        priority TEXT NOT NULL DEFAULT 'Normal',
        status TEXT NOT NULL DEFAULT 'Open',
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
    );
    """)

    # Upgrade columns from older BAM databases.
    ensure_column(conn, "users", "is_active", "INTEGER DEFAULT 1")
    ensure_column(conn, "users", "created_at", "TEXT")
    ensure_column(conn, "users", "last_login", "TEXT")

    ensure_column(conn, "vehicles", "ppsr_number", "TEXT")
    ensure_column(conn, "vehicles", "roadworthy_status", "TEXT DEFAULT 'Not Checked'")
    ensure_column(conn, "vehicles", "service_due_date", "TEXT")
    ensure_column(conn, "vehicles", "service_history", "TEXT")
    ensure_column(conn, "vehicles", "asking_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "advertisement_title", "TEXT")
    ensure_column(conn, "vehicles", "advertisement_description", "TEXT")

    ensure_column(conn, "vehicles", "featured_photo_id", "INTEGER")
    ensure_column(conn, "vehicles", "estimated_sale_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "minimum_sale_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "valuation_notes", "TEXT")

    ensure_column(conn, "vehicles", "advertised_date", "TEXT")
    ensure_column(conn, "vehicles", "ready_for_sale_date", "TEXT")
    ensure_column(conn, "vehicles", "target_profit", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "negotiated_price", "REAL DEFAULT 0")

    ensure_column(conn, "vehicles", "asset_type", "TEXT DEFAULT 'Car'")
    ensure_column(conn, "vehicles", "length_m", "REAL")
    ensure_column(conn, "vehicles", "tare_weight_kg", "REAL")
    ensure_column(conn, "vehicles", "atm_kg", "REAL")
    ensure_column(conn, "vehicles", "gtm_kg", "REAL")
    ensure_column(conn, "vehicles", "berths", "INTEGER")
    ensure_column(conn, "vehicles", "axles", "INTEGER")
    ensure_column(conn, "vehicles", "caravan_features", "TEXT")
    ensure_column(conn, "vehicles", "boat_type", "TEXT")
    ensure_column(conn, "vehicles", "hull_material", "TEXT")
    ensure_column(conn, "vehicles", "engine_make", "TEXT")
    ensure_column(conn, "vehicles", "engine_model", "TEXT")
    ensure_column(conn, "vehicles", "engine_hours", "REAL")
    ensure_column(conn, "vehicles", "horsepower", "REAL")
    ensure_column(conn, "vehicles", "fuel_type", "TEXT")
    ensure_column(conn, "vehicles", "hin", "TEXT")
    ensure_column(conn, "vehicles", "trailer_included", "INTEGER DEFAULT 0")
    ensure_column(conn, "vehicles", "trailer_registration", "TEXT")
    ensure_column(conn, "vehicles", "capacity_people", "INTEGER")
    ensure_column(conn, "vehicles", "boat_features", "TEXT")

    # Version 22 - Dealer Intelligence
    ensure_column(conn, "vehicles", "vin_decode_json", "TEXT")
    ensure_column(conn, "vehicles", "decoded_at", "TEXT")
    ensure_column(conn, "vehicles", "manufacturer_name", "TEXT")
    ensure_column(conn, "vehicles", "body_class", "TEXT")
    ensure_column(conn, "vehicles", "drive_type", "TEXT")
    ensure_column(conn, "vehicles", "transmission_style", "TEXT")
    ensure_column(conn, "vehicles", "engine_cylinders", "TEXT")
    ensure_column(conn, "vehicles", "engine_displacement_l", "TEXT")
    ensure_column(conn, "vehicles", "fuel_type_primary", "TEXT")
    ensure_column(conn, "vehicles", "decoded_model_year", "TEXT")
    ensure_column(conn, "vehicles", "decoded_series", "TEXT")
    ensure_column(conn, "vehicles", "decoded_trim", "TEXT")
    ensure_column(conn, "vehicles", "market_price_low", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "market_price_mid", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "market_price_high", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "market_price_checked_at", "TEXT")

    # Version 18 - Parts Vehicle / Dismantling
    ensure_column(conn, "vehicles", "vehicle_purpose", "TEXT DEFAULT 'Retail Sale'")
    ensure_column(conn, "vehicles", "dismantling_status", "TEXT DEFAULT 'Not Started'")
    ensure_column(conn, "vehicles", "shell_sale_price", "REAL DEFAULT 0")
    ensure_column(conn, "vehicles", "shell_sale_date", "TEXT")

    ensure_column(conn, "sales", "invoice_number", "TEXT")
    ensure_column(conn, "sales", "buyer_email", "TEXT")
    ensure_column(conn, "sales", "buyer_address", "TEXT")

    ensure_column(conn, "sales", "deposit_amount", "REAL DEFAULT 0")
    ensure_column(conn, "sales", "deposit_date", "TEXT")
    ensure_column(conn, "sales", "payment_method", "TEXT")
    ensure_column(conn, "sales", "trade_in_description", "TEXT")
    ensure_column(conn, "sales", "trade_in_value", "REAL DEFAULT 0")
    ensure_column(conn, "sales", "warranty_type", "TEXT")
    ensure_column(conn, "sales", "warranty_expiry", "TEXT")
    ensure_column(conn, "sales", "contract_number", "TEXT")
    ensure_column(conn, "sales", "invoice_status", "TEXT DEFAULT 'Draft'")
    ensure_column(conn, "sales", "updated_at", "TEXT")
    ensure_column(conn, "sales", "delivery_status", "TEXT DEFAULT 'Preparing'")
    ensure_column(conn, "sales", "delivery_date", "TEXT")
    ensure_column(conn, "sales", "keys_handed_over", "INTEGER DEFAULT 0")
    ensure_column(conn, "sales", "registration_transferred", "INTEGER DEFAULT 0")
    ensure_column(conn, "sales", "customer_signature_received", "INTEGER DEFAULT 0")
    ensure_column(conn, "sales", "finance_documents_complete", "INTEGER DEFAULT 0")
    ensure_column(conn, "sales", "warranty_documents_complete", "INTEGER DEFAULT 0")
    ensure_column(conn, "sales", "deal_notes", "TEXT")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS invoice_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            vehicle_id INTEGER NOT NULL,
            invoice_number TEXT,
            changed_by TEXT,
            change_note TEXT,
            old_values TEXT,
            new_values TEXT,
            changed_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
        )
    """)

    conn.executescript("""
        CREATE TABLE IF NOT EXISTS part_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            caption TEXT,
            is_featured INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (part_id)
                REFERENCES parts(id)
                ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS part_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_id INTEGER NOT NULL,
            quantity INTEGER DEFAULT 1,
            customer_name TEXT,
            customer_phone TEXT,
            customer_email TEXT,
            sale_price REAL DEFAULT 0,
            freight_cost REAL DEFAULT 0,
            payment_method TEXT,
            warranty TEXT,
            invoice_number TEXT,
            sale_date TEXT DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY (part_id) REFERENCES parts(id)
        );


        CREATE TABLE IF NOT EXISTS part_shipments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            shipment_number TEXT UNIQUE NOT NULL,
            part_id INTEGER,
            quantity REAL NOT NULL DEFAULT 1,
            customer_name TEXT NOT NULL,
            customer_phone TEXT,
            customer_email TEXT,
            address_line TEXT NOT NULL,
            suburb TEXT,
            state TEXT,
            postcode TEXT,
            courier TEXT,
            tracking_number TEXT,
            parcel_weight_kg REAL DEFAULT 0,
            parcel_length_cm REAL DEFAULT 0,
            parcel_width_cm REAL DEFAULT 0,
            parcel_height_cm REAL DEFAULT 0,
            shipping_cost REAL DEFAULT 0,
            freight_charged REAL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'Ready to Pack',
            date_created TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            date_sent TEXT,
            date_delivered TEXT,
            invoice_included INTEGER NOT NULL DEFAULT 0,
            bubble_wrapped INTEGER NOT NULL DEFAULT 0,
            box_sealed INTEGER NOT NULL DEFAULT 0,
            tracking_sent INTEGER NOT NULL DEFAULT 0,
            stock_adjusted INTEGER NOT NULL DEFAULT 0,
            notes TEXT,
            FOREIGN KEY(part_id) REFERENCES parts(id) ON DELETE SET NULL
        );
    """)
    ensure_column(conn, "parts", "vehicle_id", "INTEGER")
    ensure_column(conn, "parts", "vehicle_stock_no", "TEXT")
    ensure_column(conn, "parts", "vin", "TEXT")
    ensure_column(conn, "parts", "make", "TEXT")
    ensure_column(conn, "parts", "model", "TEXT")
    ensure_column(conn, "parts", "year", "INTEGER")
    ensure_column(conn, "parts", "subcategory", "TEXT")
    ensure_column(conn, "parts", "description", "TEXT")
    ensure_column(conn, "parts", "condition", "TEXT DEFAULT 'Used'")
    ensure_column(conn, "parts", "selling_price", "REAL DEFAULT 0")
    ensure_column(conn, "parts", "status", "TEXT DEFAULT 'In Stock'")
    ensure_column(conn, "parts", "engine_code", "TEXT")
    ensure_column(conn, "parts", "transmission_code", "TEXT")
    ensure_column(conn, "parts", "barcode", "TEXT")
    ensure_column(conn, "parts", "date_added", "TEXT")

    # Version 19 - Professional Parts Inventory
    ensure_column(conn, "parts", "position", "TEXT")
    ensure_column(conn, "parts", "fitment", "TEXT")
    ensure_column(conn, "parts", "manufacturer_part_no", "TEXT")
    ensure_column(conn, "parts", "reserved_for", "TEXT")
    ensure_column(conn, "parts", "reserved_until", "TEXT")
    ensure_column(conn, "parts", "updated_at", "TEXT")
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS equipment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_no TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            category TEXT,
            brand TEXT,
            model TEXT,
            serial_number TEXT,
            purchase_date TEXT,
            purchase_price REAL NOT NULL DEFAULT 0,
            current_value REAL NOT NULL DEFAULT 0,
            supplier TEXT,
            warranty_expiry TEXT,
            location TEXT,
            assigned_to TEXT,
            condition TEXT NOT NULL DEFAULT 'Good',
            status TEXT NOT NULL DEFAULT 'Available',
            next_service_date TEXT,
            calibration_due TEXT,
            test_tag_due TEXT,
            photo_filename TEXT,
            receipt_filename TEXT,
            notes TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS equipment_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            assigned_to TEXT,
            location TEXT,
            condition TEXT,
            notes TEXT,
            action_date TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            user_name TEXT,
            FOREIGN KEY(equipment_id) REFERENCES equipment(id) ON DELETE CASCADE
        );
    """)

    # Version 23 - Workshop Intelligence
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS workshop_operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER,
            part_id INTEGER,
            operation_code TEXT,
            system_name TEXT,
            operation_name TEXT NOT NULL,
            labour_hours REAL NOT NULL DEFAULT 0,
            labour_rate REAL NOT NULL DEFAULT 0,
            source_name TEXT,
            source_reference TEXT,
            procedure_url TEXT,
            notes TEXT,
            make TEXT,
            model TEXT,
            year INTEGER,
            engine_hint TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE,
            FOREIGN KEY(part_id) REFERENCES parts(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS workshop_references (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER NOT NULL,
            reference_type TEXT NOT NULL,
            title TEXT NOT NULL,
            provider_name TEXT,
            reference_url TEXT,
            reference_code TEXT,
            notes TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE CASCADE
        );
    """)
    ensure_column(conn, "job_cards", "labour_operation_id", "INTEGER")
    ensure_column(conn, "job_cards", "labour_hours", "REAL DEFAULT 0")
    ensure_column(conn, "job_cards", "labour_rate", "REAL DEFAULT 0")
    ensure_column(conn, "job_cards", "labour_source", "TEXT")
    ensure_column(conn, "job_cards", "labour_code", "TEXT")
    ensure_column(conn, "job_cards", "procedure_url", "TEXT")

    # Version 23.1 - eManualOnline + Email Centre
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS email_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            direction TEXT NOT NULL DEFAULT 'Incoming',
            message_date TEXT,
            sender TEXT,
            recipient TEXT,
            subject TEXT NOT NULL,
            body_excerpt TEXT,
            status TEXT NOT NULL DEFAULT 'Unread',
            follow_up_date TEXT,
            vehicle_id INTEGER,
            contact_id INTEGER,
            external_reference TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE SET NULL,
            FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
        );
    """)

    # Version 24.1 - Professional Workshop Scheduler
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS workshop_bookings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER,
            contact_id INTEGER,
            booking_date TEXT NOT NULL,
            start_time TEXT,
            end_time TEXT,
            technician TEXT,
            bay TEXT,
            job_type TEXT,
            description TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Booked',
            quoted_hours REAL NOT NULL DEFAULT 0,
            labour_rate REAL NOT NULL DEFAULT 0,
            customer_name TEXT,
            customer_phone TEXT,
            notes TEXT,
            job_card_id INTEGER,
            created_by TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id) ON DELETE SET NULL,
            FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL,
            FOREIGN KEY(job_card_id) REFERENCES job_cards(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS workshop_time_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id INTEGER NOT NULL,
            technician TEXT,
            clock_in TEXT NOT NULL,
            clock_out TEXT,
            minutes INTEGER NOT NULL DEFAULT 0,
            notes TEXT,
            created_by TEXT,
            FOREIGN KEY(booking_id) REFERENCES workshop_bookings(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_workshop_bookings_date ON workshop_bookings(booking_date);
        CREATE INDEX IF NOT EXISTS idx_workshop_bookings_vehicle ON workshop_bookings(vehicle_id);
        CREATE INDEX IF NOT EXISTS idx_workshop_time_booking ON workshop_time_entries(booking_id);
    """)

    count = conn.execute(
        "SELECT COUNT(*) AS c FROM users"
    ).fetchone()["c"]

    if count == 0:
        conn.execute(
            "INSERT INTO users(username,password_hash,display_name,role) VALUES(?,?,?,?)",
            ("barry", generate_password_hash("ChangeMe123!"), "Barry", "owner"),
        )
        conn.execute(
            "INSERT INTO users(username,password_hash,display_name,role) VALUES(?,?,?,?)",
            ("matt", generate_password_hash("ChangeMe123!"), "Matt", "owner"),
        )

    conn.commit()
    conn.close()



def next_stock_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT stock_no FROM vehicles WHERE stock_no LIKE 'BAM-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["stock_no"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"BAM-{highest + 1:05d}"


def next_invoice_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT invoice_number FROM sales WHERE invoice_number LIKE 'INV-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["invoice_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"INV-{highest + 1:05d}"



def next_part_number(conn=None):
    own_conn = conn is None
    conn = conn or db()
    highest = 0
    rows = conn.execute("SELECT part_number FROM parts WHERE part_number LIKE 'PRT-%'").fetchall()
    for row in rows:
        match = re.search(r"(\d+)$", row["part_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))
    if own_conn:
        conn.close()
    return f"PRT-{highest + 1:06d}"


def next_shipment_number(conn=None):
    own_conn = conn is None
    conn = conn or db()
    highest = 0
    rows = conn.execute(
        "SELECT shipment_number FROM part_shipments WHERE shipment_number LIKE 'SHP-%'"
    ).fetchall()
    for row in rows:
        match = re.search(r"(\d+)$", row["shipment_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))
    if own_conn:
        conn.close()
    return f"SHP-{highest + 1:05d}"

def local_network_ip():
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.connect(("8.8.8.8", 80))
        ip = sock.getsockname()[0]
        sock.close()
        return ip
    except OSError:
        return "127.0.0.1"


def log_action(action, entity_type=None, entity_id=None, details=None):
    conn = db()
    conn.execute("""
        INSERT INTO audit_log(user_name,action,entity_type,entity_id,details)
        VALUES(?,?,?,?,?)
    """, (
        session.get("display_name"),
        action,
        entity_type,
        entity_id,
        details,
    ))
    conn.commit()
    conn.close()


def next_contract_number(conn=None):
    own_conn = conn is None
    conn = conn or db()

    highest = 0
    rows = conn.execute(
        "SELECT contract_number FROM sales WHERE contract_number LIKE 'CON-%'"
    ).fetchall()

    for row in rows:
        match = re.search(r"(\d+)$", row["contract_number"] or "")
        if match:
            highest = max(highest, int(match.group(1)))

    if own_conn:
        conn.close()

    return f"CON-{highest + 1:05d}"



def _clean_text(value):
    value = str(value or "").strip()
    return value if value and value.lower() not in {"not applicable", "not available", "null", "none"} else ""


def decode_vin(vin):
    """Decode a VIN using a configurable external decoder. Returns (data, error)."""
    vin = re.sub(r"[^A-Za-z0-9]", "", vin or "").upper()
    if len(vin) != 17:
        return None, "VIN must contain exactly 17 characters."
    try:
        url = VIN_DECODER_URL.format(vin=urllib.parse.quote(vin))
        req = urllib.request.Request(url, headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"})
        with urllib.request.urlopen(req, timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8"))
        result = (payload.get("Results") or [{}])[0]
        if not result:
            return None, "VIN decoder returned no vehicle data."
        specs = {
            "vin": vin,
            "make": _clean_text(result.get("Make")),
            "model": _clean_text(result.get("Model")),
            "year": _clean_text(result.get("ModelYear")),
            "manufacturer_name": _clean_text(result.get("Manufacturer")),
            "body_class": _clean_text(result.get("BodyClass")),
            "drive_type": _clean_text(result.get("DriveType")),
            "transmission_style": _clean_text(result.get("TransmissionStyle")),
            "engine_cylinders": _clean_text(result.get("EngineCylinders")),
            "engine_displacement_l": _clean_text(result.get("DisplacementL")),
            "fuel_type_primary": _clean_text(result.get("FuelTypePrimary")),
            "series": _clean_text(result.get("Series")),
            "trim": _clean_text(result.get("Trim")),
            "raw": result,
        }
        error_text = _clean_text(result.get("ErrorText"))
        if not any(specs.get(k) for k in ("make", "model", "year", "body_class")):
            return specs, error_text or "VIN decoded, but the provider returned limited specifications."
        return specs, None
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, ValueError, json.JSONDecodeError) as exc:
        return None, f"VIN service unavailable: {exc}"


def duplicate_vehicle_matches(conn, vin="", registration="", stock_no="", exclude_id=None):
    checks = []
    params = []
    vin = (vin or "").strip().upper()
    registration = (registration or "").strip().upper()
    stock_no = (stock_no or "").strip().upper()
    if vin:
        checks.append("UPPER(COALESCE(vin,''))=?")
        params.append(vin)
    if registration:
        checks.append("UPPER(COALESCE(registration,''))=?")
        params.append(registration)
    if stock_no:
        checks.append("UPPER(COALESCE(stock_no,''))=?")
        params.append(stock_no)
    if not checks:
        return []
    sql = "SELECT id,stock_no,year,make,model,vin,registration,status FROM vehicles WHERE (" + " OR ".join(checks) + ")"
    if exclude_id:
        sql += " AND id<>?"
        params.append(exclude_id)
    return conn.execute(sql, params).fetchall()


def market_price_suggestion(conn, vehicle):
    rows = conn.execute("""
        SELECT s.sale_price_inc_gst AS price, v.year, v.make, v.model
        FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
        WHERE s.sale_price_inc_gst>0
          AND LOWER(v.make)=LOWER(?) AND LOWER(v.model)=LOWER(?)
          AND v.id<>?
        ORDER BY s.sale_date DESC
        LIMIT 20
    """, (vehicle["make"], vehicle["model"], vehicle["id"])).fetchall()
    source = "same make/model sales in BAM"
    if not rows:
        rows = conn.execute("""
            SELECT s.sale_price_inc_gst AS price, v.year, v.make, v.model
            FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
            WHERE s.sale_price_inc_gst>0 AND LOWER(v.make)=LOWER(?) AND v.id<>?
            ORDER BY s.sale_date DESC LIMIT 20
        """, (vehicle["make"], vehicle["id"])).fetchall()
        source = "same-make sales in BAM"
    prices = sorted(float(r["price"] or 0) for r in rows if float(r["price"] or 0)>0)
    floor = max(float(vehicle["purchase_price_inc_gst"] or 0), float(vehicle["minimum_sale_price"] or 0))
    if prices:
        mid = statistics.median(prices)
        low = prices[max(0, int((len(prices)-1)*0.25))]
        high = prices[min(len(prices)-1, int((len(prices)-1)*0.75))]
        if floor and mid < floor:
            mid = floor * 1.08
            low = floor
            high = max(high, mid * 1.08)
        confidence = "Good" if len(prices) >= 5 else "Limited"
    else:
        base = floor or float(vehicle["asking_price"] or 0) or 0
        low, mid, high = base, base * 1.10 if base else 0, base * 1.20 if base else 0
        source = "vehicle cost/asking price (no comparable BAM sales yet)"
        confidence = "Estimate only"
    return {"low": round(low,2), "mid": round(mid,2), "high": round(high,2), "count": len(prices), "source": source, "confidence": confidence}


def parts_price_suggestion(conn, part):
    rows = conn.execute("""
        SELECT ps.sale_price, ps.quantity
        FROM part_sales ps JOIN parts p ON p.id=ps.part_id
        WHERE ps.sale_price>0 AND (LOWER(p.part_name)=LOWER(?) OR (COALESCE(?, '')<>'' AND LOWER(COALESCE(p.category,''))=LOWER(?)))
        ORDER BY ps.sale_date DESC LIMIT 30
    """, (part["part_name"], part["category"], part["category"])).fetchall()
    unit_prices = [float(r["sale_price"] or 0)/max(float(r["quantity"] or 1), 1) for r in rows if float(r["sale_price"] or 0)>0]
    if unit_prices:
        mid = statistics.median(unit_prices)
        return {"suggested": round(mid,2), "low": round(min(unit_prices),2), "high": round(max(unit_prices),2), "count": len(unit_prices), "source": "recorded BAM parts sales"}
    current = float(part["selling_price"] or 0)
    cost = float(part["unit_cost_inc_gst"] or 0)
    suggestion = current or (cost * 1.5 if cost else 0)
    return {"suggested": round(suggestion,2), "low": round(suggestion*0.9,2) if suggestion else 0, "high": round(suggestion*1.15,2) if suggestion else 0, "count": 0, "source": "current price/cost (no comparable sales yet)"}


def generate_vehicle_ad_text(vehicle):
    title = f"{vehicle['year'] or ''} {vehicle['make']} {vehicle['model']} {vehicle['variant'] or ''}".strip()
    facts = []
    if vehicle["odometer_km"]:
        facts.append(f"{int(vehicle['odometer_km']):,} km")
    if vehicle["colour"]:
        facts.append(str(vehicle["colour"]))
    if vehicle["fuel_type_primary"] or vehicle["fuel_type"]:
        facts.append(str(vehicle["fuel_type_primary"] or vehicle["fuel_type"]))
    fallback = title + "\n\n" + " • ".join(facts) + ("\n\n" + str(vehicle["notes"]) if vehicle["notes"] else "") + "\n\nContact BAM Motor Group — Buy • Sell • Trade."
    if not OPENAI_API_KEY:
        return fallback, "Professional local draft (set OPENAI_API_KEY in Azure to enable AI generation)."
    prompt = (
        "Write a clear, professional Australian used-vehicle advertisement. Do not invent features. "
        "Use only the supplied facts. Include a short headline, a concise paragraph and bullet points. "
        "Avoid exaggerated claims. Facts: " + json.dumps({
            "stock_no": vehicle["stock_no"], "year": vehicle["year"], "make": vehicle["make"],
            "model": vehicle["model"], "variant": vehicle["variant"], "odometer_km": vehicle["odometer_km"],
            "colour": vehicle["colour"], "registration": vehicle["registration"], "roadworthy": vehicle["roadworthy_status"],
            "body_class": vehicle["body_class"], "drive_type": vehicle["drive_type"], "fuel": vehicle["fuel_type_primary"] or vehicle["fuel_type"],
            "asking_price": vehicle["asking_price"], "notes": vehicle["notes"],
        }, ensure_ascii=False)
    )
    try:
        payload = json.dumps({"model": OPENAI_MODEL, "input": prompt, "store": False}).encode("utf-8")
        req = urllib.request.Request("https://api.openai.com/v1/responses", data=payload, headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=35) as response:
            data = json.loads(response.read().decode("utf-8"))
        text = data.get("output_text")
        if not text:
            parts = []
            for item in data.get("output", []):
                for content in item.get("content", []):
                    if content.get("type") == "output_text" and content.get("text"):
                        parts.append(content["text"])
            text = "\n".join(parts).strip()
        return text or fallback, f"AI draft using {OPENAI_MODEL}"
    except Exception as exc:
        return fallback, f"AI service unavailable; local draft used ({exc})."


def ensure_smart_reminders(conn):
    today = date.today()
    def add_once(vehicle_id, due_date, title, notes):
        exists = conn.execute("SELECT id FROM reminders WHERE vehicle_id=? AND title=? AND completed=0", (vehicle_id, title)).fetchone()
        if not exists:
            conn.execute("INSERT INTO reminders(vehicle_id,reminder_date,reminder_type,title,notes) VALUES(?,?,?,?,?)", (vehicle_id, due_date, "Smart Reminder", title, notes))

    vehicles = conn.execute("SELECT * FROM vehicles WHERE status NOT IN ('Sold','BER')").fetchall()
    for v in vehicles:
        if v["rego_expiry"]:
            try:
                d = date.fromisoformat(v["rego_expiry"][:10])
                if today <= d <= today + timedelta(days=30):
                    add_once(v["id"], d.isoformat(), "Registration expires soon", f"{v['stock_no']} registration expires {d.isoformat()}.")
            except ValueError:
                pass
        if v["service_due_date"]:
            try:
                d = date.fromisoformat(v["service_due_date"][:10])
                if d <= today + timedelta(days=14):
                    add_once(v["id"], d.isoformat(), "Service due", f"{v['stock_no']} service is due {d.isoformat()}.")
            except ValueError:
                pass
        purchase_text = v["purchase_date"] or (str(v["created_at"] or "")[:10])
        try:
            bought = date.fromisoformat(purchase_text[:10])
            days = (today - bought).days
            if days >= 90:
                add_once(v["id"], today.isoformat(), "Stock aged 90+ days", f"{v['stock_no']} has been in stock {days} days. Review pricing/advertising.")
            elif days >= 60:
                add_once(v["id"], today.isoformat(), "Stock aged 60+ days", f"{v['stock_no']} has been in stock {days} days. Review next action.")
        except (ValueError, TypeError):
            pass
    drafts = conn.execute("""
        SELECT s.vehicle_id,s.sale_date,v.stock_no FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
        WHERE COALESCE(s.invoice_status,'Draft')='Draft' AND COALESCE(s.sale_date,'')<>''
    """).fetchall()
    for row in drafts:
        try:
            d = date.fromisoformat(row["sale_date"][:10])
            if d <= today - timedelta(days=7):
                add_once(row["vehicle_id"], today.isoformat(), "Invoice still draft", f"Review the sale invoice for {row['stock_no']}.")
        except ValueError:
            pass
    reserved_parts = conn.execute("""
        SELECT id,vehicle_id,part_number,part_name,reserved_until,reserved_for
        FROM parts WHERE status='Reserved' AND vehicle_id IS NOT NULL AND COALESCE(reserved_until,'')<>''
    """).fetchall()
    for part in reserved_parts:
        try:
            d = date.fromisoformat(part["reserved_until"][:10])
            if d <= today + timedelta(days=2):
                add_once(part["vehicle_id"], d.isoformat(), f"Reserved part due: {part['part_number'] or part['part_name']}", f"Reservation for {part['reserved_for'] or 'customer'} expires {d.isoformat()}.")
        except ValueError:
            pass
    conn.commit()


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def owner_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "owner":
            flash("Owner access is required.", "error")
            return redirect(url_for("dashboard"))
        return fn(*args, **kwargs)
    return wrapper

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def save_upload(file):
    if not file or not file.filename:
        return None
    if not allowed_file(file.filename):
        raise ValueError("Unsupported file type.")
    name = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(file.filename)}"
    file.save(UPLOAD_DIR / name)
    return name

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip().lower()
        password = request.form["password"]
        conn = db()
        user = conn.execute("SELECT * FROM users WHERE username=? AND COALESCE(is_active,1)=1", (username,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            conn.execute("UPDATE users SET last_login=CURRENT_TIMESTAMP WHERE id=?", (user["id"],))
            conn.commit()
            session.clear()
            session["user_id"] = user["id"]
            session["display_name"] = user["display_name"]
            session["role"] = user["role"]
            conn.close()
            return redirect(url_for("dashboard"))
        conn.close()
        flash("Incorrect username or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    conn = db()
    ensure_smart_reminders(conn)

    metrics = conn.execute("""
        SELECT
          COUNT(*) AS total_vehicles,
          SUM(CASE WHEN status='Sold' THEN 1 ELSE 0 END) AS sold,
          SUM(CASE WHEN status='BER' THEN 1 ELSE 0 END) AS ber,
          SUM(CASE WHEN status NOT IN ('Sold','BER') THEN 1 ELSE 0 END) AS active_stock,
          SUM(CASE WHEN status NOT IN ('Sold','BER') THEN purchase_price_inc_gst ELSE 0 END) AS stock_value,
          SUM(purchase_price_inc_gst) AS purchase_total
        FROM vehicles
    """).fetchone()

    expense_total = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM expenses").fetchone()["v"]
    job_total = conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) AS v FROM job_cards").fetchone()["v"]
    service_total = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM service_entries").fetchone()["v"]
    parts_total = conn.execute("SELECT COALESCE(SUM(quantity_used*unit_cost_inc_gst),0) AS v FROM part_usage").fetchone()["v"]
    sales_total = conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) AS v FROM sales").fetchone()["v"]

    gst_paid = conn.execute("SELECT COALESCE(SUM(purchase_gst),0) AS v FROM vehicles").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM expenses").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM job_cards").fetchone()["v"]
    gst_paid += conn.execute("SELECT COALESCE(SUM(gst_amount),0) AS v FROM service_entries").fetchone()["v"]
    gst_collected = conn.execute("SELECT COALESCE(SUM(sale_gst),0) AS v FROM sales").fetchone()["v"]

    barry_total = conn.execute("SELECT COALESCE(SUM(barry_contribution),0) AS v FROM vehicles").fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM expenses").fetchone()["v"]
    barry_total += conn.execute(
        "SELECT COALESCE(SUM(CASE "
        "WHEN paid_by='Barry' THEN CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END "
        "WHEN paid_by='Shared' THEN (CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)/2 "
        "ELSE 0 END),0) AS v FROM job_cards"
    ).fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM service_entries").fetchone()["v"]
    barry_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Barry' THEN quantity_used*unit_cost_inc_gst WHEN paid_by='Shared' THEN quantity_used*unit_cost_inc_gst/2 ELSE 0 END),0) AS v FROM part_usage").fetchone()["v"]

    matt_total = conn.execute("SELECT COALESCE(SUM(matt_contribution),0) AS v FROM vehicles").fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM expenses").fetchone()["v"]
    matt_total += conn.execute(
        "SELECT COALESCE(SUM(CASE "
        "WHEN paid_by='Matt' THEN CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END "
        "WHEN paid_by='Shared' THEN (CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)/2 "
        "ELSE 0 END),0) AS v FROM job_cards"
    ).fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN cost_inc_gst WHEN paid_by='Shared' THEN cost_inc_gst/2 ELSE 0 END),0) AS v FROM service_entries").fetchone()["v"]
    matt_total += conn.execute("SELECT COALESCE(SUM(CASE WHEN paid_by='Matt' THEN quantity_used*unit_cost_inc_gst WHEN paid_by='Shared' THEN quantity_used*unit_cost_inc_gst/2 ELSE 0 END),0) AS v FROM part_usage").fetchone()["v"]

    status_counts = conn.execute("""
        SELECT status, COUNT(*) AS total
        FROM vehicles
        GROUP BY status
        ORDER BY status
    """).fetchall()

    today = date.today()
    rego_30 = (today + timedelta(days=30)).isoformat()
    rego_60 = (today + timedelta(days=60)).isoformat()
    rego_90 = (today + timedelta(days=90)).isoformat()

    rego_alerts = conn.execute("""
        SELECT id,stock_no,make,model,registration,rego_expiry
        FROM vehicles
        WHERE rego_expiry IS NOT NULL AND rego_expiry!=''
          AND rego_expiry<=?
          AND status NOT IN ('Sold','BER')
        ORDER BY rego_expiry
        LIMIT 8
    """, (rego_90,)).fetchall()

    service_alerts = conn.execute("""
        SELECT id,stock_no,make,model,service_due_date
        FROM vehicles
        WHERE service_due_date IS NOT NULL AND service_due_date!=''
          AND service_due_date<=?
          AND status NOT IN ('Sold','BER')
        ORDER BY service_due_date
        LIMIT 8
    """, (rego_90,)).fetchall()

    recent_vehicles = conn.execute("""
        SELECT v.*,
          COALESCE((SELECT SUM(cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
          COALESCE((SELECT SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END)
                    FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
          COALESCE((SELECT sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v
        ORDER BY v.id DESC
        LIMIT 8
    """).fetchall()

    monthly_rows = conn.execute("""
        SELECT substr(sale_date,1,7) AS month,
               COUNT(*) AS vehicles_sold,
               COALESCE(SUM(sale_price_inc_gst),0) AS sales_total
        FROM sales
        WHERE sale_date IS NOT NULL AND sale_date!=''
        GROUP BY substr(sale_date,1,7)
        ORDER BY month DESC
        LIMIT 12
    """).fetchall()
    monthly_rows = list(reversed(monthly_rows))

    next_stock = next_stock_number(conn)

    today_text = today.isoformat()
    open_tasks = conn.execute("""
        SELECT t.*,v.stock_no,v.make,v.model
        FROM tasks t
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        WHERE t.status!='Completed'
        ORDER BY CASE t.priority WHEN 'Urgent' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
                 COALESCE(t.due_date,'9999-12-31'),t.id
        LIMIT 8
    """).fetchall()

    attention_vehicles = conn.execute("""
        SELECT v.*,
          CAST(julianday(?) - julianday(COALESCE(v.purchase_date,substr(v.created_at,1,10))) AS INTEGER) AS days_in_stock,
          COALESCE((SELECT COUNT(*) FROM reminders r WHERE r.vehicle_id=v.id AND r.completed=0),0) AS open_reminders,
          COALESCE((SELECT COUNT(*) FROM tasks t WHERE t.vehicle_id=v.id AND t.status!='Completed'),0) AS open_tasks
        FROM vehicles v
        WHERE v.status NOT IN ('Sold','BER')
        ORDER BY open_reminders DESC,open_tasks DESC,days_in_stock DESC
        LIMIT 8
    """, (today_text,)).fetchall()

    advertised_count = conn.execute(
        "SELECT COUNT(*) AS c FROM vehicles WHERE status='Advertised'"
    ).fetchone()["c"]

    sold_this_month = conn.execute("""
        SELECT COUNT(*) AS c FROM sales
        WHERE substr(sale_date,1,7)=substr(?,1,7)
    """, (today_text,)).fetchone()["c"]

    parts_metrics = conn.execute("""
        SELECT
          COUNT(*) AS part_lines,
          COALESCE(SUM(quantity_on_hand),0) AS parts_on_hand,
          COALESCE(SUM(quantity_on_hand * unit_cost_inc_gst),0) AS parts_value,
          SUM(CASE WHEN quantity_on_hand <= reorder_level THEN 1 ELSE 0 END) AS low_stock
        FROM parts
    """).fetchone()

    missing_photo_count = conn.execute("""
        SELECT COUNT(*) AS c
        FROM vehicles
        WHERE status NOT IN ('Sold','BER')
          AND COALESCE(photo_filename,'')=''
          AND NOT EXISTS (SELECT 1 FROM vehicle_photos p WHERE p.vehicle_id=vehicles.id)
    """).fetchone()["c"]

    workshop_metrics = conn.execute("""
        SELECT
          COUNT(*) AS total_jobs,
          SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END) AS open_jobs,
          SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) AS in_progress,
          SUM(CASE WHEN status='Waiting Parts' THEN 1 ELSE 0 END) AS waiting_parts,
          SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) AS completed_jobs
        FROM job_cards
    """).fetchone()

    priority_tasks = conn.execute("""
        SELECT t.*,v.stock_no,v.make,v.model
        FROM tasks t
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        WHERE t.status!='Completed'
          AND (t.priority IN ('Urgent','High') OR COALESCE(t.due_date,'9999-12-31')<=?)
        ORDER BY CASE t.priority WHEN 'Urgent' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
                 COALESCE(t.due_date,'9999-12-31'),t.id
        LIMIT 6
    """, (today_text,)).fetchall()

    ready_to_advertise = conn.execute("""
        SELECT id,stock_no,year,make,model,status,asking_price
        FROM vehicles
        WHERE status='Ready for Sale'
        ORDER BY COALESCE(ready_for_sale_date,purchase_date,substr(created_at,1,10)),id
        LIMIT 6
    """).fetchall()

    parts_sales_month = conn.execute("""
        SELECT COALESCE(SUM(sale_price),0) AS revenue, COUNT(*) AS sales_count
        FROM part_sales
        WHERE substr(sale_date,1,7)=substr(?,1,7)
    """, (today_text,)).fetchone()

    parts_retail_value = conn.execute("""
        SELECT COALESCE(SUM(quantity_on_hand * selling_price),0) AS value
        FROM parts
    """).fetchone()["value"]

    asset_summary = conn.execute("""
        SELECT COALESCE(asset_type,'Car') AS asset_type,
               COUNT(*) AS total,
               SUM(CASE WHEN status NOT IN ('Sold','BER') THEN 1 ELSE 0 END) AS active,
               COALESCE(SUM(CASE WHEN status NOT IN ('Sold','BER') THEN purchase_price_inc_gst ELSE 0 END),0) AS stock_value
        FROM vehicles
        GROUP BY COALESCE(asset_type,'Car')
        ORDER BY asset_type
    """).fetchall()

    asset_totals = {row["asset_type"]: dict(row) for row in asset_summary}

    equipment_metrics = conn.execute("""
        SELECT
          COUNT(*) AS total_equipment,
          COALESCE(SUM(current_value),0) AS current_value,
          SUM(CASE WHEN status='Checked Out' THEN 1 ELSE 0 END) AS checked_out,
          SUM(CASE WHEN status IN ('Needs Service','Under Repair') OR condition IN ('Needs Repair','Unserviceable') THEN 1 ELSE 0 END) AS needs_attention,
          SUM(CASE WHEN COALESCE(photo_filename,'')='' THEN 1 ELSE 0 END) AS missing_photos
        FROM equipment
    """).fetchone()

    equipment_due = conn.execute("""
        SELECT id,equipment_no,name,next_service_date,calibration_due,test_tag_due
        FROM equipment
        WHERE (next_service_date IS NOT NULL AND next_service_date!='' AND next_service_date<=?)
           OR (calibration_due IS NOT NULL AND calibration_due!='' AND calibration_due<=?)
           OR (test_tag_due IS NOT NULL AND test_tag_due!='' AND test_tag_due<=?)
        ORDER BY COALESCE(next_service_date,calibration_due,test_tag_due),equipment_no
        LIMIT 5
    """, (rego_90,rego_90,rego_90)).fetchall()

    # Version 22 - Business Intelligence KPIs
    business_kpis = {}
    business_kpis["customers"] = conn.execute("""
        SELECT COUNT(*) AS c FROM (
            SELECT LOWER(TRIM(COALESCE(buyer_email,buyer_phone,buyer_name,''))) AS customer_key FROM sales WHERE COALESCE(buyer_name,'')<>''
            UNION
            SELECT LOWER(TRIM(COALESCE(customer_email,customer_phone,customer_name,''))) FROM part_sales WHERE COALESCE(customer_name,'')<>''
        ) WHERE customer_key<>''
    """).fetchone()["c"]
    business_kpis["smart_reminders"] = conn.execute("SELECT COUNT(*) AS c FROM reminders WHERE completed=0 AND reminder_type='Smart Reminder'").fetchone()["c"]
    business_kpis["donor_vehicles"] = conn.execute("SELECT COUNT(*) AS c FROM vehicles WHERE COALESCE(vehicle_purpose,'Retail Sale')='Parts Vehicle' OR status='BER'").fetchone()["c"]
    business_kpis["inventory_retail"] = parts_retail_value
    business_kpis["parts_month"] = float(parts_sales_month["revenue"] or 0)
    business_kpis["vehicle_month"] = float(conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) AS v FROM sales WHERE substr(sale_date,1,7)=substr(?,1,7)", (today_text,)).fetchone()["v"] or 0)
    business_kpis["avg_stock_days"] = float(conn.execute("""
        SELECT COALESCE(AVG(julianday(?) - julianday(COALESCE(purchase_date,substr(created_at,1,10)))),0) AS v
        FROM vehicles WHERE status NOT IN ('Sold','BER')
    """, (today_text,)).fetchone()["v"] or 0)
    business_kpis["advertised"] = advertised_count

    email_metrics = conn.execute("""
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN status='Unread' THEN 1 ELSE 0 END) AS unread,
               SUM(CASE WHEN status='Follow Up' THEN 1 ELSE 0 END) AS follow_up
        FROM email_messages
    """).fetchone()
    recent_emails = conn.execute("""
        SELECT em.*, v.stock_no, c.name AS contact_name
        FROM email_messages em
        LEFT JOIN vehicles v ON v.id=em.vehicle_id
        LEFT JOIN contacts c ON c.id=em.contact_id
        ORDER BY COALESCE(em.message_date, substr(em.created_at,1,10)) DESC, em.id DESC
        LIMIT 5
    """).fetchall()

    conn.close()

    net_profit = sales_total - (metrics["purchase_total"] or 0) - expense_total - job_total - service_total - parts_total

    status_labels = [row["status"] for row in status_counts]
    status_values = [row["total"] for row in status_counts]
    month_labels = [row["month"] for row in monthly_rows]
    month_values = [row["sales_total"] for row in monthly_rows]

    return render_template(
        "dashboard.html",
        metrics=metrics,
        expense_total=expense_total,
        job_total=job_total,
        service_total=service_total,
        parts_total=parts_total,
        sales_total=sales_total,
        gst_paid=gst_paid,
        gst_collected=gst_collected,
        net_profit=net_profit,
        barry_total=barry_total,
        matt_total=matt_total,
        rego_alerts=rego_alerts,
        service_alerts=service_alerts,
        recent_vehicles=recent_vehicles,
        next_stock=next_stock,
        status_labels=status_labels,
        status_values=status_values,
        month_labels=month_labels,
        month_values=month_values,
        today=today.isoformat(),
        open_tasks=open_tasks,
        attention_vehicles=attention_vehicles,
        advertised_count=advertised_count,
        sold_this_month=sold_this_month,
        parts_metrics=parts_metrics,
        missing_photo_count=missing_photo_count,
        workshop_metrics=workshop_metrics,
        priority_tasks=priority_tasks,
        ready_to_advertise=ready_to_advertise,
        parts_sales_month=parts_sales_month,
        parts_retail_value=parts_retail_value,
        asset_summary=asset_summary,
        asset_totals=asset_totals,
        equipment_metrics=equipment_metrics,
        equipment_due=equipment_due,
        business_kpis=business_kpis,
        email_metrics=email_metrics,
        recent_emails=recent_emails,
    )

@app.route("/vehicles")
@login_required
def vehicle_list():
    q = request.args.get("q", "").strip()
    asset_type = request.args.get("asset_type", "").strip()
    conn = db()
    base_query = """
        SELECT v.*,
               COALESCE(
                 v.photo_filename,
                 (SELECT fp.filename FROM vehicle_photos fp WHERE fp.id=v.featured_photo_id LIMIT 1),
                 (SELECT p.filename FROM vehicle_photos p WHERE p.vehicle_id=v.id ORDER BY p.id DESC LIMIT 1)
               ) AS thumbnail
        FROM vehicles v
    """
    conditions = []
    params = []
    if q:
        conditions.append("(v.stock_no LIKE ? OR v.make LIKE ? OR v.model LIKE ? OR v.vin LIKE ? OR v.registration LIKE ? OR v.hin LIKE ?)")
        params.extend([f"%{q}%"] * 6)
    if asset_type:
        conditions.append("COALESCE(v.asset_type,'Car')=?")
        params.append(asset_type)
    if conditions:
        base_query += " WHERE " + " AND ".join(conditions)
    rows = conn.execute(base_query + " ORDER BY v.id DESC", tuple(params)).fetchall()
    type_counts = conn.execute("""
        SELECT COALESCE(asset_type,'Car') AS asset_type, COUNT(*) AS total
        FROM vehicles GROUP BY COALESCE(asset_type,'Car')
    """).fetchall()
    conn.close()
    return render_template("vehicles.html", vehicles=rows, q=q, asset_type=asset_type, type_counts=type_counts)

@app.route("/vehicles/new", methods=["GET", "POST"])
@login_required
def vehicle_new():
    conn = db()
    suggested_stock = next_stock_number(conn)
    conn.close()
    if request.method == "POST":
        try:
            photo = save_upload(request.files.get("photo"))
            price = float(request.form.get("purchase_price_inc_gst") or 0)
            gst = round(price / 11, 2)
            conn = db()
            stock_no = request.form.get("stock_no", "").strip() or next_stock_number(conn)
            duplicate_rows = duplicate_vehicle_matches(conn, request.form.get("vin"), request.form.get("registration"), stock_no)
            if duplicate_rows:
                hit = duplicate_rows[0]
                raise ValueError(f"Possible duplicate stock: {hit['stock_no']} {hit['year'] or ''} {hit['make']} {hit['model']} (VIN/registration/stock match).")
            fields = [
                "stock_no","status","purchase_date","make","model","variant","year","vin","registration",
                "odometer_km","colour","purchase_price_inc_gst","purchase_gst","barry_contribution",
                "matt_contribution","rego_expiry","photo_filename","notes","ppsr_number","roadworthy_status",
                "service_due_date","service_history","asset_type","length_m","tare_weight_kg","atm_kg","gtm_kg",
                "berths","axles","caravan_features","boat_type","hull_material","engine_make","engine_model",
                "engine_hours","horsepower","fuel_type","hin","trailer_included","trailer_registration",
                "capacity_people","boat_features","vehicle_purpose","dismantling_status"
            ]
            values = [
                stock_no, request.form.get("status") or "In Stock", request.form.get("purchase_date"),
                (request.form.get("make") or "").strip(), (request.form.get("model") or "").strip(),
                request.form.get("variant"), request.form.get("year") or None, request.form.get("vin") or None,
                request.form.get("registration"), request.form.get("odometer_km") or None, request.form.get("colour"),
                price, gst, float(request.form.get("barry_contribution") or 0),
                float(request.form.get("matt_contribution") or 0), request.form.get("rego_expiry"), photo,
                request.form.get("notes"), request.form.get("ppsr_number"),
                request.form.get("roadworthy_status") or "Not Checked", request.form.get("service_due_date"),
                request.form.get("service_history"), request.form.get("asset_type") or "Car",
                request.form.get("length_m") or None, request.form.get("tare_weight_kg") or None,
                request.form.get("atm_kg") or None, request.form.get("gtm_kg") or None,
                request.form.get("berths") or None, request.form.get("axles") or None,
                request.form.get("caravan_features"), request.form.get("boat_type"),
                request.form.get("hull_material"), request.form.get("engine_make"),
                request.form.get("engine_model"), request.form.get("engine_hours") or None,
                request.form.get("horsepower") or None, request.form.get("fuel_type"),
                request.form.get("hin"), 1 if request.form.get("trailer_included") else 0,
                request.form.get("trailer_registration"), request.form.get("capacity_people") or None,
                request.form.get("boat_features"),
                request.form.get("vehicle_purpose") or ("Parts Vehicle" if (request.form.get("status") == "BER") else "Retail Sale"),
                request.form.get("dismantling_status") or "Not Started"
            ]
            if not values[3] or not values[4]:
                raise ValueError("Make and model are required.")
            placeholders = ",".join(["?"] * len(fields))
            conn.execute(f"INSERT INTO vehicles({','.join(fields)}) VALUES({placeholders})", values)
            conn.commit()
            conn.close()
            flash(f"{request.form.get('asset_type') or 'Asset'} added.", "success")
            return redirect(url_for("vehicle_list"))
        except (sqlite3.IntegrityError, ValueError) as exc:
            flash(str(exc), "error")
    return render_template("vehicle_form.html", suggested_stock=suggested_stock)

@app.route("/vehicles/<int:vehicle_id>/edit", methods=["GET", "POST"])
@login_required
def vehicle_edit(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    if request.method == "POST":
        try:
            old_values = dict(vehicle)
            price = float(request.form.get("purchase_price_inc_gst") or 0)
            gst = round(price / 11, 2)
            new_photo = save_upload(request.files.get("photo"))
            photo_filename = new_photo or vehicle["photo_filename"]

            values = {
                "stock_no": (request.form.get("stock_no") or vehicle["stock_no"]).strip(),
                "status": request.form.get("status") or "In Stock",
                "purchase_date": request.form.get("purchase_date") or None,
                "make": (request.form.get("make") or "").strip(),
                "model": (request.form.get("model") or "").strip(),
                "variant": request.form.get("variant") or None,
                "year": request.form.get("year") or None,
                "vin": (request.form.get("vin") or "").strip() or None,
                "registration": (request.form.get("registration") or "").strip() or None,
                "odometer_km": request.form.get("odometer_km") or None,
                "colour": request.form.get("colour") or None,
                "purchase_price_inc_gst": price,
                "purchase_gst": gst,
                "asking_price": float(request.form.get("asking_price") or 0),
                "minimum_sale_price": float(request.form.get("minimum_sale_price") or 0),
                "negotiated_price": float(request.form.get("negotiated_price") or 0),
                "barry_contribution": float(request.form.get("barry_contribution") or 0),
                "matt_contribution": float(request.form.get("matt_contribution") or 0),
                "rego_expiry": request.form.get("rego_expiry") or None,
                "photo_filename": photo_filename,
                "notes": request.form.get("notes") or None,
                "ppsr_number": request.form.get("ppsr_number") or None,
                "roadworthy_status": request.form.get("roadworthy_status") or "Not Checked",
                "service_due_date": request.form.get("service_due_date") or None,
                "service_history": request.form.get("service_history") or None,
                "asset_type": request.form.get("asset_type") or "Car",
                "length_m": request.form.get("length_m") or None,
                "tare_weight_kg": request.form.get("tare_weight_kg") or None,
                "atm_kg": request.form.get("atm_kg") or None,
                "gtm_kg": request.form.get("gtm_kg") or None,
                "berths": request.form.get("berths") or None,
                "axles": request.form.get("axles") or None,
                "caravan_features": request.form.get("caravan_features") or None,
                "boat_type": request.form.get("boat_type") or None,
                "hull_material": request.form.get("hull_material") or None,
                "engine_make": request.form.get("engine_make") or None,
                "engine_model": request.form.get("engine_model") or None,
                "engine_hours": request.form.get("engine_hours") or None,
                "horsepower": request.form.get("horsepower") or None,
                "fuel_type": request.form.get("fuel_type") or None,
                "hin": request.form.get("hin") or None,
                "trailer_included": 1 if request.form.get("trailer_included") else 0,
                "trailer_registration": request.form.get("trailer_registration") or None,
                "capacity_people": request.form.get("capacity_people") or None,
                "boat_features": request.form.get("boat_features") or None,
                "vehicle_purpose": request.form.get("vehicle_purpose") or ("Parts Vehicle" if (request.form.get("status") == "BER") else "Retail Sale"),
                "dismantling_status": request.form.get("dismantling_status") or "Not Started",
                "shell_sale_price": float(request.form.get("shell_sale_price") or 0),
                "shell_sale_date": request.form.get("shell_sale_date") or None,
            }
            if not values["make"] or not values["model"]:
                raise ValueError("Make and model are required.")
            duplicate_rows = duplicate_vehicle_matches(conn, values["vin"], values["registration"], values["stock_no"], exclude_id=vehicle_id)
            if duplicate_rows:
                hit = duplicate_rows[0]
                raise ValueError(f"Possible duplicate stock: {hit['stock_no']} {hit['year'] or ''} {hit['make']} {hit['model']} (VIN/registration/stock match).")

            conn.execute("""
                UPDATE vehicles SET
                    stock_no=?, status=?, purchase_date=?, make=?, model=?, variant=?, year=?,
                    vin=?, registration=?, odometer_km=?, colour=?, purchase_price_inc_gst=?,
                    purchase_gst=?, asking_price=?, minimum_sale_price=?, negotiated_price=?,
                    barry_contribution=?, matt_contribution=?, rego_expiry=?,
                    photo_filename=?, notes=?, ppsr_number=?, roadworthy_status=?,
                    service_due_date=?, service_history=?, asset_type=?, length_m=?, tare_weight_kg=?,
                    atm_kg=?, gtm_kg=?, berths=?, axles=?, caravan_features=?, boat_type=?, hull_material=?,
                    engine_make=?, engine_model=?, engine_hours=?, horsepower=?, fuel_type=?, hin=?,
                    trailer_included=?, trailer_registration=?, capacity_people=?, boat_features=?,
                    vehicle_purpose=?, dismantling_status=?, shell_sale_price=?, shell_sale_date=?
                WHERE id=?
            """, (
                values["stock_no"], values["status"], values["purchase_date"],
                values["make"], values["model"], values["variant"], values["year"],
                values["vin"], values["registration"], values["odometer_km"],
                values["colour"], values["purchase_price_inc_gst"], values["purchase_gst"],
                values["asking_price"], values["minimum_sale_price"], values["negotiated_price"],
                values["barry_contribution"], values["matt_contribution"],
                values["rego_expiry"], values["photo_filename"], values["notes"],
                values["ppsr_number"], values["roadworthy_status"],
                values["service_due_date"], values["service_history"], values["asset_type"],
                values["length_m"], values["tare_weight_kg"], values["atm_kg"], values["gtm_kg"],
                values["berths"], values["axles"], values["caravan_features"], values["boat_type"],
                values["hull_material"], values["engine_make"], values["engine_model"],
                values["engine_hours"], values["horsepower"], values["fuel_type"], values["hin"],
                values["trailer_included"], values["trailer_registration"], values["capacity_people"],
                values["boat_features"], values["vehicle_purpose"], values["dismantling_status"],
                values["shell_sale_price"], values["shell_sale_date"], vehicle_id,
            ))
            if values["asking_price"] > 0 and values["minimum_sale_price"] > 0:
                conn.execute("""
                    UPDATE tasks
                    SET status='Completed'
                    WHERE vehicle_id=?
                      AND title='Confirm asking price and minimum sale price'
                      AND status!='Completed'
                """, (vehicle_id,))
            conn.commit()
            conn.close()

            changed = []
            watched = [
                "stock_no", "status", "purchase_date", "make", "model", "variant",
                "year", "vin", "registration", "odometer_km", "colour",
                "purchase_price_inc_gst", "asking_price", "minimum_sale_price", "negotiated_price",
                "barry_contribution", "matt_contribution",
                "rego_expiry", "ppsr_number", "roadworthy_status",
                "service_due_date", "service_history", "asset_type", "length_m", "tare_weight_kg",
                "atm_kg", "gtm_kg", "berths", "axles", "caravan_features", "boat_type",
                "hull_material", "engine_make", "engine_model", "engine_hours", "horsepower",
                "fuel_type", "hin", "trailer_included", "trailer_registration", "capacity_people",
                "boat_features", "vehicle_purpose", "dismantling_status", "shell_sale_price",
                "shell_sale_date", "notes",
            ]
            for field in watched:
                old = old_values.get(field)
                new = values.get(field)
                if str(old or "") != str(new or ""):
                    changed.append(f"{field}: {old or ''} -> {new or ''}")
            if new_photo:
                changed.append("main photo replaced")
            log_action(
                "Vehicle updated",
                "vehicle",
                vehicle_id,
                "; ".join(changed) if changed else "Vehicle saved with no field changes",
            )
            flash("Vehicle updated.", "success")
            if request.form.get("continue_editing"):
                return redirect(url_for("vehicle_edit", vehicle_id=vehicle_id))
            return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))
        except (sqlite3.IntegrityError, ValueError) as exc:
            conn.close()
            flash(str(exc), "error")
            vehicle = {**dict(vehicle), **request.form.to_dict()}

    else:
        conn.close()

    return render_template("vehicle_edit.html", vehicle=vehicle)

@app.errorhandler(413)
def upload_too_large(_error):
    flash("The selected files are too large. Upload fewer photos at a time or use smaller files.", "error")
    return redirect(request.referrer or url_for("vehicle_list"))

@app.route("/api/vin-decode")
@login_required
def api_vin_decode():
    specs, error = decode_vin(request.args.get("vin", ""))
    return jsonify({"specs": specs, "error": error})


@app.route("/api/vehicle-duplicate-check")
@login_required
def vehicle_duplicate_check():
    conn = db()
    rows = duplicate_vehicle_matches(conn, request.args.get("vin"), request.args.get("registration"), request.args.get("stock_no"), request.args.get("exclude_id", type=int))
    conn.close()
    return jsonify({"matches": [dict(r) for r in rows]})


@app.route("/vehicles/<int:vehicle_id>/intelligence")
@login_required
def dealer_intelligence(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404
    market = market_price_suggestion(conn, vehicle)
    duplicates = duplicate_vehicle_matches(conn, vehicle["vin"], vehicle["registration"], vehicle["stock_no"], exclude_id=vehicle_id)
    comparable_sales = conn.execute("""
        SELECT s.sale_date,s.sale_price_inc_gst,v.stock_no,v.year,v.make,v.model
        FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
        WHERE s.sale_price_inc_gst>0 AND LOWER(v.make)=LOWER(?)
        ORDER BY s.sale_date DESC LIMIT 8
    """, (vehicle["make"],)).fetchall()
    conn.close()
    return render_template("dealer_intelligence.html", vehicle=vehicle, market=market, duplicates=duplicates,
                           comparable_sales=comparable_sales, ai_enabled=bool(OPENAI_API_KEY))


@app.route("/vehicles/<int:vehicle_id>/decode-vin", methods=["POST"])
@login_required
def vehicle_decode_vin(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404
    specs, error = decode_vin(vehicle["vin"])
    if specs:
        conn.execute("""
            UPDATE vehicles SET vin_decode_json=?,decoded_at=CURRENT_TIMESTAMP,manufacturer_name=?,body_class=?,drive_type=?,
                transmission_style=?,engine_cylinders=?,engine_displacement_l=?,fuel_type_primary=?,decoded_model_year=?,decoded_series=?,decoded_trim=?,
                make=CASE WHEN COALESCE(make,'')='' THEN ? ELSE make END,
                model=CASE WHEN COALESCE(model,'')='' THEN ? ELSE model END,
                year=CASE WHEN year IS NULL OR year=0 THEN ? ELSE year END
            WHERE id=?
        """, (json.dumps(specs.get("raw", {})), specs.get("manufacturer_name"), specs.get("body_class"), specs.get("drive_type"),
              specs.get("transmission_style"), specs.get("engine_cylinders"), specs.get("engine_displacement_l"), specs.get("fuel_type_primary"),
              specs.get("year"), specs.get("series"), specs.get("trim"), specs.get("make"), specs.get("model"), specs.get("year") or None, vehicle_id))
        conn.commit()
        log_action("VIN decoded", "vehicle", vehicle_id, f"VIN {vehicle['vin']}; provider specs saved")
        flash("VIN decoded and available specifications were saved." + (f" Note: {error}" if error else ""), "success")
    else:
        flash(error or "VIN could not be decoded.", "error")
    conn.close()
    return redirect(url_for("dealer_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/market-suggestion", methods=["POST"])
@login_required
def vehicle_market_suggestion_save(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found", 404
    market = market_price_suggestion(conn, vehicle)
    conn.execute("UPDATE vehicles SET market_price_low=?,market_price_mid=?,market_price_high=?,market_price_checked_at=CURRENT_TIMESTAMP WHERE id=?",
                 (market["low"], market["mid"], market["high"], vehicle_id))
    conn.commit(); conn.close()
    log_action("Market pricing suggestion refreshed", "vehicle", vehicle_id, json.dumps(market))
    flash("Pricing suggestion refreshed from BAM's recorded sales and vehicle cost data.", "success")
    return redirect(url_for("dealer_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/generate-ai-ad", methods=["POST"])
@login_required
def vehicle_generate_ai_ad(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found", 404
    text, source = generate_vehicle_ad_text(vehicle)
    title = f"{vehicle['year'] or ''} {vehicle['make']} {vehicle['model']} {vehicle['variant'] or ''}".strip()
    conn.execute("UPDATE vehicles SET advertisement_title=?,advertisement_description=? WHERE id=?", (title, text, vehicle_id))
    conn.commit(); conn.close()
    log_action("Advertisement generated", "vehicle", vehicle_id, source)
    flash(f"Advertisement generated. {source}", "success")
    return redirect(url_for("advertisement_pro", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>")
@login_required
def vehicle_detail(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Asset not found", 404

    expenses = conn.execute(
        "SELECT * FROM expenses WHERE vehicle_id=? ORDER BY expense_date DESC,id DESC",
        (vehicle_id,),
    ).fetchall()
    job_cards = conn.execute(
        "SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY job_date DESC,id DESC",
        (vehicle_id,),
    ).fetchall()
    photos = conn.execute(
        """SELECT * FROM vehicle_photos
           WHERE vehicle_id=?
           ORDER BY CASE WHEN id=(SELECT featured_photo_id FROM vehicles WHERE id=?) THEN 0 ELSE 1 END,
                    id DESC""",
        (vehicle_id, vehicle_id),
    ).fetchall()
    documents = conn.execute(
        "SELECT * FROM vehicle_documents WHERE vehicle_id=? ORDER BY uploaded_at DESC,id DESC",
        (vehicle_id,),
    ).fetchall()
    reminders = conn.execute(
        "SELECT * FROM reminders WHERE vehicle_id=? ORDER BY completed,reminder_date,id DESC",
        (vehicle_id,),
    ).fetchall()
    services = conn.execute(
        "SELECT * FROM service_entries WHERE vehicle_id=? ORDER BY service_date DESC,id DESC",
        (vehicle_id,),
    ).fetchall()
    parts_used = conn.execute(
        """SELECT u.*,p.part_number,p.part_name
           FROM part_usage u JOIN parts p ON p.id=u.part_id
           WHERE u.vehicle_id=? ORDER BY u.usage_date DESC,u.id DESC""",
        (vehicle_id,),
    ).fetchall()
    available_parts = conn.execute(
        "SELECT * FROM parts WHERE quantity_on_hand>0 ORDER BY part_name",
    ).fetchall()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()

    expense_total = sum(float(row["cost_inc_gst"] or 0) for row in expenses)
    job_total = sum(
        float(row["actual_cost_inc_gst"] or 0)
        if float(row["actual_cost_inc_gst"] or 0) > 0
        else float(row["estimated_cost"] or 0)
        for row in job_cards
    )
    service_total = sum(float(row["cost_inc_gst"] or 0) for row in services)
    parts_total = sum(float(row["quantity_used"] or 0) * float(row["unit_cost_inc_gst"] or 0) for row in parts_used)
    sale_price = float(sale["sale_price_inc_gst"] or 0) if sale else 0
    selling_costs = float((sale["advertising_cost"] or 0) + (sale["transfer_cost"] or 0)) if sale else 0
    total_invested = float(vehicle["purchase_price_inc_gst"] or 0) + expense_total + job_total + service_total + parts_total + selling_costs

    # Version 21 - every dismantled-part sale feeds back to the donor vehicle.
    donor_parts = conn.execute(
        "SELECT * FROM parts WHERE vehicle_id=? ORDER BY id DESC", (vehicle_id,)
    ).fetchall()
    donor_metrics = conn.execute("""
        SELECT COUNT(*) AS part_lines,
               COALESCE(SUM(quantity_on_hand),0) AS units_on_hand,
               COALESCE(SUM(quantity_on_hand * selling_price),0) AS remaining_retail
        FROM parts WHERE vehicle_id=?
    """, (vehicle_id,)).fetchone()
    donor_sales = conn.execute("""
        SELECT COALESCE(SUM(ps.sale_price),0) AS revenue,
               COALESCE(SUM(ps.quantity),0) AS units_sold,
               COUNT(ps.id) AS sale_count
        FROM part_sales ps JOIN parts p ON p.id=ps.part_id
        WHERE p.vehicle_id=?
    """, (vehicle_id,)).fetchone()
    donor_revenue = float(donor_sales["revenue"] or 0)
    shell_revenue = float(vehicle["shell_sale_price"] or 0)
    total_revenue = sale_price + donor_revenue + shell_revenue
    profit = total_revenue - total_invested
    donor_profit = donor_revenue + shell_revenue - total_invested
    recovery_pct = (total_revenue / total_invested * 100) if total_invested > 0 else (100.0 if total_revenue > 0 else 0.0)
    remaining_to_break_even = max(0.0, total_invested - total_revenue)
    break_even_status = "Recovered / Profitable" if total_revenue >= total_invested and total_invested > 0 else "Recovering Investment"
    conn.close()

    return render_template(
        "vehicle_detail.html",
        vehicle=vehicle,
        expenses=expenses,
        job_cards=job_cards,
        photos=photos,
        documents=documents,
        reminders=reminders,
        services=services,
        parts_used=parts_used,
        available_parts=available_parts,
        sale=sale,
        expense_total=expense_total,
        job_total=job_total,
        service_total=service_total,
        parts_total=parts_total,
        total_invested=total_invested,
        profit=profit,
        donor_parts=donor_parts,
        donor_metrics=donor_metrics,
        donor_revenue=donor_revenue,
        donor_units_sold=float(donor_sales["units_sold"] or 0),
        donor_sale_count=int(donor_sales["sale_count"] or 0),
        donor_profit=donor_profit,
        total_revenue=total_revenue,
        recovery_pct=recovery_pct,
        remaining_to_break_even=remaining_to_break_even,
        break_even_status=break_even_status,
    )

@app.route("/vehicles/<int:vehicle_id>/documents", methods=["POST"])
@login_required
def vehicle_document_add(vehicle_id):
    upload = request.files.get("document")
    try:
        filename = save_upload(upload)
        if not filename:
            raise ValueError("Choose a document to upload.")
        conn = db()
        conn.execute(
            """INSERT INTO vehicle_documents(vehicle_id,document_type,filename,description)
               VALUES(?,?,?,?)""",
            (
                vehicle_id,
                request.form.get("document_type") or "Other",
                filename,
                request.form.get("description"),
            ),
        )
        conn.commit()
        conn.close()
        log_action("Document uploaded", "vehicle", vehicle_id, request.form.get("document_type") or "Other")
        flash("Document uploaded.", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id) + "#documents")

@app.route("/vehicles/<int:vehicle_id>/documents/<int:document_id>/delete", methods=["POST"])
@login_required
def vehicle_document_delete(vehicle_id, document_id):
    conn = db()
    document = conn.execute(
        "SELECT * FROM vehicle_documents WHERE id=? AND vehicle_id=?",
        (document_id, vehicle_id),
    ).fetchone()
    if document:
        conn.execute("DELETE FROM vehicle_documents WHERE id=?", (document_id,))
        conn.commit()
    conn.close()
    if document:
        try:
            (UPLOAD_DIR / document["filename"]).unlink(missing_ok=True)
        except OSError:
            pass
        log_action("Document deleted", "vehicle", vehicle_id, document["document_type"])
        flash("Document deleted.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id) + "#documents")

@app.route("/vehicles/<int:vehicle_id>/expense", methods=["POST"])
@login_required
def expense_add(vehicle_id):
    try:
        receipt = save_upload(request.files.get("receipt"))
        cost = float(request.form.get("cost_inc_gst") or 0)
        gst = round(cost / 11, 2)
        conn = db()
        conn.execute("""
            INSERT INTO expenses(vehicle_id,expense_date,category,description,supplier,paid_by,
                                 cost_inc_gst,gst_amount,receipt_filename,notes)
            VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (vehicle_id, request.form.get("expense_date"), request.form["category"],
              request.form["description"], request.form.get("supplier"),
              request.form["paid_by"], cost, gst, receipt, request.form.get("notes")))
        conn.commit()
        conn.close()
        flash("Expense added.", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

@app.route("/vehicles/<int:vehicle_id>/photos", methods=["POST"])
@login_required
def photo_add(vehicle_id):
    files = request.files.getlist("photos")
    caption = request.form.get("caption")
    saved = 0
    conn = db()
    try:
        for file in files:
            filename = save_upload(file)
            if filename:
                conn.execute(
                    "INSERT INTO vehicle_photos(vehicle_id,filename,caption) VALUES(?,?,?)",
                    (vehicle_id, filename, caption),
                )
                saved += 1
        conn.commit()
        flash(f"{saved} photo(s) added.", "success")
    except ValueError as exc:
        flash(str(exc), "error")
    finally:
        conn.close()
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/job-card", methods=["POST"])
@login_required
def job_card_add(vehicle_id):
    actual = float(request.form.get("actual_cost_inc_gst") or 0)
    gst = round(actual / 11, 2)
    conn = db()
    cursor = conn.execute("""
        INSERT INTO job_cards(
            vehicle_id,job_date,category,description,supplier,paid_by,
            estimated_cost,actual_cost_inc_gst,gst_amount,status,notes,
            labour_operation_id,labour_hours,labour_rate,labour_source,labour_code,procedure_url
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        vehicle_id,
        request.form.get("job_date"),
        request.form.get("category"),
        request.form.get("description"),
        request.form.get("supplier"),
        request.form.get("paid_by"),
        float(request.form.get("estimated_cost") or 0),
        actual,
        gst,
        request.form.get("status") or "Open",
        request.form.get("notes"),
        int(request.form.get("labour_operation_id")) if request.form.get("labour_operation_id") else None,
        float(request.form.get("labour_hours") or 0),
        float(request.form.get("labour_rate") or 0),
        request.form.get("labour_source"),
        request.form.get("labour_code"),
        request.form.get("procedure_url"),
    ))
    job_id = cursor.lastrowid
    conn.execute("""
        INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
        VALUES(?,?,?,?,?)
    """, (
        job_id, None, request.form.get("status") or "Open",
        session.get("display_name"), "Job card created"
    ))
    conn.commit()
    conn.close()
    flash("Job card added.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))



@app.route("/job-cards/<int:job_id>/edit", methods=["GET", "POST"])
@login_required
def job_card_edit(job_id):
    conn = db()
    job = conn.execute("SELECT * FROM job_cards WHERE id=?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return "Job card not found", 404

    if request.method == "POST":
        old_status = job["status"]
        new_status = request.form.get("status") or old_status
        actual = float(request.form.get("actual_cost_inc_gst") or 0)
        gst = round(actual / 11, 2)

        conn.execute("""
            UPDATE job_cards
            SET job_date=?, category=?, description=?, supplier=?, paid_by=?,
                estimated_cost=?, actual_cost_inc_gst=?, gst_amount=?, status=?, notes=?,
                labour_hours=?,labour_rate=?,labour_source=?,labour_code=?,procedure_url=?
            WHERE id=?
        """, (
            request.form.get("job_date"),
            request.form.get("category"),
            request.form.get("description"),
            request.form.get("supplier"),
            request.form.get("paid_by"),
            float(request.form.get("estimated_cost") or 0),
            actual,
            gst,
            new_status,
            request.form.get("notes"),
            float(request.form.get("labour_hours") or 0),
            float(request.form.get("labour_rate") or 0),
            request.form.get("labour_source"),
            request.form.get("labour_code"),
            request.form.get("procedure_url"),
            job_id,
        ))

        change_note = request.form.get("change_note") or "Job card updated"
        conn.execute("""
            INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
            VALUES(?,?,?,?,?)
        """, (
            job_id, old_status, new_status,
            session.get("display_name"), change_note
        ))
        conn.commit()
        vehicle_id = job["vehicle_id"]
        conn.close()
        flash("Job card updated.", "success")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    history = conn.execute(
        "SELECT * FROM job_card_history WHERE job_card_id=? ORDER BY changed_at DESC,id DESC",
        (job_id,)
    ).fetchall()
    conn.close()
    return render_template("job_card_edit.html", job=job, history=history)


@app.route("/job-cards/<int:job_id>/reopen", methods=["POST"])
@login_required
def job_card_reopen(job_id):
    conn = db()
    job = conn.execute("SELECT * FROM job_cards WHERE id=?", (job_id,)).fetchone()
    if not job:
        conn.close()
        return "Job card not found", 404

    old_status = job["status"]
    new_status = request.form.get("status") or "Open"
    note = request.form.get("note") or "Job card reopened"

    conn.execute("UPDATE job_cards SET status=? WHERE id=?", (new_status, job_id))
    conn.execute("""
        INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note)
        VALUES(?,?,?,?,?)
    """, (
        job_id, old_status, new_status,
        session.get("display_name"), note
    ))
    conn.commit()
    vehicle_id = job["vehicle_id"]
    conn.close()
    flash(f"Job card reopened as {new_status}.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/sale", methods=["POST"])
@login_required
def sale_add(vehicle_id):
    price = float(request.form.get("sale_price_inc_gst") or 0)
    gst = round(price / 11, 2)
    conn = db()
    existing = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    invoice_number = existing["invoice_number"] if existing and existing["invoice_number"] else next_invoice_number(conn)
    contract_number = existing["contract_number"] if existing and existing["contract_number"] else next_contract_number(conn)

    conn.execute("""
        INSERT INTO sales(
            vehicle_id,sale_date,buyer_name,buyer_phone,buyer_email,buyer_address,
            sale_price_inc_gst,sale_gst,advertising_cost,transfer_cost,invoice_number,
            deposit_amount,deposit_date,payment_method,trade_in_description,trade_in_value,
            warranty_type,warranty_expiry,contract_number,notes
        )
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(vehicle_id) DO UPDATE SET
          sale_date=excluded.sale_date,
          buyer_name=excluded.buyer_name,
          buyer_phone=excluded.buyer_phone,
          buyer_email=excluded.buyer_email,
          buyer_address=excluded.buyer_address,
          sale_price_inc_gst=excluded.sale_price_inc_gst,
          sale_gst=excluded.sale_gst,
          advertising_cost=excluded.advertising_cost,
          transfer_cost=excluded.transfer_cost,
          invoice_number=COALESCE(sales.invoice_number,excluded.invoice_number),
          deposit_amount=excluded.deposit_amount,
          deposit_date=excluded.deposit_date,
          payment_method=excluded.payment_method,
          trade_in_description=excluded.trade_in_description,
          trade_in_value=excluded.trade_in_value,
          warranty_type=excluded.warranty_type,
          warranty_expiry=excluded.warranty_expiry,
          contract_number=COALESCE(sales.contract_number,excluded.contract_number),
          notes=excluded.notes
    """, (
        vehicle_id,
        request.form.get("sale_date"),
        request.form.get("buyer_name"),
        request.form.get("buyer_phone"),
        request.form.get("buyer_email"),
        request.form.get("buyer_address"),
        price,
        gst,
        float(request.form.get("advertising_cost") or 0),
        float(request.form.get("transfer_cost") or 0),
        invoice_number,
        float(request.form.get("deposit_amount") or 0),
        request.form.get("deposit_date"),
        request.form.get("payment_method"),
        request.form.get("trade_in_description"),
        float(request.form.get("trade_in_value") or 0),
        request.form.get("warranty_type"),
        request.form.get("warranty_expiry"),
        contract_number,
        request.form.get("notes"),
    ))
    conn.execute("UPDATE vehicles SET status='Sold' WHERE id=?", (vehicle_id,))
    conn.commit()
    conn.close()
    log_action("Sale recorded", "vehicle", vehicle_id, f"{invoice_number} / {contract_number}")
    flash(f"Sale recorded. Invoice {invoice_number} and contract {contract_number} created.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/import-excel", methods=["GET", "POST"])
@login_required
def import_excel():
    result = None
    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        if not uploaded or not uploaded.filename.lower().endswith(".xlsx"):
            flash("Please choose an .xlsx Excel workbook.", "error")
            return redirect(url_for("import_excel"))

        temp_path = BASE_DIR / f"import_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(uploaded.filename)}"
        uploaded.save(temp_path)
        imported = 0
        skipped = 0
        warnings = []
        workbook = None

        try:
            workbook = load_workbook(temp_path, data_only=True)
            sheet = None
            for candidate in ["Vehicles", "Vehicle Inventory"]:
                if candidate in workbook.sheetnames:
                    sheet = workbook[candidate]
                    break
            if sheet is None:
                raise ValueError("The workbook needs a sheet named Vehicles or Vehicle Inventory.")

            headers = {}
            header_row = None
            for row_number in range(1, min(sheet.max_row, 15) + 1):
                values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
                if any(str(v).strip() in ("Stock No.", "Stock Number") for v in values if v is not None):
                    header_row = row_number
                    headers = {
                        str(sheet.cell(row_number, col).value).strip(): col
                        for col in range(1, sheet.max_column + 1)
                        if sheet.cell(row_number, col).value is not None
                    }
                    break
            if not header_row:
                raise ValueError("Could not find the vehicle heading row.")

            def value(row, *names):
                for name in names:
                    col = headers.get(name)
                    if col:
                        return sheet.cell(row, col).value
                return None

            conn = db()
            for row in range(header_row + 1, sheet.max_row + 1):
                make = str(value(row, "Make") or "").strip()
                model = str(value(row, "Model") or "").strip()
                if not make and not model:
                    continue

                stock = str(value(row, "Stock No.", "Stock Number") or "").strip()
                if not stock:
                    stock = next_stock_number(conn)

                existing = conn.execute(
                    "SELECT id FROM vehicles WHERE stock_no=?",
                    (stock,)
                ).fetchone()
                if existing:
                    skipped += 1
                    continue

                vin = str(value(row, "VIN") or "").strip() or None
                if vin and conn.execute("SELECT id FROM vehicles WHERE vin=?", (vin,)).fetchone():
                    warnings.append(f"{stock}: duplicate VIN skipped")
                    skipped += 1
                    continue

                price = clean_number(value(row, "Purchase Price Inc GST", "Purchase Price"))
                gst = clean_number(value(row, "Purchase GST"))
                if not gst and price:
                    gst = round(price / 11, 2)

                purchase_date = value(row, "Purchase Date")
                rego_expiry = value(row, "Registration Expiry", "Rego Expiry")
                for_date = lambda x: x.date().isoformat() if hasattr(x, "date") else (str(x) if x else None)

                conn.execute("""
                    INSERT INTO vehicles(
                        stock_no,status,purchase_date,make,model,variant,year,vin,registration,
                        odometer_km,colour,purchase_price_inc_gst,purchase_gst,
                        barry_contribution,matt_contribution,rego_expiry,photo_filename,notes
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    stock,
                    value(row, "Status") or "In Stock",
                    for_date(purchase_date),
                    make,
                    model,
                    value(row, "Variant"),
                    int(clean_number(value(row, "Year"))) or None,
                    vin,
                    value(row, "Registration"),
                    int(clean_number(value(row, "Odometer km", "Odometer"))) or None,
                    value(row, "Colour"),
                    price,
                    gst,
                    clean_number(value(row, "Barry Contribution", "Barry Purchase Contribution")),
                    clean_number(value(row, "Matt Contribution", "Matt Purchase Contribution")),
                    for_date(rego_expiry),
                    value(row, "Main Photo Link"),
                    value(row, "Notes"),
                ))
                imported += 1

            conn.commit()
            conn.close()
            result = {"imported": imported, "skipped": skipped, "warnings": warnings}
            flash(f"Import completed: {imported} vehicle(s) added, {skipped} skipped.", "success")
        except Exception as exc:
            flash(f"Import failed: {exc}", "error")
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except PermissionError:
                    # Windows may briefly retain the file handle. The next import
                    # uses a new temporary name, so leaving this temp copy is safe.
                    pass

    return render_template("import_excel.html", result=result)


@app.route("/vehicles/<int:vehicle_id>/report")
@login_required
def vehicle_report(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    expenses = conn.execute("SELECT * FROM expenses WHERE vehicle_id=? ORDER BY expense_date,id", (vehicle_id,)).fetchall()
    jobs = conn.execute("SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY job_date,id", (vehicle_id,)).fetchall()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404
    expense_total = sum(x["cost_inc_gst"] for x in expenses)
    job_total = sum((x["actual_cost_inc_gst"] if x["actual_cost_inc_gst"] > 0 else x["estimated_cost"]) for x in jobs)
    sale_price = sale["sale_price_inc_gst"] if sale else 0
    selling_costs = (sale["advertising_cost"] + sale["transfer_cost"]) if sale else 0
    profit = sale_price - vehicle["purchase_price_inc_gst"] - expense_total - job_total - selling_costs
    return render_template("vehicle_report.html", vehicle=vehicle, expenses=expenses, jobs=jobs,
                           sale=sale, expense_total=expense_total, job_total=job_total, profit=profit)

@app.route("/backup")
@login_required
def backup_database():
    return redirect(url_for("backup_centre"))


def create_full_backup(prefix="bam_full_backup"):
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{prefix}_{stamp}.zip"
    temp_db = BACKUP_DIR / f".snapshot_{stamp}.db"
    source = sqlite3.connect(DB_PATH)
    target = sqlite3.connect(temp_db)
    with target:
        source.backup(target)
    source.close(); target.close()
    with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(temp_db, "database/bam_motor_group.db")
        for folder_name in ("uploads", "reports"):
            folder = PROJECT_DIR / folder_name
            if folder.exists():
                for file in folder.rglob("*"):
                    if file.is_file():
                        zf.write(file, file.relative_to(PROJECT_DIR))
        zf.writestr("BACKUP_INFO.txt", f"{APP_NAME} v{APP_VERSION}\nCreated: {datetime.now().isoformat()}\n")
    temp_db.unlink(missing_ok=True)
    return backup_path


@app.route("/backups")
@login_required
def backup_centre():
    backups = sorted(BACKUP_DIR.glob("*.zip"), key=lambda x: x.stat().st_mtime, reverse=True)
    rows = [{"name": b.name, "size": b.stat().st_size, "modified": datetime.fromtimestamp(b.stat().st_mtime)} for b in backups]
    return render_template("backup_centre.html", backups=rows)


@app.route("/backups/create", methods=["POST"])
@login_required
def backup_create():
    path = create_full_backup()
    log_action("Full backup created", "system", None, path.name)
    return send_file(path, as_attachment=True, download_name=path.name)


@app.route("/backups/download/<path:filename>")
@login_required
def backup_download(filename):
    safe = Path(filename).name
    path = BACKUP_DIR / safe
    if not path.exists() or path.suffix.lower() != ".zip":
        return "Backup not found", 404
    return send_file(path, as_attachment=True, download_name=path.name)


@app.route("/backups/upload", methods=["POST"])
@owner_required
def backup_upload():
    file = request.files.get("backup_file")
    if not file or not file.filename or not file.filename.lower().endswith(".zip"):
        flash("Choose a BAM ZIP backup.", "error")
        return redirect(url_for("backup_centre"))
    name = secure_filename(file.filename)
    destination = BACKUP_DIR / name
    file.save(destination)
    try:
        with zipfile.ZipFile(destination) as zf:
            if "database/bam_motor_group.db" not in zf.namelist():
                raise ValueError("This is not a valid BAM full backup.")
    except Exception as exc:
        destination.unlink(missing_ok=True)
        flash(f"Backup upload failed: {exc}", "error")
        return redirect(url_for("backup_centre"))
    flash("Backup uploaded and ready to restore.", "success")
    return redirect(url_for("backup_centre"))


@app.route("/backups/restore", methods=["POST"])
@owner_required
def backup_restore():
    filename = Path(request.form.get("filename") or "").name
    confirmation = (request.form.get("confirmation") or "").strip().upper()
    if confirmation != "RESTORE":
        flash("Type RESTORE to confirm.", "error")
        return redirect(url_for("backup_centre"))
    path = BACKUP_DIR / filename
    if not path.exists() or path.suffix.lower() != ".zip":
        flash("Backup file not found.", "error")
        return redirect(url_for("backup_centre"))
    create_full_backup("pre_restore_backup")
    restore_dir = BACKUP_DIR / ".restore_temp"
    if restore_dir.exists(): shutil.rmtree(restore_dir)
    restore_dir.mkdir()
    try:
        with zipfile.ZipFile(path) as zf:
            for member in zf.infolist():
                target = (restore_dir / member.filename).resolve()
                if restore_dir.resolve() not in target.parents and target != restore_dir.resolve():
                    raise ValueError("Unsafe backup archive.")
            zf.extractall(restore_dir)
        restored_db = restore_dir / "database" / "bam_motor_group.db"
        if not restored_db.exists(): raise ValueError("The backup does not contain a database.")
        shutil.copy2(restored_db, DB_PATH)
        for folder_name in ("uploads", "reports"):
            source_folder = restore_dir / folder_name
            if source_folder.exists():
                destination = PROJECT_DIR / folder_name
                destination.mkdir(exist_ok=True)
                for file in source_folder.rglob("*"):
                    if file.is_file():
                        out = destination / file.relative_to(source_folder)
                        out.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(file, out)
        init_db()
        session.clear()
        flash("Backup restored successfully. Please sign in again.", "success")
        return redirect(url_for("login"))
    except Exception as exc:
        flash(f"Restore failed: {exc}", "error")
        return redirect(url_for("backup_centre"))
    finally:
        if restore_dir.exists(): shutil.rmtree(restore_dir)


@app.route("/contacts", methods=["GET", "POST"])
@login_required
def contacts_page():
    conn = db()
    if request.method == "POST":
        conn.execute("""
            INSERT INTO contacts(contact_type,name,phone,email,address,licence_no,notes)
            VALUES(?,?,?,?,?,?,?)
        """, (
            request.form.get("contact_type"),
            request.form.get("name"),
            request.form.get("phone"),
            request.form.get("email"),
            request.form.get("address"),
            request.form.get("licence_no"),
            request.form.get("notes"),
        ))
        conn.commit()
        flash("Contact saved.", "success")
        return redirect(url_for("contacts_page"))

    q = request.args.get("q", "").strip()
    if q:
        rows = conn.execute("""
            SELECT * FROM contacts
            WHERE name LIKE ? OR phone LIKE ? OR email LIKE ? OR contact_type LIKE ?
            ORDER BY name
        """, tuple([f"%{q}%"] * 4)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM contacts ORDER BY name").fetchall()
    conn.close()
    return render_template("contacts.html", contacts=rows, q=q)


@app.route("/reports/profit")
@login_required
def profit_report():
    conn = db()
    rows = conn.execute("""
        SELECT v.id,v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price,
               COALESCE((SELECT s.advertising_cost+s.transfer_cost FROM sales s WHERE s.vehicle_id=v.id),0) AS selling_costs
        FROM vehicles v
        ORDER BY v.id DESC
    """).fetchall()
    conn.close()
    report_rows = []
    for row in rows:
        total_cost = row["purchase_price_inc_gst"] + row["expenses"] + row["jobs"] + row["selling_costs"]
        profit = row["sale_price"] - total_cost
        report_rows.append(dict(row) | {"total_cost": total_cost, "profit": profit})
    totals = {
        "purchase": sum(r["purchase_price_inc_gst"] for r in report_rows),
        "expenses": sum(r["expenses"] for r in report_rows),
        "jobs": sum(r["jobs"] for r in report_rows),
        "sales": sum(r["sale_price"] for r in report_rows),
        "profit": sum(r["profit"] for r in report_rows),
    }
    return render_template("profit_report.html", rows=report_rows, totals=totals)


@app.route("/reports/dismantling-profit")
@login_required
def dismantling_profit_report():
    conn = db()
    rows = conn.execute("""
        SELECT v.id,v.stock_no,v.year,v.make,v.model,v.status,v.dismantling_status,
               v.purchase_price_inc_gst,v.shell_sale_price,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT SUM(se.cost_inc_gst) FROM service_entries se WHERE se.vehicle_id=v.id),0) AS services,
               COALESCE((SELECT SUM(pu.quantity_used*pu.unit_cost_inc_gst) FROM part_usage pu WHERE pu.vehicle_id=v.id),0) AS parts_used,
               COALESCE((SELECT SUM(ps.sale_price) FROM part_sales ps JOIN parts p ON p.id=ps.part_id WHERE p.vehicle_id=v.id),0) AS parts_revenue,
               COALESCE((SELECT SUM(ps.quantity) FROM part_sales ps JOIN parts p ON p.id=ps.part_id WHERE p.vehicle_id=v.id),0) AS units_sold,
               COALESCE((SELECT SUM(p.quantity_on_hand*p.selling_price) FROM parts p WHERE p.vehicle_id=v.id),0) AS remaining_retail,
               COALESCE((SELECT COUNT(*) FROM parts p WHERE p.vehicle_id=v.id),0) AS part_lines,
               COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS vehicle_sale
        FROM vehicles v
        WHERE COALESCE(v.vehicle_purpose,'Retail Sale')='Parts Vehicle' OR v.status='BER'
        ORDER BY v.id DESC
    """).fetchall()
    conn.close()
    report_rows = []
    for row in rows:
        item = dict(row)
        invested = sum(float(item[k] or 0) for k in ["purchase_price_inc_gst","expenses","jobs","services","parts_used"])
        revenue = float(item["parts_revenue"] or 0) + float(item["shell_sale_price"] or 0) + float(item["vehicle_sale"] or 0)
        item["invested"] = invested
        item["total_revenue"] = revenue
        item["profit"] = revenue - invested
        item["recovery_pct"] = (revenue / invested * 100) if invested > 0 else (100.0 if revenue > 0 else 0.0)
        item["remaining_to_break_even"] = max(0.0, invested - revenue)
        report_rows.append(item)
    totals = {
        "invested": sum(r["invested"] for r in report_rows),
        "parts_revenue": sum(float(r["parts_revenue"] or 0) for r in report_rows),
        "shell_revenue": sum(float(r["shell_sale_price"] or 0) for r in report_rows),
        "remaining_retail": sum(float(r["remaining_retail"] or 0) for r in report_rows),
        "profit": sum(r["profit"] for r in report_rows),
    }
    return render_template("dismantling_profit_report.html", rows=report_rows, totals=totals)


@app.route("/invoices")
@login_required
def invoice_centre():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    conn = db()
    sql = """
        SELECT s.*, v.stock_no, v.asset_type, v.year, v.make, v.model, v.registration
        FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
        WHERE 1=1
    """
    params = []
    if q:
        like = f"%{q}%"
        sql += " AND (s.invoice_number LIKE ? OR s.buyer_name LIKE ? OR v.stock_no LIKE ? OR v.registration LIKE ?)"
        params.extend([like, like, like, like])
    if status:
        sql += " AND COALESCE(s.invoice_status,'Draft')=?"
        params.append(status)
    sql += " ORDER BY COALESCE(s.sale_date, s.id) DESC, s.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return render_template("invoice_centre.html", invoices=rows, q=q, status=status)


@app.route("/vehicles/<int:vehicle_id>/invoice/edit", methods=["GET", "POST"])
@login_required
def sale_invoice_edit(vehicle_id):
    import json
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    if not vehicle or not sale:
        conn.close()
        return "Sale invoice is not available until a sale is recorded.", 404

    if request.method == "POST":
        tracked = [
            "sale_date", "buyer_name", "buyer_phone", "buyer_email", "buyer_address",
            "sale_price_inc_gst", "advertising_cost", "transfer_cost", "deposit_amount",
            "deposit_date", "payment_method", "trade_in_description", "trade_in_value",
            "warranty_type", "warranty_expiry", "invoice_status", "notes"
        ]
        old_values = {key: sale[key] for key in tracked}
        price = float(request.form.get("sale_price_inc_gst") or 0)
        new_values = {
            "sale_date": request.form.get("sale_date"),
            "buyer_name": request.form.get("buyer_name"),
            "buyer_phone": request.form.get("buyer_phone"),
            "buyer_email": request.form.get("buyer_email"),
            "buyer_address": request.form.get("buyer_address"),
            "sale_price_inc_gst": price,
            "advertising_cost": float(request.form.get("advertising_cost") or 0),
            "transfer_cost": float(request.form.get("transfer_cost") or 0),
            "deposit_amount": float(request.form.get("deposit_amount") or 0),
            "deposit_date": request.form.get("deposit_date"),
            "payment_method": request.form.get("payment_method"),
            "trade_in_description": request.form.get("trade_in_description"),
            "trade_in_value": float(request.form.get("trade_in_value") or 0),
            "warranty_type": request.form.get("warranty_type"),
            "warranty_expiry": request.form.get("warranty_expiry"),
            "invoice_status": request.form.get("invoice_status") or "Draft",
            "notes": request.form.get("notes"),
        }
        gst = round(price / 11, 2)
        conn.execute("""
            UPDATE sales SET
              sale_date=?,buyer_name=?,buyer_phone=?,buyer_email=?,buyer_address=?,
              sale_price_inc_gst=?,sale_gst=?,advertising_cost=?,transfer_cost=?,
              deposit_amount=?,deposit_date=?,payment_method=?,trade_in_description=?,trade_in_value=?,
              warranty_type=?,warranty_expiry=?,invoice_status=?,notes=?,updated_at=CURRENT_TIMESTAMP
            WHERE vehicle_id=?
        """, (
            new_values["sale_date"], new_values["buyer_name"], new_values["buyer_phone"],
            new_values["buyer_email"], new_values["buyer_address"], price, gst,
            new_values["advertising_cost"], new_values["transfer_cost"], new_values["deposit_amount"],
            new_values["deposit_date"], new_values["payment_method"], new_values["trade_in_description"],
            new_values["trade_in_value"], new_values["warranty_type"], new_values["warranty_expiry"],
            new_values["invoice_status"], new_values["notes"], vehicle_id
        ))
        changed = {k: {"from": old_values.get(k), "to": new_values.get(k)} for k in tracked if str(old_values.get(k) or "") != str(new_values.get(k) or "")}
        note = request.form.get("change_note") or "Invoice updated"
        conn.execute("""
            INSERT INTO invoice_history(sale_id,vehicle_id,invoice_number,changed_by,change_note,old_values,new_values)
            VALUES(?,?,?,?,?,?,?)
        """, (sale["id"], vehicle_id, sale["invoice_number"], session.get("display_name"), note,
              json.dumps(old_values, default=str), json.dumps(new_values, default=str)))
        conn.commit()
        conn.close()
        log_action("Invoice edited", "vehicle", vehicle_id, f"{sale['invoice_number']}: {note}; {len(changed)} field(s) changed")
        flash(f"Invoice {sale['invoice_number']} updated.", "success")
        if request.form.get("save_continue"):
            return redirect(url_for("sale_invoice_edit", vehicle_id=vehicle_id))
        return redirect(url_for("sale_invoice", vehicle_id=vehicle_id))

    history = conn.execute("SELECT * FROM invoice_history WHERE sale_id=? ORDER BY changed_at DESC,id DESC", (sale["id"],)).fetchall()
    conn.close()
    return render_template("invoice_edit.html", vehicle=vehicle, sale=sale, history=history)


@app.route("/vehicles/<int:vehicle_id>/deal-file", methods=["GET", "POST"])
@login_required
def deal_file(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    if not vehicle or not sale:
        conn.close()
        return "A deal file is available after a sale has been recorded.", 404

    if request.method == "POST":
        conn.execute("""
            UPDATE sales SET delivery_status=?,delivery_date=?,keys_handed_over=?,
              registration_transferred=?,customer_signature_received=?,
              finance_documents_complete=?,warranty_documents_complete=?,deal_notes=?,
              updated_at=CURRENT_TIMESTAMP
            WHERE vehicle_id=?
        """, (
            request.form.get("delivery_status") or "Preparing",
            request.form.get("delivery_date") or None,
            1 if request.form.get("keys_handed_over") else 0,
            1 if request.form.get("registration_transferred") else 0,
            1 if request.form.get("customer_signature_received") else 0,
            1 if request.form.get("finance_documents_complete") else 0,
            1 if request.form.get("warranty_documents_complete") else 0,
            request.form.get("deal_notes"),
            vehicle_id,
        ))
        conn.commit()
        conn.close()
        log_action("Deal file updated", "vehicle", vehicle_id, request.form.get("delivery_status") or "Preparing")
        flash("Deal file updated.", "success")
        return redirect(url_for("deal_file", vehicle_id=vehicle_id))

    documents = conn.execute(
        "SELECT * FROM vehicle_documents WHERE vehicle_id=? ORDER BY uploaded_at DESC,id DESC",
        (vehicle_id,),
    ).fetchall()
    photos = conn.execute(
        "SELECT * FROM vehicle_photos WHERE vehicle_id=? ORDER BY id DESC", (vehicle_id,)
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM invoice_history WHERE sale_id=? ORDER BY changed_at DESC,id DESC", (sale["id"],)
    ).fetchall()
    expenses = conn.execute(
        "SELECT * FROM expenses WHERE vehicle_id=? ORDER BY expense_date DESC,id DESC", (vehicle_id,)
    ).fetchall()
    jobs = conn.execute(
        "SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY job_date DESC,id DESC", (vehicle_id,)
    ).fetchall()
    conn.close()
    balance_due = max(float(sale["sale_price_inc_gst"] or 0) - float(sale["deposit_amount"] or 0), 0)
    return render_template(
        "deal_file.html", vehicle=vehicle, sale=sale, documents=documents, photos=photos,
        history=history, expenses=expenses, jobs=jobs, balance_due=balance_due
    )


@app.route("/vehicles/<int:vehicle_id>/invoice")
@login_required
def sale_invoice(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Sale invoice is not available until a sale is recorded.", 404
    sale_ex_gst = sale["sale_price_inc_gst"] - sale["sale_gst"]
    return render_template("sale_invoice.html", vehicle=vehicle, sale=sale, sale_ex_gst=sale_ex_gst)


@app.route("/export/vehicles.csv")
@login_required
def export_vehicles_csv():
    conn = db()
    rows = conn.execute("""
        SELECT stock_no,status,purchase_date,year,make,model,variant,vin,registration,
               odometer_km,purchase_price_inc_gst,purchase_gst,barry_contribution,
               matt_contribution,rego_expiry,ppsr_number,roadworthy_status,service_due_date
        FROM vehicles ORDER BY id
    """).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(rows[0].keys() if rows else [
        "stock_no","status","purchase_date","year","make","model","variant","vin",
        "registration","odometer_km","purchase_price_inc_gst","purchase_gst",
        "barry_contribution","matt_contribution","rego_expiry","ppsr_number",
        "roadworthy_status","service_due_date"
    ])
    for row in rows:
        writer.writerow(tuple(row))
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=BAM_vehicle_export.csv"},
    )


@app.route("/export/profit.csv")
@login_required
def export_profit_csv():
    conn = db()
    rows = conn.execute("""
        SELECT v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v ORDER BY v.id
    """).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Stock","Year","Make","Model","Status","Purchase","Expenses","Jobs","Sale","Profit/Loss"])
    for row in rows:
        profit = row["sale_price"] - row["purchase_price_inc_gst"] - row["expenses"] - row["jobs"]
        writer.writerow([
            row["stock_no"],row["year"],row["make"],row["model"],row["status"],
            row["purchase_price_inc_gst"],row["expenses"],row["jobs"],row["sale_price"],profit
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=BAM_profit_export.csv"},
    )


@app.route("/vehicles/<int:vehicle_id>/dismantling", methods=["GET", "POST"])
@login_required
def vehicle_dismantling(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    if request.method == "POST":
        action = request.form.get("action") or "add_part"
        if action == "update_status":
            status = request.form.get("dismantling_status") or "Not Started"
            conn.execute(
                "UPDATE vehicles SET vehicle_purpose='Parts Vehicle', dismantling_status=? WHERE id=?",
                (status, vehicle_id),
            )
            conn.commit()
            conn.close()
            flash("Dismantling status updated.", "success")
            return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))

        if action == "shell_sale":
            price = float(request.form.get("shell_sale_price") or 0)
            conn.execute(
                "UPDATE vehicles SET vehicle_purpose='Parts Vehicle', shell_sale_price=?, shell_sale_date=? WHERE id=?",
                (price, request.form.get("shell_sale_date") or None, vehicle_id),
            )
            conn.commit()
            conn.close()
            flash("Shell sale recorded.", "success")
            return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))

        part_name = (request.form.get("part_name") or "").strip()
        if not part_name:
            conn.close()
            flash("Part name is required.", "error")
            return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))
        qty = float(request.form.get("quantity_on_hand") or 1)
        cost = float(request.form.get("unit_cost_inc_gst") or 0)
        selling = float(request.form.get("selling_price") or 0)
        gst = round(cost / 11, 2)
        # PRT numbers are allocated by the server from one business-wide sequence.
        # Ignore stale PRT values posted by an open browser form so a duplicate SKU
        # can never be reused accidentally. Custom non-PRT numbers are still allowed.
        requested_number = (request.form.get("part_number") or "").strip()
        part_number = requested_number if requested_number and not requested_number.upper().startswith("PRT-") else None
        try:
            conn.execute("BEGIN IMMEDIATE")
            if not part_number:
                part_number = next_part_number(conn)
            conn.execute("""
                INSERT INTO parts(
                    part_number,part_name,category,supplier,quantity_on_hand,reorder_level,
                    unit_cost_inc_gst,gst_amount_per_unit,storage_location,notes,
                    vehicle_id,vehicle_stock_no,vin,make,model,year,condition,selling_price,status,date_added,
                    position,fitment,manufacturer_part_no,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,?,?,?,CURRENT_TIMESTAMP)
            """, (
                part_number, part_name, request.form.get("category"),
                "Donor vehicle", qty, 0, cost, gst, request.form.get("storage_location"),
                request.form.get("notes"), vehicle_id, vehicle["stock_no"], vehicle["vin"],
                vehicle["make"], vehicle["model"], vehicle["year"],
                request.form.get("condition") or "Used", selling,
                request.form.get("status") or "In Stock", request.form.get("position"),
                request.form.get("fitment"), request.form.get("manufacturer_part_no"),
            ))
            conn.execute(
                "UPDATE vehicles SET vehicle_purpose='Parts Vehicle', dismantling_status=CASE WHEN COALESCE(dismantling_status,'Not Started')='Not Started' THEN 'Dismantling' ELSE dismantling_status END WHERE id=?",
                (vehicle_id,),
            )
            conn.commit()
        except sqlite3.IntegrityError as exc:
            conn.rollback()
            conn.close()
            flash(f"Could not add part: {exc}", "error")
            return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))
        except sqlite3.OperationalError as exc:
            conn.rollback()
            conn.close()
            flash(f"Database busy while adding part. Please try again: {exc}", "error")
            return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))
        conn.close()
        flash(f"{part_name} added to Parts Centre.", "success")
        return redirect(url_for("vehicle_dismantling", vehicle_id=vehicle_id))

    parts = conn.execute("SELECT * FROM parts WHERE vehicle_id=? ORDER BY id DESC", (vehicle_id,)).fetchall()
    revenue = float(conn.execute("""
        SELECT COALESCE(SUM(ps.sale_price),0) AS revenue
        FROM part_sales ps JOIN parts p ON p.id=ps.part_id
        WHERE p.vehicle_id=?
    """, (vehicle_id,)).fetchone()["revenue"] or 0)
    remaining_retail = float(conn.execute(
        "SELECT COALESCE(SUM(quantity_on_hand*selling_price),0) AS v FROM parts WHERE vehicle_id=?",
        (vehicle_id,),
    ).fetchone()["v"] or 0)
    expense_total = float(conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM expenses WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"] or 0)
    job_total = float(conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) AS v FROM job_cards WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"] or 0)
    service_total = float(conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM service_entries WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"] or 0)
    parts_used_total = float(conn.execute("SELECT COALESCE(SUM(quantity_used*unit_cost_inc_gst),0) AS v FROM part_usage WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"] or 0)
    invested = float(vehicle["purchase_price_inc_gst"] or 0) + expense_total + job_total + service_total + parts_used_total
    realised_revenue = revenue + float(vehicle["shell_sale_price"] or 0)
    profit = realised_revenue - invested
    recovery_pct = (realised_revenue / invested * 100) if invested > 0 else (100.0 if realised_revenue > 0 else 0.0)
    remaining_to_break_even = max(0.0, invested - realised_revenue)
    break_even_status = "Investment recovered" if realised_revenue >= invested and invested > 0 else "Recovering investment"
    conn.close()
    return render_template(
        "dismantling.html", vehicle=vehicle, parts=parts, revenue=revenue, invested=invested, profit=profit,
        remaining_retail=remaining_retail, recovery_pct=recovery_pct,
        remaining_to_break_even=remaining_to_break_even, break_even_status=break_even_status,
    )


@app.route("/parts/<int:part_id>/sell", methods=["POST"])
@login_required
def part_sell(part_id):
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()
    if not part:
        conn.close()
        flash("Part not found.", "error")
        return redirect(url_for("parts_page"))
    qty = float(request.form.get("quantity") or 1)
    if qty <= 0 or qty > float(part["quantity_on_hand"] or 0):
        conn.close()
        flash("Sale quantity is invalid.", "error")
        return redirect(request.referrer or url_for("parts_page"))
    sale_price = float(request.form.get("sale_price") or part["selling_price"] or 0)
    remaining = max(0, float(part["quantity_on_hand"] or 0) - qty)
    try:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("""
            INSERT INTO part_sales(
                part_id,quantity,customer_name,customer_phone,customer_email,sale_price,freight_cost,
                payment_method,warranty,invoice_number,notes
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
        """, (
            part_id, qty, request.form.get("customer_name"), request.form.get("customer_phone"),
            request.form.get("customer_email"), sale_price, float(request.form.get("freight_cost") or 0),
            request.form.get("payment_method"), request.form.get("warranty"),
            request.form.get("invoice_number"), request.form.get("notes"),
        ))
        conn.execute(
            "UPDATE parts SET quantity_on_hand=?, status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (remaining, "Sold" if remaining <= 0 else "In Stock", part_id),
        )
        conn.commit()
    except (sqlite3.IntegrityError, sqlite3.OperationalError, ValueError) as exc:
        conn.rollback()
        conn.close()
        flash(f"Could not record part sale: {exc}", "error")
        return redirect(request.referrer or url_for("part_detail", part_id=part_id))
    vehicle_id = part["vehicle_id"]
    part_number = part["part_number"] or f"Part {part_id}"
    conn.close()
    log_action(
        "Part sold", "part", part_id,
        f"{part_number}; qty {qty:g}; sale ${sale_price:,.2f}; donor {part['vehicle_stock_no'] or 'none'}",
    )
    flash(f"Part sale recorded. {part_number} revenue is now included in the donor vehicle Financial Snapshot.", "success")
    if vehicle_id:
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))
    return redirect(url_for("part_detail", part_id=part_id))


@app.route("/parts", methods=["GET", "POST"])
@login_required
def parts_page():
    conn = db()
    if request.method == "POST":
        part_name = (request.form.get("part_name") or "").strip()
        if not part_name:
            conn.close()
            flash("Part name is required.", "error")
            return redirect(url_for("parts_page"))

        quantity = float(request.form.get("quantity_on_hand") or 0)
        unit_cost = float(request.form.get("unit_cost_inc_gst") or 0)
        gst = round(unit_cost / 11, 2)
        requested_number = (request.form.get("part_number") or "").strip()
        part_number = requested_number if requested_number and not requested_number.upper().startswith("PRT-") else None
        source_stock = (request.form.get("vehicle_stock_no") or "").strip()
        donor = None
        if source_stock:
            donor = conn.execute("SELECT * FROM vehicles WHERE stock_no=?", (source_stock,)).fetchone()

        try:
            conn.execute("BEGIN IMMEDIATE")
            if not part_number:
                part_number = next_part_number(conn)
            conn.execute("""
                INSERT INTO parts(
                    part_number,part_name,category,supplier,quantity_on_hand,reorder_level,
                    unit_cost_inc_gst,gst_amount_per_unit,storage_location,notes,
                    vehicle_id,vehicle_stock_no,vin,make,model,year,condition,selling_price,status,
                    position,fitment,manufacturer_part_no,barcode,date_added,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)
            """, (
                part_number, part_name, request.form.get("category"), request.form.get("supplier"),
                quantity, float(request.form.get("reorder_level") or 0), unit_cost, gst,
                request.form.get("storage_location"), request.form.get("notes"),
                donor["id"] if donor else None, source_stock or None,
                donor["vin"] if donor else request.form.get("vin"),
                donor["make"] if donor else request.form.get("make"),
                donor["model"] if donor else request.form.get("model"),
                donor["year"] if donor else (request.form.get("year") or None),
                request.form.get("condition") or "Used", float(request.form.get("selling_price") or 0),
                request.form.get("status") or "In Stock", request.form.get("position"),
                request.form.get("fitment"), request.form.get("manufacturer_part_no"), request.form.get("barcode"),
            ))
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            flash(f"Part number {part_number} already exists.", "error")
            return redirect(url_for("parts_page"))
        except sqlite3.OperationalError as exc:
            conn.rollback()
            conn.close()
            flash(f"Database busy while adding part. Please try again: {exc}", "error")
            return redirect(url_for("parts_page"))
        conn.commit()
        part_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.close()
        log_action("Part added", "part", part_id, f"{part_number} - {part_name}")
        flash(f"{part_number} added to Parts Centre.", "success")
        return redirect(url_for("part_detail", part_id=part_id))

    q = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    category_filter = request.args.get("category", "").strip()
    source_filter = request.args.get("source", "").strip()

    where = []
    params = []
    if q:
        token = f"%{q}%"
        where.append("(part_number LIKE ? OR part_name LIKE ? OR category LIKE ? OR supplier LIKE ? OR storage_location LIKE ? OR vehicle_stock_no LIKE ? OR manufacturer_part_no LIKE ? OR barcode LIKE ?)")
        params.extend([token] * 8)
    if status_filter:
        where.append("status=?")
        params.append(status_filter)
    if category_filter:
        where.append("category=?")
        params.append(category_filter)
    if source_filter:
        where.append("vehicle_stock_no=?")
        params.append(source_filter)
    sql = "SELECT * FROM parts"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY CASE WHEN status='In Stock' THEN 0 WHEN status='Reserved' THEN 1 ELSE 2 END, part_name, id DESC"
    rows = conn.execute(sql, params).fetchall()

    part_metrics = conn.execute("""
        SELECT
          COUNT(*) AS part_lines,
          COALESCE(SUM(quantity_on_hand),0) AS parts_on_hand,
          COALESCE(SUM(quantity_on_hand * unit_cost_inc_gst),0) AS stock_cost,
          COALESCE(SUM(quantity_on_hand * selling_price),0) AS retail_value,
          SUM(CASE WHEN quantity_on_hand <= reorder_level AND status NOT IN ('Sold','Scrap') THEN 1 ELSE 0 END) AS low_stock,
          SUM(CASE WHEN status='Reserved' THEN 1 ELSE 0 END) AS reserved_lines
        FROM parts
    """).fetchone()
    sales_metrics = conn.execute("""
        SELECT COALESCE(SUM(quantity),0) AS units_sold,
               COALESCE(SUM(sale_price),0) AS revenue,
               COALESCE(SUM(freight_cost),0) AS freight
        FROM part_sales
    """).fetchone()
    donor_vehicles = conn.execute("""
        SELECT v.*,
               COUNT(p.id) AS part_lines,
               COALESCE(SUM(p.quantity_on_hand),0) AS units_remaining,
               COALESCE(SUM(p.quantity_on_hand * p.selling_price),0) AS remaining_retail
        FROM vehicles v
        LEFT JOIN parts p ON p.vehicle_id=v.id
        WHERE COALESCE(v.vehicle_purpose,'Retail Sale')='Parts Vehicle' OR v.status='BER'
        GROUP BY v.id
        ORDER BY CASE WHEN COALESCE(v.dismantling_status,'Not Started')='Complete' THEN 1 ELSE 0 END, v.id DESC
    """).fetchall()
    categories = conn.execute("SELECT DISTINCT category FROM parts WHERE COALESCE(category,'')!='' ORDER BY category").fetchall()
    sources = conn.execute("SELECT DISTINCT vehicle_stock_no FROM parts WHERE COALESCE(vehicle_stock_no,'')!='' ORDER BY vehicle_stock_no").fetchall()
    next_part = next_part_number(conn)
    conn.close()
    return render_template(
        "parts.html", parts=rows, q=q, status_filter=status_filter, category_filter=category_filter,
        source_filter=source_filter, part_metrics=part_metrics, sales_metrics=sales_metrics,
        donor_vehicles=donor_vehicles, categories=categories, sources=sources, next_part=next_part,
    )


@app.route("/parts/<int:part_id>")
@login_required
def part_detail(part_id):
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()
    if not part:
        conn.close()
        return "Part not found", 404
    photos = conn.execute("SELECT * FROM part_photos WHERE part_id=? ORDER BY is_featured DESC,id DESC", (part_id,)).fetchall()
    sales = conn.execute("SELECT * FROM part_sales WHERE part_id=? ORDER BY sale_date DESC,id DESC", (part_id,)).fetchall()
    sold_qty = sum(float(r["quantity"] or 0) for r in sales)
    revenue = sum(float(r["sale_price"] or 0) for r in sales)
    freight = sum(float(r["freight_cost"] or 0) for r in sales)
    sold_cost = sold_qty * float(part["unit_cost_inc_gst"] or 0)
    gross_profit = revenue - sold_cost - freight
    donor = conn.execute("SELECT * FROM vehicles WHERE id=?", (part["vehicle_id"],)).fetchone() if part["vehicle_id"] else None
    price_suggestion = parts_price_suggestion(conn, part)
    conn.close()
    return render_template("part_detail.html", part=part, photos=photos, sales=sales, donor=donor,
                           sold_qty=sold_qty, revenue=revenue, freight=freight, gross_profit=gross_profit, price_suggestion=price_suggestion)


@app.route("/parts/<int:part_id>/edit", methods=["GET", "POST"])
@login_required
def part_edit(part_id):
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()
    if not part:
        conn.close()
        return "Part not found", 404
    if request.method == "POST":
        part_number = (request.form.get("part_number") or "").strip() or part["part_number"] or next_part_number(conn)
        source_stock = (request.form.get("vehicle_stock_no") or "").strip()
        donor = conn.execute("SELECT * FROM vehicles WHERE stock_no=?", (source_stock,)).fetchone() if source_stock else None
        try:
            conn.execute("""
                UPDATE parts SET
                    part_number=?,part_name=?,category=?,subcategory=?,description=?,supplier=?,quantity_on_hand=?,reorder_level=?,
                    unit_cost_inc_gst=?,gst_amount_per_unit=?,selling_price=?,storage_location=?,notes=?,vehicle_id=?,vehicle_stock_no=?,
                    vin=?,make=?,model=?,year=?,condition=?,status=?,engine_code=?,transmission_code=?,barcode=?,position=?,fitment=?,
                    manufacturer_part_no=?,reserved_for=?,reserved_until=?,updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (
                part_number, request.form.get("part_name"), request.form.get("category"), request.form.get("subcategory"),
                request.form.get("description"), request.form.get("supplier"), float(request.form.get("quantity_on_hand") or 0),
                float(request.form.get("reorder_level") or 0), float(request.form.get("unit_cost_inc_gst") or 0),
                round(float(request.form.get("unit_cost_inc_gst") or 0)/11, 2), float(request.form.get("selling_price") or 0),
                request.form.get("storage_location"), request.form.get("notes"), donor["id"] if donor else part["vehicle_id"],
                source_stock or None, donor["vin"] if donor else request.form.get("vin"), donor["make"] if donor else request.form.get("make"),
                donor["model"] if donor else request.form.get("model"), donor["year"] if donor else (request.form.get("year") or None),
                request.form.get("condition") or "Used", request.form.get("status") or "In Stock", request.form.get("engine_code"),
                request.form.get("transmission_code"), request.form.get("barcode"), request.form.get("position"), request.form.get("fitment"),
                request.form.get("manufacturer_part_no"), request.form.get("reserved_for"), request.form.get("reserved_until") or None, part_id,
            ))
        except sqlite3.IntegrityError:
            conn.rollback()
            conn.close()
            flash(f"Part number {part_number} already exists.", "error")
            return redirect(url_for("part_edit", part_id=part_id))
        except sqlite3.OperationalError as exc:
            conn.rollback()
            conn.close()
            flash(f"Could not update part: {exc}", "error")
            return redirect(url_for("part_edit", part_id=part_id))
        conn.commit()
        conn.close()
        log_action(
            "Part updated", "part", part_id,
            f"{part_number}; {request.form.get('part_name') or ''}; qty {request.form.get('quantity_on_hand') or 0}; status {request.form.get('status') or 'In Stock'}",
        )
        flash("Part updated.", "success")
        return redirect(url_for("part_detail", part_id=part_id))
    donor_vehicles = conn.execute("SELECT id,stock_no,year,make,model FROM vehicles WHERE vehicle_purpose='Parts Vehicle' OR status='BER' ORDER BY stock_no").fetchall()
    conn.close()
    return render_template("part_edit.html", part=part, donor_vehicles=donor_vehicles)


@app.route("/parts/<int:part_id>/photos", methods=["POST"])
@login_required
def part_photo_add(part_id):
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()
    if not part:
        conn.close()
        return "Part not found", 404
    files = request.files.getlist("photos")
    added = 0
    image_exts = {"png", "jpg", "jpeg", "webp"}
    for file in files:
        if not file or not file.filename:
            continue
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in image_exts:
            continue
        filename = save_upload(file)
        featured = 1 if added == 0 and not conn.execute("SELECT id FROM part_photos WHERE part_id=? LIMIT 1", (part_id,)).fetchone() else 0
        conn.execute("INSERT INTO part_photos(part_id,filename,caption,is_featured) VALUES(?,?,?,?)",
                     (part_id, filename, request.form.get("caption"), featured))
        added += 1
    conn.commit()
    conn.close()
    flash(f"{added} part photo(s) uploaded." if added else "No valid images selected.", "success" if added else "error")
    return redirect(url_for("part_detail", part_id=part_id))


@app.route("/parts/<int:part_id>/photos/<int:photo_id>/feature", methods=["POST"])
@login_required
def part_photo_feature(part_id, photo_id):
    conn = db()
    conn.execute("UPDATE part_photos SET is_featured=0 WHERE part_id=?", (part_id,))
    conn.execute("UPDATE part_photos SET is_featured=1 WHERE id=? AND part_id=?", (photo_id, part_id))
    conn.commit(); conn.close()
    flash("Featured part photo updated.", "success")
    return redirect(url_for("part_detail", part_id=part_id))


@app.route("/parts/<int:part_id>/photos/<int:photo_id>/delete", methods=["POST"])
@login_required
def part_photo_delete(part_id, photo_id):
    conn = db()
    photo = conn.execute("SELECT * FROM part_photos WHERE id=? AND part_id=?", (photo_id, part_id)).fetchone()
    if photo:
        conn.execute("DELETE FROM part_photos WHERE id=?", (photo_id,))
        conn.commit()
    conn.close()
    if photo:
        try:
            (UPLOAD_DIR / photo["filename"]).unlink(missing_ok=True)
        except OSError:
            pass
    flash("Part photo deleted.", "success")
    return redirect(url_for("part_detail", part_id=part_id))


@app.route("/vehicles/<int:vehicle_id>/use-part", methods=["POST"])
@login_required
def use_part(vehicle_id):
    part_id = int(request.form.get("part_id"))
    quantity = float(request.form.get("quantity_used") or 0)
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()

    if not part:
        conn.close()
        flash("Part not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    if quantity <= 0:
        conn.close()
        flash("Quantity must be greater than zero.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    if quantity > part["quantity_on_hand"]:
        conn.close()
        flash("Not enough stock available.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute("""
        INSERT INTO part_usage(
            part_id,vehicle_id,job_card_id,usage_date,quantity_used,unit_cost_inc_gst,paid_by,notes
        ) VALUES(?,?,?,?,?,?,?,?)
    """, (
        part_id,
        vehicle_id,
        int(request.form.get("job_card_id")) if request.form.get("job_card_id") else None,
        request.form.get("usage_date"),
        quantity,
        part["unit_cost_inc_gst"],
        request.form.get("paid_by"),
        request.form.get("notes"),
    ))
    conn.execute(
        "UPDATE parts SET quantity_on_hand=quantity_on_hand-? WHERE id=?",
        (quantity, part_id),
    )
    conn.commit()
    conn.close()
    flash("Part allocated to vehicle.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/timeline")
@login_required
def vehicle_timeline(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    events = []
    def add_event(event_date, event_type, title, details="", amount=None, status=""):
        if event_date:
            events.append({
                "date": str(event_date), "type": event_type, "title": title,
                "details": details or "", "amount": amount, "status": status or ""
            })

    add_event(vehicle["created_at"], "Vehicle", "Vehicle created", f"{vehicle['stock_no']} - {vehicle['year'] or ''} {vehicle['make']} {vehicle['model']}")
    add_event(vehicle["purchase_date"], "Purchase", "Vehicle purchased", "Purchase recorded", vehicle["purchase_price_inc_gst"])

    for row in conn.execute("SELECT * FROM expenses WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        add_event(row["expense_date"], "Expense", row["category"] or "Expense", row["description"], row["cost_inc_gst"])
    for row in conn.execute("SELECT * FROM job_cards WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        amount = row["actual_cost_inc_gst"] if float(row["actual_cost_inc_gst"] or 0) > 0 else row["estimated_cost"]
        add_event(row["job_date"], "Workshop", row["description"], row["category"] or "Workshop job", amount, row["status"] or "")
    for row in conn.execute("SELECT * FROM service_entries WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        add_event(row["service_date"], "Service", row["service_type"] or "Service", row["description"], row["cost_inc_gst"])
    for row in conn.execute("SELECT * FROM reminders WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        add_event(row["reminder_date"], "Reminder", row["title"], row["reminder_type"], None, "Completed" if row["completed"] else "Open")
    for row in conn.execute("SELECT * FROM tasks WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        add_event(row["task_date"] or row["created_at"], "Task", row["title"], row["category"] or "", None, row["status"] or "")
    for row in conn.execute("SELECT * FROM parts WHERE vehicle_id=?", (vehicle_id,)).fetchall():
        add_event(row["date_added"] or row["updated_at"], "Dismantling", f"Part added: {row['part_number'] or ''} {row['part_name']}", row["storage_location"] or "", row["selling_price"], row["status"] or "")
    for row in conn.execute("""
        SELECT ps.*,p.part_number,p.part_name FROM part_sales ps
        JOIN parts p ON p.id=ps.part_id WHERE p.vehicle_id=?
    """, (vehicle_id,)).fetchall():
        add_event(row["sale_date"], "Part Sale", f"Sold {row['part_number'] or ''} {row['part_name']}", row["customer_name"] or "", row["sale_price"], "Sold")
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    if sale:
        add_event(sale["sale_date"], "Vehicle Sale", f"Vehicle sold to {sale['buyer_name'] or 'buyer'}", sale["invoice_number"] or "", sale["sale_price_inc_gst"], sale["invoice_status"] or "")
    for row in conn.execute("SELECT * FROM audit_log WHERE entity_type='vehicle' AND entity_id=?", (vehicle_id,)).fetchall():
        add_event(row["created_at"], "Activity", row["action"], row["details"] or "")

    events.sort(key=lambda x: x["date"] or "", reverse=True)
    conn.close()
    return render_template("vehicle_timeline.html", vehicle=vehicle, events=events)


@app.route("/search")
@login_required
def global_search():
    q = request.args.get("q", "").strip()
    conn = db()
    vehicles = []
    contacts = []
    parts = []
    invoices = []
    if q:
        pattern = f"%{q}%"
        vehicles = conn.execute("""
            SELECT * FROM vehicles
            WHERE stock_no LIKE ? OR vin LIKE ? OR registration LIKE ? OR make LIKE ? OR model LIKE ?
            ORDER BY id DESC LIMIT 50
        """, (pattern, pattern, pattern, pattern, pattern)).fetchall()
        contacts = conn.execute("""
            SELECT * FROM contacts
            WHERE name LIKE ? OR phone LIKE ? OR email LIKE ?
            ORDER BY name LIMIT 50
        """, (pattern, pattern, pattern)).fetchall()
        parts = conn.execute("""
            SELECT * FROM parts
            WHERE part_number LIKE ? OR part_name LIKE ? OR manufacturer_part_no LIKE ?
               OR barcode LIKE ? OR vehicle_stock_no LIKE ? OR storage_location LIKE ? OR fitment LIKE ?
            ORDER BY id DESC LIMIT 50
        """, (pattern, pattern, pattern, pattern, pattern, pattern, pattern)).fetchall()
        invoices = conn.execute("""
            SELECT s.*,v.stock_no,v.make,v.model,v.year FROM sales s
            JOIN vehicles v ON v.id=s.vehicle_id
            WHERE s.invoice_number LIKE ? OR s.buyer_name LIKE ? OR s.buyer_phone LIKE ?
               OR v.stock_no LIKE ? OR v.registration LIKE ?
            ORDER BY s.id DESC LIMIT 50
        """, (pattern, pattern, pattern, pattern, pattern)).fetchall()
    conn.close()
    return render_template("global_search.html", q=q, vehicles=vehicles, contacts=contacts, parts=parts, invoices=invoices)


@app.route("/reports/finance")
@login_required
def finance_summary():
    conn = db()
    rows = conn.execute("""
        SELECT v.id,v.stock_no,v.year,v.make,v.model,v.status,
               v.purchase_price_inc_gst,v.barry_contribution,v.matt_contribution,
               COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
               COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                         FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs,
               COALESCE((SELECT SUM(s.cost_inc_gst) FROM service_entries s WHERE s.vehicle_id=v.id),0) AS services,
               COALESCE((SELECT SUM(u.quantity_used*u.unit_cost_inc_gst) FROM part_usage u WHERE u.vehicle_id=v.id),0) AS parts,
               COALESCE((SELECT sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0) AS sale_price
        FROM vehicles v
        ORDER BY v.id DESC
    """).fetchall()
    conn.close()

    report = []
    total_barry = 0
    total_matt = 0
    total_profit = 0
    money_in_stock = 0

    for r in rows:
        total_cost = r["purchase_price_inc_gst"] + r["expenses"] + r["jobs"] + r["services"] + r["parts"]
        profit = r["sale_price"] - total_cost
        if r["status"] not in ("Sold", "BER"):
            money_in_stock += total_cost
        total_barry += r["barry_contribution"]
        total_matt += r["matt_contribution"]
        total_profit += profit
        report.append(dict(r) | {"total_cost": total_cost, "profit": profit})

    return render_template(
        "finance_summary.html",
        rows=report,
        total_barry=total_barry,
        total_matt=total_matt,
        total_profit=total_profit,
        money_in_stock=money_in_stock,
        barry_profit_share=total_profit/2,
        matt_profit_share=total_profit/2,
    )


@app.route("/reports/performance")
@login_required
def performance_report():
    conn = db()
    monthly = conn.execute("""
        SELECT substr(sale_date,1,7) AS month,
               COUNT(*) AS vehicles_sold,
               SUM(sale_price_inc_gst) AS sales_total
        FROM sales
        WHERE sale_date IS NOT NULL AND sale_date!=''
        GROUP BY substr(sale_date,1,7)
        ORDER BY month
    """).fetchall()

    makes = conn.execute("""
        SELECT make, COUNT(*) AS total
        FROM vehicles
        GROUP BY make
        ORDER BY total DESC, make
        LIMIT 10
    """).fetchall()
    conn.close()

    max_sales = max([r["sales_total"] or 0 for r in monthly], default=1)
    return render_template("performance_report.html", monthly=monthly, makes=makes, max_sales=max_sales)


@app.route("/vehicles/<int:vehicle_id>/contract")
@login_required
def sale_contract(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Sales contract is not available until a sale is recorded.", 404
    balance_due = sale["sale_price_inc_gst"] - sale["deposit_amount"] - sale["trade_in_value"]
    return render_template(
        "sale_contract.html",
        vehicle=vehicle,
        sale=sale,
        balance_due=balance_due,
    )


@app.route("/customers")
@login_required
def customer_history():
    conn = db()
    q = request.args.get("q", "").strip()
    pattern = f"%{q}%"
    vehicle_sql = """
        SELECT 'Vehicle' AS purchase_type,s.id AS sale_id,s.vehicle_id,s.sale_date AS purchase_date,
               s.buyer_name AS customer_name,s.buyer_phone AS customer_phone,s.buyer_email AS customer_email,
               s.sale_price_inc_gst AS amount,s.invoice_number,
               v.stock_no AS item_no,(COALESCE(v.year,'') || ' ' || v.make || ' ' || v.model) AS item_name,
               v.registration
        FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
    """
    part_sql = """
        SELECT 'Part' AS purchase_type,ps.id AS sale_id,p.vehicle_id,ps.sale_date AS purchase_date,
               ps.customer_name,ps.customer_phone,ps.customer_email,ps.sale_price AS amount,ps.invoice_number,
               p.part_number AS item_no,p.part_name AS item_name,'' AS registration
        FROM part_sales ps JOIN parts p ON p.id=ps.part_id
    """
    params = []
    if q:
        vehicle_sql += " WHERE s.buyer_name LIKE ? OR s.buyer_phone LIKE ? OR s.buyer_email LIKE ?"
        part_sql += " WHERE ps.customer_name LIKE ? OR ps.customer_phone LIKE ? OR ps.customer_email LIKE ?"
        params = [pattern, pattern, pattern]
    vehicle_rows = conn.execute(vehicle_sql + " ORDER BY purchase_date DESC", params).fetchall()
    part_rows = conn.execute(part_sql + " ORDER BY purchase_date DESC", params).fetchall()
    conn.close()

    transactions = [dict(r) for r in vehicle_rows] + [dict(r) for r in part_rows]
    transactions.sort(key=lambda r: str(r.get("purchase_date") or ""), reverse=True)
    customers = {}
    for row in transactions:
        email = (row.get("customer_email") or "").strip().lower()
        phone = re.sub(r"\s+", "", row.get("customer_phone") or "")
        name = (row.get("customer_name") or "Unknown customer").strip()
        key = email or phone or name.lower()
        profile = customers.setdefault(key, {"name": name, "phone": row.get("customer_phone") or "", "email": row.get("customer_email") or "", "total_spend": 0.0, "vehicle_count": 0, "part_count": 0, "last_purchase": "", "transactions": []})
        profile["total_spend"] += float(row.get("amount") or 0)
        profile["vehicle_count"] += 1 if row["purchase_type"] == "Vehicle" else 0
        profile["part_count"] += 1 if row["purchase_type"] == "Part" else 0
        profile["last_purchase"] = max(profile["last_purchase"], str(row.get("purchase_date") or ""))
        profile["transactions"].append(row)
    customer_rows = sorted(customers.values(), key=lambda x: (x["total_spend"], x["last_purchase"]), reverse=True)
    return render_template("customer_history.html", customers=customer_rows, transactions=transactions, q=q)


@app.route("/reports/bas")
@login_required
def bas_report():
    start = request.args.get("start", "")
    end = request.args.get("end", "")
    conn = db()

    clauses = []
    params = []
    if start:
        clauses.append("date_value>=?")
        params.append(start)
    if end:
        clauses.append("date_value<=?")
        params.append(end)
    where = "WHERE " + " AND ".join(clauses) if clauses else ""

    sales_rows = conn.execute(f"""
        SELECT sale_date AS date_value,sale_price_inc_gst,sale_gst
        FROM sales
        {where.replace('date_value','sale_date')}
    """, params).fetchall()

    vehicle_rows = conn.execute(f"""
        SELECT purchase_date AS date_value,purchase_price_inc_gst,purchase_gst
        FROM vehicles
        {where.replace('date_value','purchase_date')}
    """, params).fetchall()

    expense_rows = conn.execute(f"""
        SELECT expense_date AS date_value,cost_inc_gst,gst_amount
        FROM expenses
        {where.replace('date_value','expense_date')}
    """, params).fetchall()

    job_rows = conn.execute(f"""
        SELECT job_date AS date_value,
               CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END AS cost_value,
               gst_amount
        FROM job_cards
        {where.replace('date_value','job_date')}
    """, params).fetchall()

    service_rows = conn.execute(f"""
        SELECT service_date AS date_value,cost_inc_gst,gst_amount
        FROM service_entries
        {where.replace('date_value','service_date')}
    """, params).fetchall()

    part_rows = conn.execute(f"""
        SELECT usage_date AS date_value,quantity_used*unit_cost_inc_gst AS cost_value,
               (quantity_used*unit_cost_inc_gst)/11.0 AS gst_amount
        FROM part_usage
        {where.replace('date_value','usage_date')}
    """, params).fetchall()
    conn.close()

    sales_total = sum(r["sale_price_inc_gst"] for r in sales_rows)
    gst_collected = sum(r["sale_gst"] for r in sales_rows)
    purchase_total = sum(r["purchase_price_inc_gst"] for r in vehicle_rows)
    gst_vehicle = sum(r["purchase_gst"] for r in vehicle_rows)
    expense_total = sum(r["cost_inc_gst"] for r in expense_rows)
    gst_expenses = sum(r["gst_amount"] for r in expense_rows)
    job_total = sum(r["cost_value"] for r in job_rows)
    gst_jobs = sum(r["gst_amount"] for r in job_rows)
    service_total = sum(r["cost_inc_gst"] for r in service_rows)
    gst_services = sum(r["gst_amount"] for r in service_rows)
    parts_total = sum(r["cost_value"] for r in part_rows)
    gst_parts = sum(r["gst_amount"] for r in part_rows)

    gst_paid = gst_vehicle + gst_expenses + gst_jobs + gst_services + gst_parts
    net_gst = gst_collected - gst_paid

    return render_template(
        "bas_report.html",
        start=start,
        end=end,
        sales_total=sales_total,
        gst_collected=gst_collected,
        purchase_total=purchase_total,
        expense_total=expense_total,
        job_total=job_total,
        service_total=service_total,
        parts_total=parts_total,
        gst_paid=gst_paid,
        net_gst=net_gst,
    )


@app.route("/audit")
@login_required
def audit_page():
    conn = db()
    rows = conn.execute("""
        SELECT * FROM audit_log
        ORDER BY created_at DESC,id DESC
        LIMIT 500
    """).fetchall()
    conn.close()
    return render_template("audit_log.html", rows=rows)


@app.route("/quick-add-vehicle", methods=["POST"])
@login_required
def quick_add_vehicle():
    conn = db()
    stock_no = next_stock_number(conn)
    price = float(request.form.get("purchase_price_inc_gst") or 0)
    gst = round(price / 11, 2)

    cursor = conn.execute("""
        INSERT INTO vehicles(
            stock_no,status,purchase_date,make,model,variant,year,vin,registration,
            odometer_km,colour,purchase_price_inc_gst,purchase_gst,
            barry_contribution,matt_contribution,rego_expiry,photo_filename,notes,
            ppsr_number,roadworthy_status,service_due_date,service_history
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        stock_no,
        request.form.get("status") or "In Stock",
        request.form.get("purchase_date"),
        request.form.get("make"),
        request.form.get("model"),
        request.form.get("variant"),
        int(request.form.get("year") or 0) or None,
        request.form.get("vin") or None,
        request.form.get("registration"),
        int(request.form.get("odometer_km") or 0) or None,
        request.form.get("colour"),
        price,
        gst,
        float(request.form.get("barry_contribution") or 0),
        float(request.form.get("matt_contribution") or 0),
        request.form.get("rego_expiry"),
        None,
        request.form.get("notes"),
        request.form.get("ppsr_number"),
        request.form.get("roadworthy_status") or "Not Checked",
        request.form.get("service_due_date"),
        request.form.get("service_history"),
    ))
    vehicle_id = cursor.lastrowid
    conn.commit()
    conn.close()

    log_action("Quick vehicle added", "vehicle", vehicle_id, stock_no)
    flash(f"{stock_no} added successfully.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/photos/featured/<int:photo_id>", methods=["POST"])
@login_required
def set_featured_photo(vehicle_id, photo_id):
    conn = db()
    photo = conn.execute(
        "SELECT * FROM vehicle_photos WHERE id=? AND vehicle_id=?",
        (photo_id, vehicle_id),
    ).fetchone()
    if not photo:
        conn.close()
        flash("Photo not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute(
        "UPDATE vehicles SET featured_photo_id=?,photo_filename=? WHERE id=?",
        (photo_id, photo["filename"], vehicle_id),
    )
    conn.commit()
    conn.close()
    log_action("Featured photo changed", "vehicle", vehicle_id, photo["filename"])
    flash("Featured photo updated.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/photos/delete/<int:photo_id>", methods=["POST"])
@login_required
def delete_vehicle_photo(vehicle_id, photo_id):
    conn = db()
    photo = conn.execute(
        "SELECT * FROM vehicle_photos WHERE id=? AND vehicle_id=?",
        (photo_id, vehicle_id),
    ).fetchone()
    if not photo:
        conn.close()
        flash("Photo not found.", "error")
        return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))

    conn.execute("DELETE FROM vehicle_photos WHERE id=?", (photo_id,))
    vehicle = conn.execute("SELECT featured_photo_id FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if vehicle and vehicle["featured_photo_id"] == photo_id:
        replacement = conn.execute(
            "SELECT id,filename FROM vehicle_photos WHERE vehicle_id=? ORDER BY id DESC LIMIT 1",
            (vehicle_id,),
        ).fetchone()
        conn.execute(
            "UPDATE vehicles SET featured_photo_id=?,photo_filename=? WHERE id=?",
            (replacement["id"] if replacement else None, replacement["filename"] if replacement else None, vehicle_id),
        )
    conn.commit()
    conn.close()

    try:
        file_path = UPLOAD_DIR / photo["filename"]
        if file_path.exists():
            file_path.unlink()
    except OSError:
        pass

    log_action("Vehicle photo deleted", "vehicle", vehicle_id, photo["filename"])
    flash("Photo deleted.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/valuation", methods=["GET", "POST"])
@login_required
def vehicle_valuation(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    expenses = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM expenses WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    jobs = conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) AS v FROM job_cards WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    services = conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) AS v FROM service_entries WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]
    parts = conn.execute("SELECT COALESCE(SUM(quantity_used*unit_cost_inc_gst),0) AS v FROM part_usage WHERE vehicle_id=?", (vehicle_id,)).fetchone()["v"]

    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    if request.method == "POST":
        conn.execute("""
            UPDATE vehicles
            SET estimated_sale_price=?,minimum_sale_price=?,valuation_notes=?
            WHERE id=?
        """, (
            float(request.form.get("estimated_sale_price") or 0),
            float(request.form.get("minimum_sale_price") or 0),
            request.form.get("valuation_notes"),
            vehicle_id,
        ))
        conn.commit()
        vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
        log_action("Vehicle valuation updated", "vehicle", vehicle_id, request.form.get("valuation_notes"))
        flash("Valuation saved.", "success")

    conn.close()
    total_cost = vehicle["purchase_price_inc_gst"] + expenses + jobs + services + parts
    estimated_profit = vehicle["estimated_sale_price"] - total_cost
    minimum_profit = vehicle["minimum_sale_price"] - total_cost
    return render_template(
        "vehicle_valuation.html",
        vehicle=vehicle,
        expenses=expenses,
        jobs=jobs,
        services=services,
        parts=parts,
        total_cost=total_cost,
        estimated_profit=estimated_profit,
        minimum_profit=minimum_profit,
    )


@app.route("/vehicles/<int:vehicle_id>/purchase-agreement")
@login_required
def purchase_agreement(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404
    return render_template("purchase_agreement.html", vehicle=vehicle)


@app.route("/vehicles/<int:vehicle_id>/receipt")
@login_required
def vehicle_receipt(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    sale = conn.execute("SELECT * FROM sales WHERE vehicle_id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle or not sale:
        return "Receipt is available after a sale is recorded.", 404
    balance_paid = sale["sale_price_inc_gst"] - sale["deposit_amount"] - sale["trade_in_value"]
    return render_template(
        "vehicle_receipt.html",
        vehicle=vehicle,
        sale=sale,
        balance_paid=balance_paid,
    )


@app.route("/vehicles/<int:vehicle_id>/reminders", methods=["POST"])
@login_required
def add_reminder(vehicle_id):
    conn = db()
    conn.execute("""
        INSERT INTO reminders(vehicle_id,reminder_date,reminder_type,title,notes)
        VALUES(?,?,?,?,?)
    """, (
        vehicle_id,
        request.form.get("reminder_date"),
        request.form.get("reminder_type"),
        request.form.get("title"),
        request.form.get("notes"),
    ))
    conn.commit()
    conn.close()
    flash("Reminder added.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/reminders/<int:reminder_id>/complete", methods=["POST"])
@login_required
def complete_reminder(reminder_id):
    conn = db()
    reminder = conn.execute("SELECT * FROM reminders WHERE id=?", (reminder_id,)).fetchone()
    if not reminder:
        conn.close()
        return "Reminder not found", 404
    conn.execute("UPDATE reminders SET completed=1 WHERE id=?", (reminder_id,))
    conn.commit()
    vehicle_id = reminder["vehicle_id"]
    conn.close()
    flash("Reminder completed.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/reminders")
@login_required
def reminders_page():
    conn = db()
    rows = conn.execute("""
        SELECT r.*,v.stock_no,v.make,v.model
        FROM reminders r
        LEFT JOIN vehicles v ON v.id=r.vehicle_id
        ORDER BY r.completed,r.reminder_date,r.id
    """).fetchall()
    conn.close()
    return render_template("reminders.html", rows=rows)


@app.route("/tasks", methods=["GET", "POST"])
@login_required
def tasks_page():
    conn = db()
    if request.method == "POST":
        conn.execute("""
            INSERT INTO tasks(
                vehicle_id,task_date,due_date,title,category,assigned_to,priority,status,notes
            ) VALUES(?,?,?,?,?,?,?,?,?)
        """, (
            int(request.form.get("vehicle_id")) if request.form.get("vehicle_id") else None,
            request.form.get("task_date"),
            request.form.get("due_date"),
            request.form.get("title"),
            request.form.get("category"),
            request.form.get("assigned_to"),
            request.form.get("priority") or "Normal",
            request.form.get("status") or "Open",
            request.form.get("notes"),
        ))
        conn.commit()
        conn.close()
        flash("Task added.", "success")
        return redirect(url_for("tasks_page"))

    rows = conn.execute("""
        SELECT t.*,v.stock_no,v.make,v.model
        FROM tasks t
        LEFT JOIN vehicles v ON v.id=t.vehicle_id
        ORDER BY CASE WHEN t.status='Completed' THEN 1 ELSE 0 END,
                 CASE t.priority WHEN 'Urgent' THEN 1 WHEN 'High' THEN 2 ELSE 3 END,
                 COALESCE(t.due_date,'9999-12-31'),t.id DESC
    """).fetchall()
    vehicles = conn.execute("""
        SELECT id,stock_no,make,model FROM vehicles
        WHERE status NOT IN ('Sold','BER')
        ORDER BY stock_no
    """).fetchall()
    conn.close()
    return render_template("tasks.html", rows=rows, vehicles=vehicles)


@app.route("/tasks/<int:task_id>/complete", methods=["POST"])
@login_required
def complete_task(task_id):
    conn = db()
    conn.execute("UPDATE tasks SET status='Completed' WHERE id=?", (task_id,))
    conn.commit()
    conn.close()
    flash("Task completed.", "success")
    return redirect(request.referrer or url_for("tasks_page"))


@app.route("/vehicles/<int:vehicle_id>/prepare-for-sale", methods=["POST"])
@login_required
def prepare_for_sale(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close()
        return "Vehicle not found", 404

    today_text = date.today().isoformat()
    tasks = [
        ("Workshop", "Complete mechanical inspection", "Barry", "High"),
        ("Compliance", "Confirm PPSR and registration details", "Barry", "High"),
        ("Presentation", "Complete detailing and final clean", "Matt", "Normal"),
        ("Photography", "Upload complete vehicle photo set", "Matt", "Normal"),
        ("Advertising", "Create advertisement and window card", "Barry", "Normal"),
        ("Sales", "Confirm asking price and minimum sale price", "Barry", "High"),
    ]
    for category, title, assigned_to, priority in tasks:
        exists = conn.execute("""
            SELECT id FROM tasks
            WHERE vehicle_id=? AND title=? AND status!='Completed'
        """, (vehicle_id, title)).fetchone()
        if not exists:
            conn.execute("""
                INSERT INTO tasks(vehicle_id,task_date,due_date,title,category,assigned_to,priority,status)
                VALUES(?,?,?,?,?,?,?,'Open')
            """, (
                vehicle_id,
                today_text,
                (date.today() + timedelta(days=7)).isoformat(),
                title,
                category,
                assigned_to,
                priority,
            ))

    conn.execute(
        "UPDATE vehicles SET status='Being Repaired' WHERE id=? AND status='In Stock'",
        (vehicle_id,),
    )
    conn.commit()
    conn.close()
    log_action("Prepare for sale workflow created", "vehicle", vehicle_id, vehicle["stock_no"])
    flash("Prepare-for-sale tasks created.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/finance-calculator", methods=["GET", "POST"])
@login_required
def finance_calculator():
    result = None
    values = {
        "vehicle_price": request.form.get("vehicle_price", ""),
        "deposit": request.form.get("deposit", ""),
        "trade_in": request.form.get("trade_in", ""),
        "annual_rate": request.form.get("annual_rate", ""),
        "term_years": request.form.get("term_years", "5"),
        "balloon": request.form.get("balloon", "0"),
    }
    if request.method == "POST":
        price = float(values["vehicle_price"] or 0)
        deposit = float(values["deposit"] or 0)
        trade_in = float(values["trade_in"] or 0)
        annual_rate = float(values["annual_rate"] or 0) / 100
        years = int(float(values["term_years"] or 0))
        balloon = float(values["balloon"] or 0)
        principal = max(price - deposit - trade_in, 0)
        months = max(years * 12, 1)
        monthly_rate = annual_rate / 12

        financed_before_balloon = max(principal - balloon, 0)
        if monthly_rate > 0:
            monthly_payment = financed_before_balloon * (
                monthly_rate * (1 + monthly_rate) ** months
            ) / ((1 + monthly_rate) ** months - 1)
        else:
            monthly_payment = financed_before_balloon / months

        total_monthly = monthly_payment * months
        total_repaid = total_monthly + balloon
        total_interest = total_repaid - principal
        result = {
            "principal": principal,
            "monthly": monthly_payment,
            "weekly": monthly_payment * 12 / 52,
            "fortnightly": monthly_payment * 12 / 26,
            "total_repaid": total_repaid,
            "total_interest": total_interest,
            "balloon": balloon,
        }

    return render_template("finance_calculator.html", result=result, values=values)


@app.route("/reports/stock-age")
@login_required
def stock_age_report():
    conn = db()
    today_text = date.today().isoformat()
    rows = conn.execute("""
        SELECT v.*,
          CAST(julianday(?) - julianday(COALESCE(v.purchase_date,substr(v.created_at,1,10))) AS INTEGER) AS days_in_stock,
          COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) AS expenses,
          COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END)
                    FROM job_cards j WHERE j.vehicle_id=v.id),0) AS jobs
        FROM vehicles v
        WHERE v.status NOT IN ('Sold','BER')
        ORDER BY days_in_stock DESC
    """, (today_text,)).fetchall()
    conn.close()
    return render_template("stock_age_report.html", rows=rows)


@app.route("/vehicles/<int:vehicle_id>/advertisement-pro")
@login_required
def advertisement_pro(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    photos = conn.execute(
        "SELECT * FROM vehicle_photos WHERE vehicle_id=? ORDER BY id",
        (vehicle_id,),
    ).fetchall()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404

    title = vehicle["advertisement_title"] or (
        f"{vehicle['year'] or ''} {vehicle['make']} {vehicle['model']} {vehicle['variant'] or ''}".strip()
    )
    description = vehicle["advertisement_description"] or (
        f"{title}\n\n"
        f"• {vehicle['odometer_km'] or 0:,} km\n"
        f"• Registration: {vehicle['registration'] or 'Not listed'}\n"
        f"• Colour: {vehicle['colour'] or 'Not listed'}\n"
        f"• Roadworthy: {vehicle['roadworthy_status'] or 'Not checked'}\n\n"
        f"{vehicle['notes'] or ''}\n\n"
        "Contact BAM Motor Group — Buy • Sell • Trade."
    )
    marketplace = description
    carsales = description + "\n\nStock Number: " + vehicle["stock_no"]
    gumtree = description + "\n\nInspection by appointment."
    return render_template(
        "advertisement_pro.html",
        vehicle=vehicle,
        photos=photos,
        title=title,
        marketplace=marketplace,
        carsales=carsales,
        gumtree=gumtree,
    )

@app.route("/vehicles/<int:vehicle_id>/window-card")
@login_required
def window_card(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    conn.close()
    if not vehicle:
        return "Vehicle not found", 404
    return render_template("window_card.html", vehicle=vehicle)


@app.route("/shipping")
@login_required
def shipping_centre():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    conn = db()
    sql = """
        SELECT sh.*, p.part_number, p.part_name, p.vehicle_stock_no
        FROM part_shipments sh
        LEFT JOIN parts p ON p.id=sh.part_id
        WHERE 1=1
    """
    params = []
    if q:
        sql += " AND (sh.shipment_number LIKE ? OR sh.customer_name LIKE ? OR sh.tracking_number LIKE ? OR p.part_number LIKE ? OR p.part_name LIKE ?)"
        params.extend([f"%{q}%"] * 5)
    if status:
        sql += " AND sh.status=?"
        params.append(status)
    sql += " ORDER BY sh.id DESC"
    shipments = conn.execute(sql, params).fetchall()
    metrics = conn.execute("""
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN status='Ready to Pack' THEN 1 ELSE 0 END) AS ready,
               SUM(CASE WHEN status='Sent' THEN 1 ELSE 0 END) AS sent,
               SUM(CASE WHEN status='Delivered' THEN 1 ELSE 0 END) AS delivered,
               COALESCE(SUM(freight_charged-shipping_cost),0) AS freight_margin
        FROM part_shipments
    """).fetchone()
    conn.close()
    return render_template("shipping.html", shipments=shipments, metrics=metrics, q=q, status=status)


@app.route("/shipping/new", methods=["GET", "POST"])
@login_required
def shipping_new():
    conn = db()
    parts = conn.execute("""
        SELECT id,part_number,part_name,quantity_on_hand,vehicle_stock_no
        FROM parts WHERE quantity_on_hand>0 ORDER BY part_name
    """).fetchall()
    suggested_number = next_shipment_number(conn)
    selected_part_id = request.args.get("part_id", type=int)

    if request.method == "POST":
        part_id = request.form.get("part_id", type=int)
        quantity = float(request.form.get("quantity") or 1)
        deduct_stock = request.form.get("deduct_stock") == "1"
        if quantity <= 0:
            flash("Shipment quantity must be greater than zero.", "error")
        else:
            part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone() if part_id else None
            if deduct_stock and part and quantity > float(part["quantity_on_hand"] or 0):
                flash("Not enough part stock is available for this shipment.", "error")
            else:
                shipment_number = request.form.get("shipment_number", "").strip() or next_shipment_number(conn)
                cursor = conn.execute("""
                    INSERT INTO part_shipments(
                        shipment_number,part_id,quantity,customer_name,customer_phone,customer_email,
                        address_line,suburb,state,postcode,courier,tracking_number,parcel_weight_kg,
                        parcel_length_cm,parcel_width_cm,parcel_height_cm,shipping_cost,freight_charged,
                        status,date_sent,invoice_included,bubble_wrapped,box_sealed,tracking_sent,
                        stock_adjusted,notes
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    shipment_number, part_id, quantity, request.form.get("customer_name"),
                    request.form.get("customer_phone"), request.form.get("customer_email"),
                    request.form.get("address_line"), request.form.get("suburb"),
                    request.form.get("state"), request.form.get("postcode"),
                    request.form.get("courier"), request.form.get("tracking_number"),
                    float(request.form.get("parcel_weight_kg") or 0),
                    float(request.form.get("parcel_length_cm") or 0),
                    float(request.form.get("parcel_width_cm") or 0),
                    float(request.form.get("parcel_height_cm") or 0),
                    float(request.form.get("shipping_cost") or 0),
                    float(request.form.get("freight_charged") or 0),
                    request.form.get("status") or "Ready to Pack",
                    request.form.get("date_sent") or None,
                    1 if request.form.get("invoice_included") else 0,
                    1 if request.form.get("bubble_wrapped") else 0,
                    1 if request.form.get("box_sealed") else 0,
                    1 if request.form.get("tracking_sent") else 0,
                    1 if deduct_stock and part else 0,
                    request.form.get("notes"),
                ))
                if deduct_stock and part:
                    conn.execute(
                        "UPDATE parts SET quantity_on_hand=quantity_on_hand-? WHERE id=?",
                        (quantity, part_id),
                    )
                conn.commit()
                shipment_id = cursor.lastrowid
                conn.close()
                log_action("Part shipment created", "shipment", shipment_id, shipment_number)
                flash(f"Shipment {shipment_number} created.", "success")
                return redirect(url_for("shipping_detail", shipment_id=shipment_id))

    conn.close()
    return render_template(
        "shipping_form.html", parts=parts, suggested_number=suggested_number,
        selected_part_id=selected_part_id
    )


@app.route("/shipping/<int:shipment_id>", methods=["GET", "POST"])
@login_required
def shipping_detail(shipment_id):
    conn = db()
    if request.method == "POST":
        status = request.form.get("status") or "Ready to Pack"
        date_sent = request.form.get("date_sent") or None
        date_delivered = request.form.get("date_delivered") or None
        conn.execute("""
            UPDATE part_shipments SET
              courier=?,tracking_number=?,status=?,date_sent=?,date_delivered=?,
              invoice_included=?,bubble_wrapped=?,box_sealed=?,tracking_sent=?,notes=?
            WHERE id=?
        """, (
            request.form.get("courier"), request.form.get("tracking_number"), status,
            date_sent, date_delivered,
            1 if request.form.get("invoice_included") else 0,
            1 if request.form.get("bubble_wrapped") else 0,
            1 if request.form.get("box_sealed") else 0,
            1 if request.form.get("tracking_sent") else 0,
            request.form.get("notes"), shipment_id,
        ))
        conn.commit()
        flash("Shipment updated.", "success")
    shipment = conn.execute("""
        SELECT sh.*,p.part_number,p.part_name,p.vehicle_stock_no
        FROM part_shipments sh LEFT JOIN parts p ON p.id=sh.part_id
        WHERE sh.id=?
    """, (shipment_id,)).fetchone()
    conn.close()
    if not shipment:
        return "Shipment not found", 404
    return render_template("shipping_detail.html", shipment=shipment)


@app.route("/shipping/<int:shipment_id>/label")
@login_required
def shipping_label(shipment_id):
    conn = db()
    shipment = conn.execute("""
        SELECT sh.*,p.part_number,p.part_name,p.vehicle_stock_no
        FROM part_shipments sh LEFT JOIN parts p ON p.id=sh.part_id
        WHERE sh.id=?
    """, (shipment_id,)).fetchone()
    conn.close()
    if not shipment:
        return "Shipment not found", 404
    return render_template("shipping_label.html", shipment=shipment)


@app.route("/iphone-access")
@login_required
def iphone_access():
    ip = local_network_ip()
    return render_template("iphone_access.html", network_ip=ip, mobile_url=f"http://{ip}:5000")



def next_equipment_number(conn=None):
    own_conn = conn is None
    conn = conn or db()
    highest = 0
    for row in conn.execute("SELECT equipment_no FROM equipment WHERE equipment_no LIKE 'EQ-%'").fetchall():
        match = re.search(r"(\d+)$", row["equipment_no"] or "")
        if match:
            highest = max(highest, int(match.group(1)))
    if own_conn:
        conn.close()
    return f"EQ-{highest + 1:05d}"


@app.route("/equipment")
@login_required
def equipment_list():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    conn = db()
    where = []
    params = []
    if q:
        where.append("(equipment_no LIKE ? OR name LIKE ? OR brand LIKE ? OR model LIKE ? OR serial_number LIKE ? OR category LIKE ? OR location LIKE ? OR assigned_to LIKE ?)")
        params.extend([f"%{q}%"] * 8)
    if status:
        where.append("status=?")
        params.append(status)
    sql = "SELECT * FROM equipment" + ((" WHERE " + " AND ".join(where)) if where else "") + " ORDER BY name, equipment_no"
    rows = conn.execute(sql, params).fetchall()
    metrics = conn.execute("""
        SELECT COUNT(*) total,
               COALESCE(SUM(current_value),0) total_value,
               SUM(CASE WHEN status='Checked Out' THEN 1 ELSE 0 END) checked_out,
               SUM(CASE WHEN next_service_date IS NOT NULL AND next_service_date!='' AND next_service_date<=date('now','+30 day') THEN 1 ELSE 0 END) due_service
        FROM equipment
    """).fetchone()
    conn.close()
    return render_template("equipment.html", equipment=rows, metrics=metrics, q=q, selected_status=status)


@app.route("/equipment/new", methods=["GET", "POST"])
@login_required
def equipment_new():
    if request.method == "POST":
        try:
            photo = save_upload(request.files.get("photo"))
            receipt = save_upload(request.files.get("receipt"))
            conn = db()
            equipment_no = request.form.get("equipment_no", "").strip() or next_equipment_number(conn)
            cursor = conn.execute("""
                INSERT INTO equipment(
                    equipment_no,name,category,brand,model,serial_number,purchase_date,purchase_price,current_value,
                    supplier,warranty_expiry,location,assigned_to,condition,status,next_service_date,calibration_due,
                    test_tag_due,photo_filename,receipt_filename,notes
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                equipment_no, request.form.get("name", "").strip(), request.form.get("category"),
                request.form.get("brand"), request.form.get("model"), request.form.get("serial_number"),
                request.form.get("purchase_date"), float(request.form.get("purchase_price") or 0),
                float(request.form.get("current_value") or 0), request.form.get("supplier"),
                request.form.get("warranty_expiry"), request.form.get("location"), request.form.get("assigned_to"),
                request.form.get("condition") or "Good", request.form.get("status") or "Available",
                request.form.get("next_service_date"), request.form.get("calibration_due"), request.form.get("test_tag_due"),
                photo, receipt, request.form.get("notes")
            ))
            conn.execute("INSERT INTO equipment_history(equipment_id,action,assigned_to,location,condition,notes,user_name) VALUES(?,?,?,?,?,?,?)",
                         (cursor.lastrowid,"Equipment added",request.form.get("assigned_to"),request.form.get("location"),request.form.get("condition"),request.form.get("notes"),session.get("display_name")))
            conn.commit(); conn.close()
            flash("Equipment added.", "success")
            return redirect(url_for("equipment_detail", equipment_id=cursor.lastrowid))
        except (ValueError, sqlite3.IntegrityError) as exc:
            flash(str(exc), "error")
    conn=db(); suggested=next_equipment_number(conn); conn.close()
    return render_template("equipment_form.html", item=None, suggested_no=suggested)


@app.route("/equipment/<int:equipment_id>")
@login_required
def equipment_detail(equipment_id):
    conn=db()
    item=conn.execute("SELECT * FROM equipment WHERE id=?",(equipment_id,)).fetchone()
    history=conn.execute("SELECT * FROM equipment_history WHERE equipment_id=? ORDER BY action_date DESC,id DESC",(equipment_id,)).fetchall()
    conn.close()
    if not item: return "Equipment not found",404
    return render_template("equipment_detail.html", item=item, history=history)


@app.route("/equipment/<int:equipment_id>/edit", methods=["GET","POST"])
@login_required
def equipment_edit(equipment_id):
    conn=db(); item=conn.execute("SELECT * FROM equipment WHERE id=?",(equipment_id,)).fetchone()
    if not item:
        conn.close(); return "Equipment not found",404
    if request.method == "POST":
        try:
            photo=save_upload(request.files.get("photo")) or item["photo_filename"]
            receipt=save_upload(request.files.get("receipt")) or item["receipt_filename"]
            conn.execute("""UPDATE equipment SET equipment_no=?,name=?,category=?,brand=?,model=?,serial_number=?,purchase_date=?,purchase_price=?,current_value=?,supplier=?,warranty_expiry=?,location=?,assigned_to=?,condition=?,status=?,next_service_date=?,calibration_due=?,test_tag_due=?,photo_filename=?,receipt_filename=?,notes=? WHERE id=?""",(
                request.form.get("equipment_no"),request.form.get("name"),request.form.get("category"),request.form.get("brand"),request.form.get("model"),request.form.get("serial_number"),request.form.get("purchase_date"),float(request.form.get("purchase_price") or 0),float(request.form.get("current_value") or 0),request.form.get("supplier"),request.form.get("warranty_expiry"),request.form.get("location"),request.form.get("assigned_to"),request.form.get("condition"),request.form.get("status"),request.form.get("next_service_date"),request.form.get("calibration_due"),request.form.get("test_tag_due"),photo,receipt,request.form.get("notes"),equipment_id))
            conn.execute("INSERT INTO equipment_history(equipment_id,action,assigned_to,location,condition,notes,user_name) VALUES(?,?,?,?,?,?,?)",(equipment_id,"Equipment updated",request.form.get("assigned_to"),request.form.get("location"),request.form.get("condition"),"Details updated",session.get("display_name")))
            conn.commit(); conn.close(); flash("Equipment updated.","success")
            return redirect(url_for("equipment_detail",equipment_id=equipment_id))
        except (ValueError,sqlite3.IntegrityError) as exc:
            flash(str(exc),"error")
    conn.close()
    return render_template("equipment_form.html",item=item,suggested_no=item["equipment_no"])


@app.route("/equipment/<int:equipment_id>/movement", methods=["POST"])
@login_required
def equipment_movement(equipment_id):
    action=request.form.get("action") or "Updated"
    status="Checked Out" if action=="Checked Out" else ("Available" if action=="Returned" else request.form.get("status") or "Available")
    conn=db()
    conn.execute("UPDATE equipment SET assigned_to=?,location=?,condition=?,status=? WHERE id=?",(request.form.get("assigned_to"),request.form.get("location"),request.form.get("condition") or "Good",status,equipment_id))
    conn.execute("INSERT INTO equipment_history(equipment_id,action,assigned_to,location,condition,notes,user_name) VALUES(?,?,?,?,?,?,?)",(equipment_id,action,request.form.get("assigned_to"),request.form.get("location"),request.form.get("condition"),request.form.get("notes"),session.get("display_name")))
    conn.commit(); conn.close(); flash(f"Equipment {action.lower()}.","success")
    return redirect(url_for("equipment_detail",equipment_id=equipment_id))


@app.route("/users")
@owner_required
def users_page():
    conn=db(); users=conn.execute("SELECT * FROM users ORDER BY display_name,username").fetchall(); conn.close()
    return render_template("users.html", users=users)


@app.route("/users/new", methods=["GET","POST"])
@owner_required
def user_new():
    if request.method == "POST":
        username=(request.form.get("username") or "").strip().lower()
        password=request.form.get("password") or ""
        display_name=(request.form.get("display_name") or "").strip()
        role=request.form.get("role") or "staff"
        if not username or not display_name or len(password)<8:
            flash("Username, display name and a password of at least 8 characters are required.","error")
        else:
            try:
                conn=db(); conn.execute("INSERT INTO users(username,password_hash,display_name,role,is_active) VALUES(?,?,?,?,1)",(username,generate_password_hash(password),display_name,role)); conn.commit(); conn.close()
                flash("User created.","success"); return redirect(url_for("users_page"))
            except sqlite3.IntegrityError:
                flash("That username is already in use.","error")
    return render_template("user_form.html", user=None)


@app.route("/users/<int:user_id>/edit", methods=["GET","POST"])
@owner_required
def user_edit(user_id):
    conn=db(); user=conn.execute("SELECT * FROM users WHERE id=?",(user_id,)).fetchone()
    if not user: conn.close(); return "User not found",404
    if request.method == "POST":
        username=(request.form.get("username") or "").strip().lower(); display_name=(request.form.get("display_name") or "").strip(); role=request.form.get("role") or "staff"; active=1 if request.form.get("is_active") else 0; password=request.form.get("password") or ""
        if user_id==session.get("user_id") and not active:
            conn.close(); flash("You cannot disable your own account.","error"); return redirect(url_for("user_edit",user_id=user_id))
        try:
            if password:
                if len(password)<8: raise ValueError("Password must be at least 8 characters.")
                conn.execute("UPDATE users SET username=?,display_name=?,role=?,is_active=?,password_hash=? WHERE id=?",(username,display_name,role,active,generate_password_hash(password),user_id))
            else:
                conn.execute("UPDATE users SET username=?,display_name=?,role=?,is_active=? WHERE id=?",(username,display_name,role,active,user_id))
            conn.commit(); conn.close(); flash("User updated.","success"); return redirect(url_for("users_page"))
        except (sqlite3.IntegrityError,ValueError) as exc:
            conn.close(); flash(str(exc),"error")
    else: conn.close()
    return render_template("user_form.html", user=user)


def _workshop_provider_vehicle_url(vehicle):
    if not WORKSHOP_PORTAL_URL:
        return ""
    replacements = {
        "vin": urllib.parse.quote(str(vehicle["vin"] or "")),
        "make": urllib.parse.quote(str(vehicle["make"] or "")),
        "model": urllib.parse.quote(str(vehicle["model"] or "")),
        "year": urllib.parse.quote(str(vehicle["year"] or "")),
        "stock_no": urllib.parse.quote(str(vehicle["stock_no"] or "")),
    }
    url = WORKSHOP_PORTAL_URL
    for key, value in replacements.items():
        url = url.replace("{" + key + "}", value)
    return url


@app.route("/workshop-intelligence")
@login_required
def workshop_centre():
    q = (request.args.get("q") or "").strip()
    conn = db()
    params = []
    where = ""
    if q:
        like = f"%{q}%"
        where = "WHERE stock_no LIKE ? OR vin LIKE ? OR registration LIKE ? OR make LIKE ? OR model LIKE ?"
        params = [like] * 5
    vehicles = conn.execute(f"""
        SELECT id,stock_no,year,make,model,variant,vin,registration,status
        FROM vehicles {where}
        ORDER BY id DESC LIMIT 50
    """, params).fetchall()
    recent_jobs = conn.execute("""
        SELECT j.*,v.stock_no,v.make,v.model
        FROM job_cards j JOIN vehicles v ON v.id=j.vehicle_id
        ORDER BY j.id DESC LIMIT 12
    """).fetchall()
    conn.close()
    return render_template("workshop_centre.html", vehicles=vehicles, q=q, recent_jobs=recent_jobs,
                           default_labour_rate=DEFAULT_LABOUR_RATE)


@app.route("/vehicles/<int:vehicle_id>/workshop-intelligence")
@login_required
def workshop_intelligence(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found", 404
    operations = conn.execute("""
        SELECT o.*,p.part_number,p.part_name
        FROM workshop_operations o
        LEFT JOIN parts p ON p.id=o.part_id
        WHERE o.vehicle_id=? OR (
            o.vehicle_id IS NULL AND LOWER(COALESCE(o.make,''))=LOWER(COALESCE(?,''))
            AND LOWER(COALESCE(o.model,''))=LOWER(COALESCE(?,''))
            AND (o.year IS NULL OR o.year=0 OR o.year=?)
        )
        ORDER BY o.system_name,o.operation_name,o.id DESC
    """, (vehicle_id, vehicle["make"], vehicle["model"], vehicle["year"] or 0)).fetchall()
    refs = conn.execute("SELECT * FROM workshop_references WHERE vehicle_id=? ORDER BY reference_type,title", (vehicle_id,)).fetchall()
    parts = conn.execute("SELECT id,part_number,part_name,manufacturer_part_no,quantity_on_hand FROM parts WHERE quantity_on_hand>0 ORDER BY part_name").fetchall()
    jobs = conn.execute("SELECT * FROM job_cards WHERE vehicle_id=? ORDER BY id DESC LIMIT 20", (vehicle_id,)).fetchall()
    conn.close()
    specs = {}
    try:
        specs = json.loads(vehicle["vin_decode_json"] or "{}") if vehicle["vin_decode_json"] else {}
    except Exception:
        specs = {}
    manual_search_phrase = " ".join(str(x).strip() for x in [vehicle["year"], vehicle["make"], vehicle["model"], vehicle["variant"], "workshop service repair manual"] if x)
    return render_template("workshop_intelligence.html", vehicle=vehicle, operations=operations, references=refs,
                           parts=parts, jobs=jobs, provider_name=WORKSHOP_PROVIDER_NAME,
                           provider_url=_workshop_provider_vehicle_url(vehicle), default_labour_rate=DEFAULT_LABOUR_RATE,
                           decoded_specs=specs, manual_search_phrase=manual_search_phrase)


@app.route("/vehicles/<int:vehicle_id>/workshop-operation", methods=["POST"])
@login_required
def workshop_operation_add(vehicle_id):
    conn = db()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found", 404
    try:
        hours = float(request.form.get("labour_hours") or 0)
        rate = float(request.form.get("labour_rate") or DEFAULT_LABOUR_RATE)
        if hours < 0 or rate < 0:
            raise ValueError("Labour hours and labour rate cannot be negative.")
        part_id = int(request.form.get("part_id")) if request.form.get("part_id") else None
        operation_name = (request.form.get("operation_name") or "").strip()
        if not operation_name and part_id:
            part = conn.execute("SELECT part_name FROM parts WHERE id=?", (part_id,)).fetchone()
            operation_name = f"Fit {part['part_name']}" if part else "Fit part"
        if not operation_name:
            raise ValueError("Enter an operation name.")
        reusable = request.form.get("scope") == "library"
        conn.execute("""
            INSERT INTO workshop_operations(vehicle_id,part_id,operation_code,system_name,operation_name,labour_hours,labour_rate,
                source_name,source_reference,procedure_url,notes,make,model,year,engine_hint,created_by)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (None if reusable else vehicle_id, part_id, request.form.get("operation_code"), request.form.get("system_name"), operation_name,
              hours, rate, request.form.get("source_name"), request.form.get("source_reference"), request.form.get("procedure_url"),
              request.form.get("notes"), vehicle["make"], vehicle["model"], vehicle["year"],
              vehicle["engine_model"] or vehicle["engine_make"] or vehicle["engine_displacement_l"], session.get("display_name")))
        conn.commit()
        log_action("Workshop labour operation added", "vehicle", vehicle_id, f"{operation_name} — {hours:.2f} h")
        flash("Workshop labour operation saved.", "success")
    except (ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc), "error")
    finally:
        conn.close()
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/workshop-operation/<int:operation_id>/job", methods=["POST"])
@login_required
def workshop_operation_to_job(vehicle_id, operation_id):
    conn = db()
    op = conn.execute("SELECT * FROM workshop_operations WHERE id=?", (operation_id,)).fetchone()
    vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not op or not vehicle:
        conn.close(); return "Workshop operation or vehicle not found", 404
    hours = float(op["labour_hours"] or 0)
    rate = float(op["labour_rate"] or DEFAULT_LABOUR_RATE)
    estimate = round(hours * rate, 2)
    notes = "\n".join(x for x in [
        f"Labour source: {op['source_name']}" if op["source_name"] else "",
        f"Reference: {op['source_reference']}" if op["source_reference"] else "",
        op["notes"] or "",
    ] if x)
    cur = conn.execute("""
        INSERT INTO job_cards(vehicle_id,job_date,category,description,supplier,paid_by,estimated_cost,actual_cost_inc_gst,
            gst_amount,status,notes,labour_operation_id,labour_hours,labour_rate,labour_source,labour_code,procedure_url)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (vehicle_id, date.today().isoformat(), op["system_name"] or "Mechanical", op["operation_name"], "", "Business", estimate, 0, 0,
          "Open", notes, op["id"], hours, rate, op["source_name"], op["operation_code"], op["procedure_url"]))
    job_id = cur.lastrowid
    conn.execute("INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note) VALUES(?,?,?,?,?)",
                 (job_id, None, "Open", session.get("display_name"), "Created from Workshop Intelligence labour operation"))
    conn.commit(); conn.close()
    log_action("Workshop operation added to job card", "vehicle", vehicle_id, f"Job #{job_id}; {hours:.2f} h @ ${rate:.2f}")
    flash(f"Job card #{job_id} created with {hours:.2f} labour hours.", "success")
    return redirect(url_for("vehicle_detail", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/workshop-operation/<int:operation_id>/delete", methods=["POST"])
@login_required
def workshop_operation_delete(vehicle_id, operation_id):
    conn = db()
    conn.execute("DELETE FROM workshop_operations WHERE id=? AND (vehicle_id=? OR vehicle_id IS NULL)", (operation_id, vehicle_id))
    conn.commit(); conn.close()
    flash("Workshop labour operation removed.", "success")
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/workshop-reference", methods=["POST"])
@login_required
def workshop_reference_add(vehicle_id):
    title = (request.form.get("title") or "").strip()
    if not title:
        flash("Reference title is required.", "error")
        return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))
    conn = db()
    conn.execute("""
        INSERT INTO workshop_references(vehicle_id,reference_type,title,provider_name,reference_url,reference_code,notes,created_by)
        VALUES(?,?,?,?,?,?,?,?)
    """, (vehicle_id, request.form.get("reference_type") or "Workshop Manual", title, request.form.get("provider_name"),
          request.form.get("reference_url"), request.form.get("reference_code"), request.form.get("notes"), session.get("display_name")))
    conn.commit(); conn.close()
    flash("Workshop technical reference saved.", "success")
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/workshop-reference/<int:reference_id>/delete", methods=["POST"])
@login_required
def workshop_reference_delete(vehicle_id, reference_id):
    conn = db(); conn.execute("DELETE FROM workshop_references WHERE id=? AND vehicle_id=?", (reference_id, vehicle_id)); conn.commit(); conn.close()
    flash("Workshop reference removed.", "success")
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))


@app.route("/vehicles/<int:vehicle_id>/workshop-labour-import", methods=["POST"])
@login_required
def workshop_labour_import(vehicle_id):
    upload = request.files.get("labour_csv")
    if not upload or not upload.filename:
        flash("Choose a CSV file first.", "error")
        return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))
    conn = db(); vehicle = conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found", 404
    try:
        raw = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw))
        required = {"operation_name", "labour_hours"}
        if not reader.fieldnames or not required.issubset({x.strip() for x in reader.fieldnames}):
            raise ValueError("CSV must include operation_name and labour_hours columns.")
        imported = 0
        for row in reader:
            name = (row.get("operation_name") or "").strip()
            if not name: continue
            hours = float(row.get("labour_hours") or 0)
            rate = float(row.get("labour_rate") or DEFAULT_LABOUR_RATE)
            conn.execute("""
                INSERT INTO workshop_operations(vehicle_id,operation_code,system_name,operation_name,labour_hours,labour_rate,
                    source_name,source_reference,procedure_url,notes,make,model,year,engine_hint,created_by)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (vehicle_id, row.get("operation_code"), row.get("system_name"), name, hours, rate, row.get("source_name"),
                  row.get("source_reference"), row.get("procedure_url"), row.get("notes"), vehicle["make"], vehicle["model"],
                  vehicle["year"], vehicle["engine_model"] or vehicle["engine_make"] or vehicle["engine_displacement_l"], session.get("display_name")))
            imported += 1
        conn.commit(); flash(f"Imported {imported} workshop labour operation(s).", "success")
    except (UnicodeDecodeError, ValueError, sqlite3.Error) as exc:
        conn.rollback(); flash(f"Labour import failed: {exc}", "error")
    finally:
        conn.close()
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))


@app.route("/email-centre", methods=["GET", "POST"])
@login_required
def email_centre():
    conn = db()
    if request.method == "POST":
        subject = (request.form.get("subject") or "").strip()
        if not subject:
            flash("Email subject is required.", "error")
        else:
            try:
                conn.execute("""
                    INSERT INTO email_messages(direction,message_date,sender,recipient,subject,body_excerpt,status,
                        follow_up_date,vehicle_id,contact_id,external_reference,created_by)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    request.form.get("direction") or "Incoming",
                    request.form.get("message_date") or date.today().isoformat(),
                    request.form.get("sender"), request.form.get("recipient"), subject,
                    request.form.get("body_excerpt"), request.form.get("status") or "Unread",
                    request.form.get("follow_up_date"),
                    int(request.form.get("vehicle_id")) if request.form.get("vehicle_id") else None,
                    int(request.form.get("contact_id")) if request.form.get("contact_id") else None,
                    request.form.get("external_reference"), session.get("display_name")
                ))
                conn.commit()
                flash("Email correspondence saved.", "success")
                return redirect(url_for("email_centre"))
            except (ValueError, sqlite3.Error) as exc:
                conn.rollback(); flash(f"Could not save email: {exc}", "error")
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    params=[]; where=[]
    if q:
        like=f"%{q}%"; where.append("(em.subject LIKE ? OR em.sender LIKE ? OR em.recipient LIKE ? OR em.body_excerpt LIKE ? OR v.stock_no LIKE ? OR c.name LIKE ?)"); params += [like]*6
    if status:
        where.append("em.status=?"); params.append(status)
    sql="""
        SELECT em.*, v.stock_no, v.make, v.model, c.name AS contact_name
        FROM email_messages em
        LEFT JOIN vehicles v ON v.id=em.vehicle_id
        LEFT JOIN contacts c ON c.id=em.contact_id
    """
    if where: sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY COALESCE(em.message_date, substr(em.created_at,1,10)) DESC, em.id DESC LIMIT 250"
    messages=conn.execute(sql, params).fetchall()
    vehicles=conn.execute("SELECT id,stock_no,year,make,model FROM vehicles ORDER BY stock_no DESC").fetchall()
    contacts=conn.execute("SELECT id,name,email FROM contacts ORDER BY name").fetchall()
    metrics=conn.execute("SELECT COUNT(*) total, SUM(CASE WHEN status='Unread' THEN 1 ELSE 0 END) unread, SUM(CASE WHEN status='Follow Up' THEN 1 ELSE 0 END) follow_up FROM email_messages").fetchone()
    conn.close()
    return render_template("email_centre.html", messages=messages, vehicles=vehicles, contacts=contacts, metrics=metrics,
                           email_webmail_url=EMAIL_WEBMAIL_URL, email_address=EMAIL_ADDRESS, q=q, status_filter=status,
                           today=date.today().isoformat())


@app.route("/email-centre/<int:message_id>/status", methods=["POST"])
@login_required
def email_message_status(message_id):
    new_status = request.form.get("status") or "Read"
    if new_status not in {"Unread","Read","Follow Up","Completed"}: new_status="Read"
    conn=db(); conn.execute("UPDATE email_messages SET status=? WHERE id=?", (new_status,message_id)); conn.commit(); conn.close()
    flash("Email status updated.", "success")
    return redirect(request.referrer or url_for("email_centre"))


@app.route("/email-centre/<int:message_id>/delete", methods=["POST"])
@login_required
def email_message_delete(message_id):
    conn=db(); conn.execute("DELETE FROM email_messages WHERE id=?", (message_id,)); conn.commit(); conn.close()
    flash("Email record removed.", "success")
    return redirect(url_for("email_centre"))


@app.route("/vehicles/<int:vehicle_id>/emanual-link", methods=["POST"])
@login_required
def emanual_link_add(vehicle_id):
    link=(request.form.get("reference_url") or "").strip()
    title=(request.form.get("title") or "eManualOnline Workshop Manual").strip()
    if not link:
        flash("Paste the purchased eManualOnline manual link first.", "error")
        return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))
    conn=db()
    vehicle=conn.execute("SELECT * FROM vehicles WHERE id=?", (vehicle_id,)).fetchone()
    if not vehicle:
        conn.close(); return "Vehicle not found",404
    conn.execute("""
        INSERT INTO workshop_references(vehicle_id,reference_type,title,provider_name,reference_url,reference_code,notes,created_by)
        VALUES(?,?,?,?,?,?,?,?)
    """, (vehicle_id,"Workshop Manual",title,"eManualOnline",link,vehicle["vin"],"Saved eManualOnline vehicle manual",session.get("display_name")))
    conn.commit(); conn.close()
    flash("eManualOnline manual linked to this vehicle.", "success")
    return redirect(url_for("workshop_intelligence", vehicle_id=vehicle_id))



# ---------------------------------------------------------------------------
# Version 24 - Intelligent Dealer Platform
# Safe local Dealer Copilot + Executive Dashboard.
# The Copilot only runs predefined, read-only queries against BAM's own data.
# ---------------------------------------------------------------------------
def _copilot_money(value):
    try:
        return f"${float(value or 0):,.2f}"
    except Exception:
        return "$0.00"


def dealer_copilot_answer(conn, question):
    q = (question or "").strip()
    low = q.lower()
    result = {"question": q, "summary": "", "items": [], "kind": "general", "tips": []}
    if not q:
        result["summary"] = "Ask about vehicles, parts, workshop jobs, customers, invoices, reminders or business performance."
        return result

    # Exact stock / VIN / registration lookup first.
    vehicle = conn.execute("""
        SELECT id,stock_no,year,make,model,status,vin,registration,asking_price,purchase_price_inc_gst
        FROM vehicles
        WHERE UPPER(COALESCE(stock_no,''))=UPPER(?) OR UPPER(COALESCE(vin,''))=UPPER(?) OR UPPER(COALESCE(registration,''))=UPPER(?)
        LIMIT 1
    """, (q,q,q)).fetchone()
    if vehicle:
        result["kind"] = "vehicle"
        result["summary"] = f"Found {vehicle['stock_no']} — {vehicle['year'] or ''} {vehicle['make']} {vehicle['model']}."
        result["items"] = [{"title": f"{vehicle['stock_no']} — {vehicle['year'] or ''} {vehicle['make']} {vehicle['model']}", "subtitle": f"{vehicle['status']} · VIN {vehicle['vin'] or 'not recorded'} · Rego {vehicle['registration'] or 'not recorded'}", "amount": _copilot_money(vehicle['asking_price']), "url": url_for('vehicle_detail', vehicle_id=vehicle['id'])}]
        return result

    if ("profit" in low or "profitable" in low) and ("vehicle" in low or "car" in low or "stock" in low):
        rows = conn.execute("""
            SELECT v.id,v.stock_no,v.year,v.make,v.model,
              COALESCE((SELECT s.sale_price_inc_gst FROM sales s WHERE s.vehicle_id=v.id),0)
              + COALESCE((SELECT SUM(ps.sale_price) FROM part_sales ps JOIN parts p ON p.id=ps.part_id WHERE p.vehicle_id=v.id),0)
              + COALESCE(v.shell_sale_price,0) AS revenue,
              COALESCE(v.purchase_price_inc_gst,0)
              + COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0)
              + COALESCE((SELECT SUM(CASE WHEN j.actual_cost_inc_gst>0 THEN j.actual_cost_inc_gst ELSE j.estimated_cost END) FROM job_cards j WHERE j.vehicle_id=v.id),0)
              + COALESCE((SELECT SUM(se.cost_inc_gst) FROM service_entries se WHERE se.vehicle_id=v.id),0) AS invested
            FROM vehicles v
            ORDER BY (revenue-invested) DESC
            LIMIT 10
        """).fetchall()
        result["kind"] = "profit"
        result["summary"] = "Most profitable vehicles based on recorded vehicle sales, donor-part sales and shell revenue less recorded investment and costs."
        for r in rows:
            profit=float(r['revenue'] or 0)-float(r['invested'] or 0)
            result["items"].append({"title": f"{r['stock_no']} — {r['year'] or ''} {r['make']} {r['model']}", "subtitle": f"Revenue {_copilot_money(r['revenue'])} · Invested {_copilot_money(r['invested'])}", "amount": _copilot_money(profit), "url": url_for('vehicle_detail', vehicle_id=r['id'])})
        return result

    days_match = re.search(r"(\d+)\s*day", low)
    days = int(days_match.group(1)) if days_match else 90
    if ("sitting" in low or "stock age" in low or "days in stock" in low or "old stock" in low):
        today_text=date.today().isoformat()
        rows=conn.execute("""
            SELECT id,stock_no,year,make,model,status,purchase_date,
                   CAST(julianday(?) - julianday(COALESCE(purchase_date,substr(created_at,1,10))) AS INTEGER) AS days_in_stock
            FROM vehicles
            WHERE status NOT IN ('Sold','BER')
              AND (julianday(?) - julianday(COALESCE(purchase_date,substr(created_at,1,10)))) >= ?
            ORDER BY days_in_stock DESC
            LIMIT 30
        """,(today_text,today_text,days)).fetchall()
        result["kind"]="stock_age"
        result["summary"]=f"{len(rows)} active vehicle(s) have been in stock for at least {days} days."
        result["items"]=[{"title":f"{r['stock_no']} — {r['year'] or ''} {r['make']} {r['model']}","subtitle":f"{r['days_in_stock']} days · {r['status']}","amount":"","url":url_for('vehicle_detail',vehicle_id=r['id'])} for r in rows]
        return result

    if "photo" in low and ("missing" in low or "need" in low or "without" in low):
        rows=conn.execute("""
            SELECT id,stock_no,year,make,model,status FROM vehicles
            WHERE status NOT IN ('Sold','BER') AND COALESCE(photo_filename,'')=''
              AND NOT EXISTS (SELECT 1 FROM vehicle_photos p WHERE p.vehicle_id=vehicles.id)
            ORDER BY id DESC LIMIT 30
        """).fetchall()
        result["kind"]="photos"; result["summary"]=f"{len(rows)} active vehicle(s) are missing photos."
        result["items"]=[{"title":f"{r['stock_no']} — {r['year'] or ''} {r['make']} {r['model']}","subtitle":r['status'],"amount":"","url":url_for('vehicle_detail',vehicle_id=r['id'])} for r in rows]
        return result

    if "workshop" in low or "job" in low:
        overdue = "overdue" in low
        sql="""SELECT j.id,j.vehicle_id,j.description,j.status,j.job_date,v.stock_no,v.make,v.model
               FROM job_cards j LEFT JOIN vehicles v ON v.id=j.vehicle_id
               WHERE COALESCE(j.status,'Open')!='Completed'"""
        params=[]
        if overdue:
            sql += " AND COALESCE(j.job_date,'9999-12-31') < ?"; params.append(date.today().isoformat())
        sql += " ORDER BY j.job_date,j.id DESC LIMIT 30"
        rows=conn.execute(sql,params).fetchall()
        result["kind"]="workshop"; result["summary"]=("Overdue" if overdue else "Active")+f" workshop jobs: {len(rows)}."
        result["items"]=[{"title":r['description'] or 'Workshop job',"subtitle":f"{r['stock_no'] or 'No vehicle'} — {r['make'] or ''} {r['model'] or ''} · {r['status']}","amount":"","url":url_for('vehicle_detail',vehicle_id=r['vehicle_id']) if r['vehicle_id'] else url_for('workshop_centre')} for r in rows]
        return result

    if "invoice" in low and ("draft" in low or "outstanding" in low or "unpaid" in low):
        rows=conn.execute("""
            SELECT s.vehicle_id,s.invoice_number,s.invoice_status,s.buyer_name,s.sale_price_inc_gst,v.stock_no,v.make,v.model
            FROM sales s JOIN vehicles v ON v.id=s.vehicle_id
            WHERE COALESCE(s.invoice_status,'Draft')!='Paid'
            ORDER BY s.sale_date DESC LIMIT 30
        """).fetchall()
        result["kind"]="invoice"; result["summary"]=f"{len(rows)} invoice(s) are not marked Paid."
        result["items"]=[{"title":r['invoice_number'] or 'Draft invoice',"subtitle":f"{r['buyer_name'] or 'Buyer'} · {r['stock_no']} {r['make']} {r['model']} · {r['invoice_status'] or 'Draft'}","amount":_copilot_money(r['sale_price_inc_gst']),"url":url_for('sale_invoice',vehicle_id=r['vehicle_id'])} for r in rows]
        return result

    if "reminder" in low or "task" in low or "attention" in low:
        rows=conn.execute("""
            SELECT r.id,r.vehicle_id,r.reminder_date,r.title,r.notes,v.stock_no,v.make,v.model
            FROM reminders r LEFT JOIN vehicles v ON v.id=r.vehicle_id
            WHERE r.completed=0
            ORDER BY COALESCE(r.reminder_date,'9999-12-31'),r.id LIMIT 30
        """).fetchall()
        result["kind"]="reminders"; result["summary"]=f"{len(rows)} open reminder(s) need attention."
        result["items"]=[{"title":r['title'],"subtitle":f"{r['stock_no'] or ''} {r['make'] or ''} {r['model'] or ''} · Due {r['reminder_date'] or 'not set'}","amount":"","url":url_for('vehicle_detail',vehicle_id=r['vehicle_id']) if r['vehicle_id'] else url_for('reminders_page')} for r in rows]
        return result

    if "part" in low or "engine" in low or "gearbox" in low or "transmission" in low:
        stop={"show","me","all","find","parts","part","in","stock","the","a","an","for","from","do","we","have"}
        terms=[t for t in re.findall(r"[a-z0-9-]+",low) if t not in stop and len(t)>1]
        search="%"+"%".join(terms[:4])+"%" if terms else "%"
        rows=conn.execute("""
            SELECT id,part_number,part_name,category,quantity_on_hand,selling_price,storage_location,make,model,vehicle_stock_no,status
            FROM parts
            WHERE (LOWER(COALESCE(part_name,'')) LIKE ? OR LOWER(COALESCE(category,'')) LIKE ? OR LOWER(COALESCE(make,'')) LIKE ? OR LOWER(COALESCE(model,'')) LIKE ? OR LOWER(COALESCE(fitment,'')) LIKE ? OR LOWER(COALESCE(manufacturer_part_no,'')) LIKE ?)
            ORDER BY quantity_on_hand DESC,id DESC LIMIT 40
        """,(search,search,search,search,search,search)).fetchall()
        result["kind"]="parts"; result["summary"]=f"Found {len(rows)} matching part(s)."
        result["items"]=[{"title":f"{r['part_number'] or ''} — {r['part_name']}","subtitle":f"{r['make'] or ''} {r['model'] or ''} · Qty {r['quantity_on_hand'] or 0} · {r['storage_location'] or 'No location'} · {r['status'] or ''}","amount":_copilot_money(r['selling_price']),"url":url_for('part_detail',part_id=r['id'])} for r in rows]
        return result

    if any(k in low for k in ("sales this month","revenue this month","monthly sales","month profit","monthly profit")):
        month=date.today().strftime('%Y-%m')
        vehicle_rev=float(conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) v FROM sales WHERE substr(sale_date,1,7)=?",(month,)).fetchone()['v'] or 0)
        parts_rev=float(conn.execute("SELECT COALESCE(SUM(sale_price),0) v FROM part_sales WHERE substr(sale_date,1,7)=?",(month,)).fetchone()['v'] or 0)
        result["kind"]="money"; result["summary"]=f"Recorded revenue this month is {_copilot_money(vehicle_rev+parts_rev)}."
        result["items"]=[{"title":"Vehicle sales","subtitle":month,"amount":_copilot_money(vehicle_rev),"url":url_for('invoice_centre')},{"title":"Parts sales","subtitle":month,"amount":_copilot_money(parts_rev),"url":url_for('parts_page')}]
        return result

    # Broad safe search fallback.
    like=f"%{q}%"
    vehicles=conn.execute("SELECT id,stock_no,year,make,model,status FROM vehicles WHERE stock_no LIKE ? OR make LIKE ? OR model LIKE ? OR vin LIKE ? OR registration LIKE ? LIMIT 12",(like,like,like,like,like)).fetchall()
    parts=conn.execute("SELECT id,part_number,part_name,make,model,quantity_on_hand,selling_price FROM parts WHERE part_number LIKE ? OR part_name LIKE ? OR make LIKE ? OR model LIKE ? OR manufacturer_part_no LIKE ? LIMIT 12",(like,like,like,like,like)).fetchall()
    result["kind"]="search"
    for r in vehicles:
        result["items"].append({"title":f"{r['stock_no']} — {r['year'] or ''} {r['make']} {r['model']}","subtitle":r['status'],"amount":"","url":url_for('vehicle_detail',vehicle_id=r['id'])})
    for r in parts:
        result["items"].append({"title":f"{r['part_number'] or ''} — {r['part_name']}","subtitle":f"{r['make'] or ''} {r['model'] or ''} · Qty {r['quantity_on_hand'] or 0}","amount":_copilot_money(r['selling_price']),"url":url_for('part_detail',part_id=r['id'])})
    result["summary"]=f"I found {len(result['items'])} BAM record(s) matching “{q}”." if result['items'] else "I couldn't match that request yet. Try asking about parts, stock age, photos, workshop jobs, reminders, invoices or profitable vehicles."
    result["tips"]=["Which vehicles have been sitting longer than 90 days?","Show vehicles missing photos","Show overdue workshop jobs","Which vehicles are most profitable?","Show draft invoices","Find Mini Cooper engine parts"]
    return result


@app.route("/copilot", methods=["GET", "POST"])
@login_required
def dealer_copilot():
    question=(request.form.get("question") if request.method=="POST" else request.args.get("q")) or ""
    conn=db()
    answer=dealer_copilot_answer(conn,question) if question.strip() else None
    conn.close()
    examples=[
        "Which vehicles have been sitting longer than 90 days?",
        "Show vehicles missing photos",
        "Show overdue workshop jobs",
        "Which vehicles are most profitable?",
        "Show draft invoices",
        "Find Mini Cooper engine parts",
    ]
    return render_template("dealer_copilot.html",question=question,answer=answer,examples=examples)


@app.route("/executive-dashboard")
@login_required
def executive_dashboard():
    conn=db(); ensure_smart_reminders(conn)
    today=date.today(); today_text=today.isoformat(); month=today.strftime('%Y-%m'); year=today.strftime('%Y')
    vehicle_today=float(conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) v FROM sales WHERE substr(sale_date,1,10)=?",(today_text,)).fetchone()['v'] or 0)
    parts_today=float(conn.execute("SELECT COALESCE(SUM(sale_price),0) v FROM part_sales WHERE substr(sale_date,1,10)=?",(today_text,)).fetchone()['v'] or 0)
    vehicle_month=float(conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) v FROM sales WHERE substr(sale_date,1,7)=?",(month,)).fetchone()['v'] or 0)
    parts_month=float(conn.execute("SELECT COALESCE(SUM(sale_price),0) v FROM part_sales WHERE substr(sale_date,1,7)=?",(month,)).fetchone()['v'] or 0)
    vehicle_year=float(conn.execute("SELECT COALESCE(SUM(sale_price_inc_gst),0) v FROM sales WHERE substr(sale_date,1,4)=?",(year,)).fetchone()['v'] or 0)
    parts_year=float(conn.execute("SELECT COALESCE(SUM(sale_price),0) v FROM part_sales WHERE substr(sale_date,1,4)=?",(year,)).fetchone()['v'] or 0)
    month_cost=float(conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) v FROM expenses WHERE substr(expense_date,1,7)=?",(month,)).fetchone()['v'] or 0)
    month_cost+=float(conn.execute("SELECT COALESCE(SUM(CASE WHEN actual_cost_inc_gst>0 THEN actual_cost_inc_gst ELSE estimated_cost END),0) v FROM job_cards WHERE substr(job_date,1,7)=?",(month,)).fetchone()['v'] or 0)
    month_cost+=float(conn.execute("SELECT COALESCE(SUM(cost_inc_gst),0) v FROM service_entries WHERE substr(service_date,1,7)=?",(month,)).fetchone()['v'] or 0)
    active=conn.execute("SELECT COUNT(*) c,COALESCE(SUM(purchase_price_inc_gst),0) value FROM vehicles WHERE status NOT IN ('Sold','BER')").fetchone()
    avg_age=float(conn.execute("SELECT COALESCE(AVG(julianday(?) - julianday(COALESCE(purchase_date,substr(created_at,1,10)))),0) v FROM vehicles WHERE status NOT IN ('Sold','BER')",(today_text,)).fetchone()['v'] or 0)
    parts_stock=conn.execute("SELECT COALESCE(SUM(quantity_on_hand),0) qty,COALESCE(SUM(quantity_on_hand*selling_price),0) retail FROM parts").fetchone()
    open_workshop=conn.execute("SELECT COUNT(*) c FROM job_cards WHERE COALESCE(status,'Open')!='Completed'").fetchone()['c']
    open_reminders=conn.execute("SELECT COUNT(*) c FROM reminders WHERE completed=0").fetchone()['c']
    draft_invoices=conn.execute("SELECT COUNT(*) c,COALESCE(SUM(sale_price_inc_gst),0) value FROM sales WHERE COALESCE(invoice_status,'Draft')!='Paid'").fetchone()
    unread_emails=conn.execute("SELECT COUNT(*) c FROM email_messages WHERE status='Unread'").fetchone()['c']
    donor_count=conn.execute("SELECT COUNT(*) c FROM vehicles WHERE COALESCE(vehicle_purpose,'Retail Sale')='Parts Vehicle' OR status='BER'").fetchone()['c']
    top_donors=conn.execute("""
      SELECT v.id,v.stock_no,v.year,v.make,v.model,
       COALESCE((SELECT SUM(ps.sale_price) FROM part_sales ps JOIN parts p ON p.id=ps.part_id WHERE p.vehicle_id=v.id),0)+COALESCE(v.shell_sale_price,0) revenue,
       COALESCE(v.purchase_price_inc_gst,0)+COALESCE((SELECT SUM(e.cost_inc_gst) FROM expenses e WHERE e.vehicle_id=v.id),0) invested
      FROM vehicles v WHERE COALESCE(v.vehicle_purpose,'Retail Sale')='Parts Vehicle' OR v.status='BER'
      ORDER BY (revenue-invested) DESC LIMIT 6
    """).fetchall()
    aged_stock=conn.execute("""
      SELECT id,stock_no,year,make,model,status,CAST(julianday(?) - julianday(COALESCE(purchase_date,substr(created_at,1,10))) AS INTEGER) days
      FROM vehicles WHERE status NOT IN ('Sold','BER') ORDER BY days DESC LIMIT 8
    """,(today_text,)).fetchall()
    monthly=conn.execute("""
      WITH months AS (
        SELECT substr(sale_date,1,7) m FROM sales WHERE sale_date IS NOT NULL
        UNION SELECT substr(sale_date,1,7) FROM part_sales WHERE sale_date IS NOT NULL
      )
      SELECT m,
        COALESCE((SELECT SUM(sale_price_inc_gst) FROM sales WHERE substr(sale_date,1,7)=m),0) vehicle_sales,
        COALESCE((SELECT SUM(sale_price) FROM part_sales WHERE substr(sale_date,1,7)=m),0) parts_sales
      FROM months WHERE m<>'' GROUP BY m ORDER BY m DESC LIMIT 12
    """).fetchall()
    monthly=list(reversed(monthly))
    conn.close()
    kpis={"today_revenue":vehicle_today+parts_today,"month_revenue":vehicle_month+parts_month,"year_revenue":vehicle_year+parts_year,"month_result":vehicle_month+parts_month-month_cost,
          "active_stock":active['c'],"stock_investment":active['value'],"avg_age":avg_age,"parts_qty":parts_stock['qty'],"parts_retail":parts_stock['retail'],"open_workshop":open_workshop,
          "open_reminders":open_reminders,"draft_invoices":draft_invoices['c'],"draft_invoice_value":draft_invoices['value'],"unread_emails":unread_emails,"donor_count":donor_count}
    return render_template("executive_dashboard.html",kpis=kpis,top_donors=top_donors,aged_stock=aged_stock,
                           month_labels=[r['m'] for r in monthly],month_values=[float(r['vehicle_sales'] or 0)+float(r['parts_sales'] or 0) for r in monthly])


# -----------------------------------------------------------------------------
# Version 24.1 - Professional Workshop Scheduler
# -----------------------------------------------------------------------------

def _parse_iso_day(value, fallback=None):
    try:
        return datetime.strptime(value or "", "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return fallback or date.today()


def _workshop_bays():
    raw = os.environ.get("BAM_WORKSHOP_BAYS", "Bay 1,Bay 2,Bay 3")
    return [item.strip() for item in raw.split(",") if item.strip()]


@app.route("/workshop-scheduler", methods=["GET", "POST"])
@login_required
def workshop_scheduler():
    if request.method == "POST":
        conn = db()
        try:
            booking_date = (request.form.get("booking_date") or "").strip()
            description = (request.form.get("description") or "").strip()
            if not booking_date:
                raise ValueError("Booking date is required.")
            _parse_iso_day(booking_date)
            if not description:
                raise ValueError("Enter the workshop work required.")
            vehicle_id = int(request.form.get("vehicle_id")) if request.form.get("vehicle_id") else None
            contact_id = int(request.form.get("contact_id")) if request.form.get("contact_id") else None
            quoted_hours = max(0.0, float(request.form.get("quoted_hours") or 0))
            labour_rate = max(0.0, float(request.form.get("labour_rate") or DEFAULT_LABOUR_RATE))
            cur = conn.execute("""
                INSERT INTO workshop_bookings(
                    vehicle_id,contact_id,booking_date,start_time,end_time,technician,bay,job_type,description,
                    status,quoted_hours,labour_rate,customer_name,customer_phone,notes,created_by,updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,'Booked',?,?,?,?,?,?,?,?)
            """, (
                vehicle_id, contact_id, booking_date, request.form.get("start_time") or None,
                request.form.get("end_time") or None, (request.form.get("technician") or "").strip() or None,
                (request.form.get("bay") or "").strip() or None, (request.form.get("job_type") or "General").strip(),
                description, quoted_hours, labour_rate, (request.form.get("customer_name") or "").strip() or None,
                (request.form.get("customer_phone") or "").strip() or None, (request.form.get("notes") or "").strip() or None,
                session.get("display_name"), datetime.now().isoformat(timespec="seconds")
            ))
            booking_id = cur.lastrowid
            conn.commit()
            log_action("Workshop booking created", "workshop_booking", booking_id, f"{booking_date} — {description}")
            flash("Workshop booking created.", "success")
            return redirect(url_for("workshop_booking_detail", booking_id=booking_id))
        except (ValueError, sqlite3.Error) as exc:
            conn.rollback(); flash(str(exc), "error")
        finally:
            conn.close()

    selected = _parse_iso_day(request.args.get("date"), date.today())
    week_start = selected - timedelta(days=selected.weekday())
    week_days = [week_start + timedelta(days=i) for i in range(7)]
    start_iso, end_iso = week_days[0].isoformat(), week_days[-1].isoformat()
    conn = db()
    rows = conn.execute("""
        SELECT b.*,v.stock_no,v.year,v.make,v.model,v.registration,c.name AS contact_name
        FROM workshop_bookings b
        LEFT JOIN vehicles v ON v.id=b.vehicle_id
        LEFT JOIN contacts c ON c.id=b.contact_id
        WHERE b.booking_date BETWEEN ? AND ?
        ORDER BY b.booking_date,COALESCE(b.start_time,'99:99'),b.id
    """, (start_iso, end_iso)).fetchall()
    vehicles = conn.execute("""
        SELECT id,stock_no,year,make,model,registration FROM vehicles
        WHERE status NOT IN ('Sold') ORDER BY stock_no
    """).fetchall()
    contacts = conn.execute("SELECT id,name,phone FROM contacts ORDER BY name").fetchall()
    users = conn.execute("SELECT display_name FROM users WHERE is_active=1 ORDER BY display_name").fetchall()
    today_iso = date.today().isoformat()
    today_stats = conn.execute("""
        SELECT COUNT(*) AS bookings,
               COALESCE(SUM(quoted_hours),0) AS quoted_hours,
               SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) AS completed,
               SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) AS in_progress
        FROM workshop_bookings WHERE booking_date=?
    """, (today_iso,)).fetchone()
    open_timers = conn.execute("""
        SELECT t.*,b.description,b.technician,b.bay,v.stock_no,v.make,v.model
        FROM workshop_time_entries t
        JOIN workshop_bookings b ON b.id=t.booking_id
        LEFT JOIN vehicles v ON v.id=b.vehicle_id
        WHERE t.clock_out IS NULL ORDER BY t.clock_in
    """).fetchall()
    conn.close()
    by_day = {d.isoformat(): [] for d in week_days}
    for row in rows:
        by_day.setdefault(row["booking_date"], []).append(row)
    prev_week=(week_start-timedelta(days=7)).isoformat(); next_week=(week_start+timedelta(days=7)).isoformat()
    return render_template("workshop_scheduler.html", week_days=week_days, by_day=by_day, selected=selected,
                           prev_week=prev_week, next_week=next_week, vehicles=vehicles, contacts=contacts,
                           technicians=[r["display_name"] for r in users if r["display_name"]], bays=_workshop_bays(),
                           default_labour_rate=DEFAULT_LABOUR_RATE, today_stats=today_stats, open_timers=open_timers,
                           today_iso=today_iso)


@app.route("/workshop-bookings/<int:booking_id>", methods=["GET", "POST"])
@login_required
def workshop_booking_detail(booking_id):
    conn = db()
    booking = conn.execute("""
        SELECT b.*,v.stock_no,v.year,v.make,v.model,v.registration,v.vin,c.name AS contact_name
        FROM workshop_bookings b
        LEFT JOIN vehicles v ON v.id=b.vehicle_id
        LEFT JOIN contacts c ON c.id=b.contact_id
        WHERE b.id=?
    """, (booking_id,)).fetchone()
    if not booking:
        conn.close(); return "Workshop booking not found", 404
    if request.method == "POST":
        try:
            quoted_hours=max(0.0,float(request.form.get("quoted_hours") or 0))
            labour_rate=max(0.0,float(request.form.get("labour_rate") or DEFAULT_LABOUR_RATE))
            status=request.form.get("status") or "Booked"
            allowed={"Booked","Checked In","In Progress","Waiting Parts","Ready","Completed","Cancelled"}
            if status not in allowed: raise ValueError("Invalid workshop status.")
            conn.execute("""
                UPDATE workshop_bookings SET booking_date=?,start_time=?,end_time=?,technician=?,bay=?,job_type=?,description=?,
                    status=?,quoted_hours=?,labour_rate=?,customer_name=?,customer_phone=?,notes=?,updated_at=? WHERE id=?
            """, (request.form.get("booking_date") or booking["booking_date"],request.form.get("start_time") or None,
                  request.form.get("end_time") or None,(request.form.get("technician") or "").strip() or None,
                  (request.form.get("bay") or "").strip() or None,request.form.get("job_type") or "General",
                  (request.form.get("description") or "").strip() or booking["description"],status,quoted_hours,labour_rate,
                  (request.form.get("customer_name") or "").strip() or None,(request.form.get("customer_phone") or "").strip() or None,
                  (request.form.get("notes") or "").strip() or None,datetime.now().isoformat(timespec="seconds"),booking_id))
            conn.commit(); log_action("Workshop booking updated","workshop_booking",booking_id,status)
            flash("Workshop booking updated.","success")
            return redirect(url_for("workshop_booking_detail",booking_id=booking_id))
        except (ValueError,sqlite3.Error) as exc:
            conn.rollback(); flash(str(exc),"error")
    time_entries=conn.execute("SELECT * FROM workshop_time_entries WHERE booking_id=? ORDER BY clock_in DESC,id DESC",(booking_id,)).fetchall()
    total_minutes=sum(int(r["minutes"] or 0) for r in time_entries)
    open_timer=next((r for r in time_entries if not r["clock_out"]),None)
    conn.close()
    return render_template("workshop_booking.html",booking=booking,time_entries=time_entries,total_minutes=total_minutes,
                           open_timer=open_timer,bays=_workshop_bays(),default_labour_rate=DEFAULT_LABOUR_RATE)


@app.post("/workshop-bookings/<int:booking_id>/clock-in")
@login_required
def workshop_booking_clock_in(booking_id):
    conn=db()
    try:
        booking=conn.execute("SELECT * FROM workshop_bookings WHERE id=?",(booking_id,)).fetchone()
        if not booking: return "Workshop booking not found",404
        existing=conn.execute("SELECT id FROM workshop_time_entries WHERE booking_id=? AND clock_out IS NULL",(booking_id,)).fetchone()
        if existing: raise ValueError("This booking already has an active timer.")
        tech=(request.form.get("technician") or booking["technician"] or session.get("display_name") or "Technician").strip()
        now=datetime.now().isoformat(timespec="seconds")
        conn.execute("INSERT INTO workshop_time_entries(booking_id,technician,clock_in,notes,created_by) VALUES(?,?,?,?,?)",
                     (booking_id,tech,now,request.form.get("notes"),session.get("display_name")))
        conn.execute("UPDATE workshop_bookings SET technician=COALESCE(technician,?),status='In Progress',updated_at=? WHERE id=?",(tech,now,booking_id))
        conn.commit(); log_action("Workshop timer started","workshop_booking",booking_id,tech); flash(f"Clocked on: {tech}.","success")
    except (ValueError,sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc),"error")
    finally: conn.close()
    return redirect(request.referrer or url_for("workshop_booking_detail",booking_id=booking_id))


@app.post("/workshop-bookings/<int:booking_id>/clock-out")
@login_required
def workshop_booking_clock_out(booking_id):
    conn=db()
    try:
        entry=conn.execute("SELECT * FROM workshop_time_entries WHERE booking_id=? AND clock_out IS NULL ORDER BY id DESC LIMIT 1",(booking_id,)).fetchone()
        if not entry: raise ValueError("There is no active timer for this booking.")
        out=datetime.now(); started=datetime.fromisoformat(entry["clock_in"])
        minutes=max(0,int(round((out-started).total_seconds()/60)))
        conn.execute("UPDATE workshop_time_entries SET clock_out=?,minutes=? WHERE id=?",(out.isoformat(timespec="seconds"),minutes,entry["id"]))
        if request.form.get("complete") == "1":
            conn.execute("UPDATE workshop_bookings SET status='Completed',updated_at=? WHERE id=?",(out.isoformat(timespec="seconds"),booking_id))
        conn.commit(); log_action("Workshop timer stopped","workshop_booking",booking_id,f"{minutes} minutes"); flash(f"Clocked off — {minutes} minutes recorded.","success")
    except (ValueError,sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc),"error")
    finally: conn.close()
    return redirect(request.referrer or url_for("workshop_booking_detail",booking_id=booking_id))


@app.post("/workshop-bookings/<int:booking_id>/create-job-card")
@login_required
def workshop_booking_create_job(booking_id):
    conn=db()
    try:
        b=conn.execute("SELECT * FROM workshop_bookings WHERE id=?",(booking_id,)).fetchone()
        if not b: return "Workshop booking not found",404
        if not b["vehicle_id"]: raise ValueError("Link a vehicle to this booking before creating a job card.")
        if b["job_card_id"]:
            flash(f"This booking is already linked to job card #{b['job_card_id']}.","error")
            return redirect(url_for("job_card_edit",job_id=b["job_card_id"]))
        hours=float(b["quoted_hours"] or 0); rate=float(b["labour_rate"] or DEFAULT_LABOUR_RATE); estimate=round(hours*rate,2)
        cur=conn.execute("""
            INSERT INTO job_cards(vehicle_id,job_date,category,description,supplier,paid_by,estimated_cost,actual_cost_inc_gst,gst_amount,status,notes,labour_hours,labour_rate,labour_source)
            VALUES(?,?,?,?,?,?,?,?,?,'Open',?,?,?,?)
        """,(b["vehicle_id"],b["booking_date"],b["job_type"] or "Workshop",b["description"],"","Business",estimate,0,0,b["notes"],hours,rate,"Workshop Scheduler"))
        job_id=cur.lastrowid
        conn.execute("INSERT INTO job_card_history(job_card_id,old_status,new_status,changed_by,note) VALUES(?,?,?,?,?)",
                     (job_id,None,"Open",session.get("display_name"),f"Created from workshop booking #{booking_id}"))
        conn.execute("UPDATE workshop_bookings SET job_card_id=?,updated_at=? WHERE id=?",(job_id,datetime.now().isoformat(timespec="seconds"),booking_id))
        conn.commit(); log_action("Workshop booking converted to job card","workshop_booking",booking_id,f"Job #{job_id}")
        flash(f"Job card #{job_id} created.","success"); return redirect(url_for("job_card_edit",job_id=job_id))
    except (ValueError,sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc),"error")
    finally: conn.close()
    return redirect(url_for("workshop_booking_detail",booking_id=booking_id))


@app.post("/workshop-bookings/<int:booking_id>/delete")
@login_required
def workshop_booking_delete(booking_id):
    conn=db()
    try:
        b=conn.execute("SELECT * FROM workshop_bookings WHERE id=?",(booking_id,)).fetchone()
        if not b: return "Workshop booking not found",404
        if b["job_card_id"]: raise ValueError("This booking is linked to a job card. Cancel it instead of deleting it.")
        conn.execute("DELETE FROM workshop_bookings WHERE id=?",(booking_id,)); conn.commit()
        log_action("Workshop booking deleted","workshop_booking",booking_id,b["description"]); flash("Workshop booking deleted.","success")
    except (ValueError,sqlite3.Error) as exc:
        conn.rollback(); flash(str(exc),"error")
    finally: conn.close()
    return redirect(url_for("workshop_scheduler"))

@app.route("/parts/<int:part_id>/label")
@login_required
def part_label(part_id):
    conn = db()
    part = conn.execute("SELECT * FROM parts WHERE id=?", (part_id,)).fetchone()
    if not part:
        conn.close()
        return "Part not found", 404
    donor = conn.execute("SELECT * FROM vehicles WHERE id=?", (part["vehicle_id"],)).fetchone() if part["vehicle_id"] else None
    conn.close()
    return render_template("part_label.html", part=part, donor=donor)


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.template_filter("money")
def money(value):
    try:
        return f"${float(value or 0):,.2f}"
    except Exception:
        return "$0.00"


@app.get("/health")
def health_check():
    return jsonify({"status": "ok", "app": APP_NAME, "version": APP_VERSION}), 200


@app.get("/ready")
def readiness_check():
    try:
        conn = db()
        conn.execute("SELECT 1").fetchone()
        conn.close()
        return jsonify({"status": "ready", "database": str(DB_PATH)}), 200
    except Exception as exc:
        app.logger.exception("Readiness check failed")
        return jsonify({"status": "not_ready", "error": str(exc)}), 503


# Gunicorn imports this module rather than executing it as __main__.
init_db()

if __name__ == "__main__":
    app.run(
        debug=os.environ.get("FLASK_DEBUG", "0") == "1",
        host=os.environ.get("BAM_HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5000")),
    )
